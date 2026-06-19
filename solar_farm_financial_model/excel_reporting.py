"""Workbook reporting helpers for the Solar Farm Streamlit application."""

from __future__ import annotations

from dataclasses import asdict, is_dataclass
from typing import Any, Dict, List, Mapping, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .model import ModelOutputs
from .schemas import Assumptions


def excel_title(ws, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:H1")
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0B3D91")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=11, color="1F2D3D", italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22


def write_styled_table(
    ws,
    df: pd.DataFrame,
    title: str,
    start_row: int,
    start_col: int = 1,
) -> int:
    """Write a styled dataframe section and return the next available row."""

    section_row = start_row
    ws.cell(row=section_row, column=start_col, value=title)
    ws.cell(row=section_row, column=start_col).font = Font(size=13, bold=True, color="0B3D91")
    section_row += 1

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    frame = df.copy()
    frame = frame.replace([np.inf, -np.inf], np.nan)
    frame = frame.where(pd.notna(frame), None)

    for col_idx, column_name in enumerate(frame.columns, start=start_col):
        cell = ws.cell(row=section_row, column=col_idx, value=str(column_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    data_start = section_row + 1
    for row_offset, values in enumerate(frame.itertuples(index=False), start=0):
        excel_row = data_start + row_offset
        for col_idx, value in enumerate(values, start=start_col):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = border
            column_name = str(frame.columns[col_idx - start_col]).lower()
            if isinstance(value, (int, float, np.floating)) and value is not None:
                if "irr" in column_name or "rate" in column_name or "share" in column_name:
                    cell.number_format = "0.0%"
                elif "month" in column_name and "payback" in column_name:
                    cell.number_format = "0"
                else:
                    cell.number_format = "#,##0.00;[Red]-#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if row_offset % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="F7FBFF")

    last_row = data_start + max(len(frame) - 1, 0)
    for col_idx, column_name in enumerate(frame.columns, start=start_col):
        desired = max(14, min(34, len(str(column_name)) + 4))
        for value in frame.iloc[:, col_idx - start_col].head(30):
            desired = max(desired, min(34, len(str(value)) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = desired

    return last_row + 2


def add_workbook_overview_sheet(
    wb: Workbook,
    outputs: ModelOutputs,
    summary_tables: Dict[str, pd.DataFrame],
    assumptions: Assumptions,
    metric_labels: Dict[str, str],
) -> None:
    if "Overview" in wb.sheetnames:
        del wb["Overview"]
    ws = wb.create_sheet("Overview", 0)
    accent = "B45309"
    ws["A1"] = assumptions.global_assumptions.project_name or "Solar Farm Financial Model"
    ws["A1"].font = Font(size=20, bold=True, color="0F172A")
    ws["A2"] = "Executive overview covering key metrics, debt profile, cash flows, and analytics output."
    ws["A2"].font = Font(size=11, color="475569")
    ws["A4"] = "Executive Snapshot"
    ws["A4"].font = Font(size=12, bold=True, color=accent)
    ws["A5"] = "Metric"
    ws["B5"] = "Value"
    for cell in ws[5]:
        cell.fill = PatternFill("solid", fgColor=accent)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    metrics_df = summary_tables.get("metrics", pd.DataFrame()).copy()
    rows: List[Tuple[str, object]] = []
    if not metrics_df.empty:
        for _, metric_row in metrics_df.head(6).iterrows():
            key = metric_row.get("metric")
            label = metric_labels.get(key, str(key).replace("_", " ").title())
            rows.append((label, metric_row.get("value")))
    rows.append(("Projection Start", assumptions.global_assumptions.start_date))
    rows.append(("Forecast Months", assumptions.global_assumptions.forecast_months))
    for row_idx, (label, value) in enumerate(rows[:8], start=6):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
    ws["D4"] = "Workbook Notes"
    ws["D4"].font = Font(size=12, bold=True, color=accent)
    notes = [
        "The detailed dashboard, statements, and analytics tabs remain intact in this export.",
        "Use this overview as the executive cover sheet for sponsor and lender circulation.",
        "Key metrics and project assumptions stay aligned to the live model run.",
    ]
    for row_idx, note in enumerate(notes, start=5):
        ws.cell(row=row_idx, column=4, value=f"- {note}")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["D"].width = 60
    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False


def _labelize(value: str) -> str:
    return str(value).replace("_", " ").strip().title()


def _serialise_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "isoformat") and not isinstance(value, (str, bytes)):
        try:
            return value.isoformat()
        except TypeError:
            pass
    if isinstance(value, (list, tuple)):
        return ", ".join(str(item) for item in value)
    return value


def _reset_index_frame(df: pd.DataFrame, index_label: str) -> pd.DataFrame:
    frame = df.reset_index()
    first_col = str(frame.columns[0])
    if first_col != index_label:
        frame = frame.rename(columns={first_col: index_label})
    return frame


def _records_frame(records: Sequence[Mapping[str, Any]], empty_message: str) -> pd.DataFrame:
    if not records:
        return pd.DataFrame([{"Message": empty_message}])
    return pd.DataFrame(
        [
            {key: _serialise_value(value) for key, value in record.items()}
            for record in records
        ]
    )


def _dataclass_records(items: Sequence[Any], empty_message: str) -> pd.DataFrame:
    records: List[Mapping[str, Any]] = []
    for item in items:
        if is_dataclass(item):
            records.append(asdict(item))
        elif isinstance(item, Mapping):
            records.append(dict(item))
    return _records_frame(records, empty_message)


def _parameter_table(rows: Sequence[Tuple[str, Any]]) -> pd.DataFrame:
    return pd.DataFrame(
        [{"Parameter": label, "Value": _serialise_value(value)} for label, value in rows]
    )


def _annual_schedule(
    monthly: pd.DataFrame,
    prefix: str,
    *,
    rename_func=None,
    include_total: bool = False,
) -> pd.DataFrame:
    matching = [column for column in monthly.columns if column.startswith(prefix)]
    if not matching:
        return pd.DataFrame([{"Message": f"No {prefix.rstrip('_')} schedule available"}])
    schedule = monthly[matching].resample("Y").sum()
    if include_total:
        schedule["total"] = schedule.sum(axis=1)
    schedule.index = schedule.index.year
    schedule = schedule.reset_index().rename(columns={"index": "Year", "month_start": "Year"})
    if rename_func is not None:
        schedule = schedule.rename(columns=rename_func)
    return schedule


def _annual_series(monthly: pd.DataFrame, column: str) -> pd.Series:
    if column not in monthly.columns:
        return pd.Series(dtype=float)
    annual = monthly[column].resample("Y").sum()
    annual.index = annual.index.year
    return annual


def _annual_sum_of_columns(monthly: pd.DataFrame, columns: Sequence[str]) -> pd.Series:
    if not columns:
        return pd.Series(dtype=float)
    annual = monthly[list(columns)].sum(axis=1).resample("Y").sum()
    annual.index = annual.index.year
    return annual


def _build_financial_performance_view(outputs: ModelOutputs) -> pd.DataFrame:
    monthly = outputs.monthly_results
    annual = outputs.annual_summary.copy()

    fixed_opex_cols = [column for column in monthly.columns if column.startswith("opex_fixed_")]
    variable_opex_cols = [column for column in monthly.columns if column.startswith("opex_variable_")]
    admin_tokens = ("admin", "g_and_a", "general", "administration")
    admin_opex_cols = [
        column
        for column in fixed_opex_cols + variable_opex_cols
        if any(token in column for token in admin_tokens)
    ]
    admin_opex_cols = list(dict.fromkeys(admin_opex_cols))

    years = list(annual.index)
    change_in_law = _annual_series(monthly, "opex_change_in_law")
    fixed_opex = _annual_sum_of_columns(monthly, fixed_opex_cols)
    variable_opex = _annual_sum_of_columns(monthly, variable_opex_cols)
    administration = _annual_sum_of_columns(monthly, admin_opex_cols)

    statement = pd.DataFrame(index=years)
    statement.index.name = "Year"
    statement["Revenue"] = annual.get("revenue_total", pd.Series(index=years, dtype=float))
    statement["Fixed Operating Expenses"] = fixed_opex.reindex(years, fill_value=0.0)
    statement["Variable Operating Expenses"] = variable_opex.reindex(years, fill_value=0.0)
    statement["Administration"] = administration.reindex(years, fill_value=0.0)
    statement["Change in Law / Other Pass-through Opex"] = change_in_law.reindex(years, fill_value=0.0)
    statement["Other Operating Expenses"] = (
        annual.get("total_opex", pd.Series(index=years, dtype=float)).fillna(0.0)
        - statement["Administration"]
        - statement["Change in Law / Other Pass-through Opex"]
    )
    statement["Total Operating Expenses"] = annual.get("total_opex", pd.Series(index=years, dtype=float))
    statement["EBITDA"] = annual.get("ebitda", pd.Series(index=years, dtype=float))
    statement["Depreciation"] = annual.get("depreciation", pd.Series(index=years, dtype=float))
    statement["EBIT"] = annual.get("ebit", pd.Series(index=years, dtype=float))
    statement["Interest Expense"] = annual.get("debt_interest", pd.Series(index=years, dtype=float))
    statement["Other Expenses"] = pd.Series(0.0, index=years)
    statement["Profit Before Tax"] = statement["EBIT"] - statement["Interest Expense"] - statement["Other Expenses"]
    statement["Taxes"] = annual.get("tax_payment", pd.Series(index=years, dtype=float))
    statement["Net Income"] = annual.get("net_income", pd.Series(index=years, dtype=float))
    return statement.reset_index().rename(columns={"index": "Year", "month_start": "Year"})


def _build_balance_sheet_view(outputs: ModelOutputs) -> pd.DataFrame:
    monthly = outputs.monthly_results
    opening_ppe = monthly.get("ppe_opening_balance", pd.Series(0.0, index=monthly.index)).cumsum()
    net_ppe = (opening_ppe + monthly["capex"].cumsum() - monthly["depreciation"].cumsum()).clip(lower=0)
    cash_balance = monthly["equity_cash_flow"].cumsum()
    debt_balance = monthly.get("debt_balance", pd.Series(0.0, index=monthly.index))
    accounts_receivable = monthly.get("accounts_receivable", pd.Series(0.0, index=monthly.index))
    prepaid_expenses = monthly.get("prepaid_expenses", pd.Series(0.0, index=monthly.index))
    other_current_assets = monthly.get("other_current_assets", pd.Series(0.0, index=monthly.index))
    inventory_balance = monthly.get("inventory_balance", pd.Series(0.0, index=monthly.index))
    accounts_payable = monthly.get("accounts_payable", pd.Series(0.0, index=monthly.index))
    dsra_required = monthly.get("dsra_required", pd.Series(0.0, index=monthly.index))

    total_assets = (
        cash_balance
        + accounts_receivable
        + prepaid_expenses
        + other_current_assets
        + inventory_balance
        + net_ppe
        + dsra_required
    )
    total_liabilities = debt_balance + accounts_payable

    balance = pd.DataFrame(
        {
            "Cash": cash_balance,
            "Accounts Receivable": accounts_receivable,
            "Prepaid Expenses": prepaid_expenses,
            "Other Current Assets": other_current_assets,
            "Inventory": inventory_balance,
            "DSRA": dsra_required,
            "Net PP&E": net_ppe,
            "Total Assets": total_assets,
            "Accounts Payable": accounts_payable,
            "Debt Outstanding": debt_balance,
            "Total Liabilities": total_liabilities,
        }
    )
    balance["Equity"] = balance["Total Assets"] - balance["Total Liabilities"]
    balance["Total Liabilities & Equity"] = balance["Total Liabilities"] + balance["Equity"]

    annual = balance.resample("Y").last()
    annual.index = annual.index.year
    return annual.reset_index().rename(columns={"index": "Year", "month_start": "Year"})


def _build_cash_flow_view(outputs: ModelOutputs) -> pd.DataFrame:
    monthly = outputs.monthly_results
    ebitda = monthly["ebitda"].resample("Y").sum()
    taxes = monthly["tax_payment"].resample("Y").sum()
    interest = monthly["debt_interest"].resample("Y").sum()
    working_capital = monthly.get("delta_working_capital", pd.Series(0.0, index=monthly.index)).resample("Y").sum()
    operating_cf = ebitda - taxes - interest - working_capital
    investing_cf = (-monthly["capex"]).resample("Y").sum()
    financing_cf = (monthly["debt_draw"] - monthly["debt_principal"]).resample("Y").sum()
    equity_cf = monthly["equity_cash_flow"].resample("Y").sum()

    cash_flow = pd.DataFrame(
        {
            "Operating Cash Flow": operating_cf,
            "Investing Cash Flow": investing_cf,
            "Financing Cash Flow": financing_cf,
            "Equity Cash Flow": equity_cf,
            "Change in Working Capital": -working_capital,
        }
    )
    cash_flow["Net Cash Flow"] = cash_flow[
        ["Operating Cash Flow", "Investing Cash Flow", "Financing Cash Flow"]
    ].sum(axis=1)
    cash_flow["Cumulative Net Cash"] = cash_flow["Net Cash Flow"].cumsum()
    cash_flow.index = cash_flow.index.year
    return cash_flow.reset_index().rename(columns={"index": "Year", "month_start": "Year"})


def _build_global_assumptions_table(assumptions: Assumptions) -> pd.DataFrame:
    global_cfg = assumptions.global_assumptions
    equity_structure = global_cfg.equity_structure
    investor_ownership_share, owner_ownership_share = equity_structure.normalized_ownership()
    return _parameter_table(
        [
            ("Project Name", global_cfg.project_name),
            ("Start Date", global_cfg.start_date),
            ("Forecast Months", global_cfg.forecast_months),
            ("Include Terminal Value", global_cfg.include_terminal_value),
            ("Exit Multiple", global_cfg.exit_multiple),
            ("Discount Rate", global_cfg.discount_rate),
            ("Income Tax Rate", global_cfg.tax.income_tax_rate),
            ("Capital Gains Tax Rate", global_cfg.tax.capital_gains_tax_rate),
            ("Investor Ownership Share", investor_ownership_share),
            ("Owner Ownership Share", owner_ownership_share),
            ("Investor Funding Basis", equity_structure.investor_funding_input_type),
            ("Investor Equity Funding Requirement (%)", equity_structure.investor_funding_share),
            ("Investor Equity Funding Requirement ($)", equity_structure.investor_funding_amount),
            ("Owner Funding Basis", equity_structure.owner_funding_input_type),
            ("Owner Equity Funding Requirement (%)", equity_structure.owner_funding_share),
            ("Owner Equity Funding Requirement ($)", equity_structure.owner_funding_amount),
            ("Terminal Growth Rate", assumptions.terminal_growth_rate),
        ]
    )


def _build_energy_assumptions_table(assumptions: Assumptions) -> pd.DataFrame:
    energy = assumptions.energy
    return _parameter_table(
        [
            ("Capacity MW", energy.capacity_mw),
            ("Capacity Factor", energy.capacity_factor),
            ("Degradation Rate", energy.degradation_rate),
            ("Annual Hours", energy.annual_hours),
            ("Energy Model Mode", energy.energy_model_mode),
            ("Monthly Expected MWh", energy.monthly_expected_mwh or []),
            ("Annual Production Growth Rate", energy.annual_production_growth_rate),
            ("Monthly Minimum MWh", energy.monthly_min_mwh),
            ("Panel Count", energy.panel_count),
            ("Panel Watt DC", energy.panel_watt_dc),
            ("Panel Unit Cost", energy.panel_unit_cost),
            ("DC/AC Ratio", energy.dc_ac_ratio),
            ("Seasonality", list(energy.seasonality)),
        ]
    )


def _build_revenue_assumptions_table(assumptions: Assumptions) -> pd.DataFrame:
    revenue = assumptions.revenue
    return pd.DataFrame(
        [
            {
                "Stream": "PPA",
                "Name": revenue.ppa.name,
                "Share of Output": revenue.ppa.share_of_output,
                "Initial Rate": revenue.ppa.rate_curve.initial,
                "Annual Escalation": revenue.ppa.rate_curve.annual_escalation,
            },
            {
                "Stream": "Merchant",
                "Name": revenue.merchant.name,
                "Share of Output": revenue.merchant.share_of_output,
                "Initial Rate": revenue.merchant.rate_curve.initial,
                "Annual Escalation": revenue.merchant.rate_curve.annual_escalation,
            },
            {
                "Stream": "REC",
                "Name": revenue.rec.name,
                "Share of Output": None,
                "Initial Rate": revenue.rec.initial,
                "Annual Escalation": revenue.rec.annual_escalation,
            },
        ]
    )


def _write_sheet_sections(
    wb: Workbook,
    sheet_name: str,
    title: str,
    subtitle: str,
    sections: Sequence[Tuple[str, pd.DataFrame]],
) -> None:
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    excel_title(ws, title, subtitle)
    row = 4
    for section_title, frame in sections:
        row = write_styled_table(ws, frame, section_title, start_row=row)
        row += 1
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False


def append_comprehensive_workbook_sheets(
    wb: Workbook,
    outputs: ModelOutputs,
    assumptions: Assumptions,
    workbook_inputs: Dict[str, pd.DataFrame] | None = None,
) -> None:
    monthly = outputs.monthly_results
    annual = outputs.annual_summary

    annual_detail = _reset_index_frame(annual, "Year")
    monthly_detail = _reset_index_frame(monthly, "Month")
    monthly_detail["Month"] = monthly_detail["Month"].map(_serialise_value)

    performance_df = _build_financial_performance_view(outputs)

    working_capital_cols = [
        "accounts_receivable",
        "prepaid_expenses",
        "other_current_assets",
        "inventory_balance",
        "accounts_payable",
        "delta_working_capital",
    ]
    working_capital_df = monthly_detail[
        [column for column in ["Month", *working_capital_cols] if column in monthly_detail.columns]
    ]

    capex_cols = [
        "capex",
        "depreciation",
        "bonus_depreciation",
        "ppe_opening_balance",
        "arrangement_fees",
        "itc_benefit",
    ]
    capex_df = monthly_detail[[column for column in ["Month", *capex_cols] if column in monthly_detail.columns]]

    financing_cols = [
        "debt_draw",
        "debt_interest",
        "debt_principal",
        "debt_service",
        "debt_balance",
        "cfads",
        "dscr",
        "dsra_required",
        "dsra_change",
        "cash_after_dsra",
        "major_maintenance_reserve_deposit",
        "inverter_reserve_deposit",
        "cash_sweep",
        "equity_contribution",
        "investor_equity_contribution",
        "owner_equity_contribution",
        "equity_distribution",
        "investor_equity_distribution",
        "owner_equity_distribution",
        "equity_cash_flow",
        "sources_total",
        "uses_total",
        "sources_uses_gap",
    ]
    financing_df = monthly_detail[[column for column in ["Month", *financing_cols] if column in monthly_detail.columns]]

    debt_cols = [
        "debt_draw",
        "debt_interest",
        "debt_principal",
        "debt_service",
        "debt_balance",
    ]
    debt_df = monthly_detail[[column for column in ["Month", *debt_cols] if column in monthly_detail.columns]]

    append_specs: List[Tuple[str, str, str, Sequence[Tuple[str, pd.DataFrame]]]] = [
        (
            "01_Monthly_Financials",
            "Monthly Financials",
            "Full monthly output table from the solar operating model.",
            [("Monthly Financial Outputs", monthly_detail)],
        ),
        (
            "02_Annual_Financials",
            "Annual Financials",
            "Annual roll-up of the solar model outputs.",
            [("Annual Financial Outputs", annual_detail)],
        ),
        (
            "03_Annual_Performance",
            "Annual Performance Statement",
            "Income statement view aligned to project finance reporting.",
            [("Statement of Financial Performance", performance_df)],
        ),
        (
            "04_Annual_Position",
            "Annual Position Statement",
            "Simplified annual balance sheet derived from the solar model.",
            [("Statement of Financial Position", _build_balance_sheet_view(outputs))],
        ),
        (
            "05_Annual_Cash_Flow",
            "Annual Cash Flow Statement",
            "Annual cash flow bridge across operations, investing, and financing.",
            [("Statement of Cash Flows", _build_cash_flow_view(outputs))],
        ),
        (
            "06_Revenue_Schedule",
            "Revenue Schedules",
            "Monthly and annual revenue schedules across all monetisation streams.",
            [
                ("Monthly Revenue Schedule", monthly_detail[[column for column in monthly_detail.columns if column == "Month" or str(column).startswith("revenue_")]]),
                (
                    "Annual Revenue Schedule",
                    _annual_schedule(
                        monthly,
                        "revenue_",
                        rename_func=lambda column: "Year"
                        if column == "Year"
                        else _labelize(str(column).replace("revenue_", "")),
                    ),
                ),
            ],
        ),
        (
            "07_OPEX_Schedule",
            "Operating Cost Schedules",
            "Monthly and annual operating cost schedules for the solar asset.",
            [
                ("Monthly OPEX Schedule", monthly_detail[[column for column in monthly_detail.columns if column == "Month" or str(column).startswith("opex_") or column == "total_opex"]]),
                (
                    "Annual OPEX Schedule",
                    _annual_schedule(
                        monthly,
                        "opex_",
                        rename_func=lambda column: "Year"
                        if column == "Year"
                        else "Total"
                        if str(column).lower() == "total"
                        else _labelize(str(column).replace("opex_", "")),
                        include_total=True,
                    ),
                ),
            ],
        ),
        (
            "08_Working_Capital",
            "Working Capital Schedule",
            "Monthly working capital balances and movements.",
            [("Working Capital Detail", working_capital_df)],
        ),
        (
            "09_CAPEX_Depreciation",
            "CAPEX and Depreciation",
            "Capital deployment, depreciation, and asset schedule outputs.",
            [
                ("Monthly CAPEX & Depreciation", capex_df),
                ("Asset Schedule", outputs.asset_summaries.reset_index(drop=True)),
            ],
        ),
        (
            "10_Financing_Cash",
            "Financing and Cash",
            "Debt service, reserves, and equity distribution schedules.",
            [("Financing & Cash Schedule", financing_df)],
        ),
        (
            "11_Debt_Schedule",
            "Debt Schedule",
            "Debt movement and debt facility assumption schedules.",
            [
                ("Debt Movement Schedule", debt_df),
                (
                    "Debt Facility Assumptions",
                    _dataclass_records(assumptions.debt_facilities, "No debt facilities configured"),
                ),
            ],
        ),
        (
            "12_Asset_Schedule",
            "Asset Schedule",
            "Detailed fixed-asset and depreciation outputs.",
            [("Asset Schedule", outputs.asset_summaries.reset_index(drop=True))],
        ),
        (
            "13_Global_Assumptions",
            "Global Assumptions",
            "Project-level forecast, valuation, and ownership settings.",
            [("Global Assumptions", _build_global_assumptions_table(assumptions))],
        ),
        (
            "14_Energy_Assumptions",
            "Energy Assumptions",
            "Core operating and generation assumptions for the solar farm.",
            [("Energy Assumptions", _build_energy_assumptions_table(assumptions))],
        ),
        (
            "15_Revenue_Assumptns",
            "Revenue Assumptions",
            "Commercial assumptions for PPA, merchant, and REC revenue streams.",
            [("Revenue Assumptions", _build_revenue_assumptions_table(assumptions))],
        ),
        (
            "16_CAPEX_Assumptions",
            "CAPEX Assumptions",
            "Capital input lines, service timing, and depreciation settings.",
            [("CAPEX Assumptions", _dataclass_records(assumptions.capex_items, "No CAPEX items configured"))],
        ),
        (
            "17_Fixed_OPEX_Assump",
            "Fixed OPEX Assumptions",
            "Fixed operating cost assumptions applied in the model.",
            [("Fixed OPEX Assumptions", _dataclass_records(assumptions.fixed_opex, "No fixed OPEX items configured"))],
        ),
        (
            "18_Var_OPEX_Assump",
            "Variable OPEX Assumptions",
            "Variable operating cost assumptions applied in the model.",
            [("Variable OPEX Assumptions", _dataclass_records(assumptions.variable_opex, "No variable OPEX items configured"))],
        ),
        (
            "19_Receivables_WC",
            "Receivables Working Capital",
            "Receivable and prepaid working-capital assumptions by year.",
            [("Receivable Settings", _dataclass_records(assumptions.receivable_settings, "No receivable settings configured"))],
        ),
        (
            "20_Inventory_WC",
            "Inventory Working Capital",
            "Inventory and payable assumptions by year.",
            [("Inventory Settings", _dataclass_records(assumptions.inventory_settings, "No inventory settings configured"))],
        ),
        (
            "21_Tax_Schedule",
            "Tax Schedule",
            "Annual tax override schedule applied to the model.",
            [("Tax Rate Schedule", _dataclass_records(assumptions.tax_schedule, "No explicit tax schedule configured"))],
        ),
        (
            "22_Risk_Schedule",
            "Risk Schedule",
            "Annual risk premia that feed the risk-adjusted discounting.",
            [("Risk Schedule", _dataclass_records(assumptions.risk_schedule, "No explicit risk schedule configured"))],
        ),
    ]

    for sheet_name, title, subtitle, sections in append_specs:
        _write_sheet_sections(wb, sheet_name, title, subtitle, sections)

    support_sheet_map = {
        "labour_schedule": (
            "24_Labour_Schedule",
            "Direct Labour Schedule",
            "Direct labour input table exported from the current workspace.",
            "Direct Labour Structure",
        ),
        "fixed_assets_input": (
            "25_Fixed_Assets_Input",
            "Fixed Assets Input",
            "Fixed-assets input schedule from the current workspace.",
            "Fixed Assets Input",
        ),
        "sensitivity_config": (
            "26_Sensitivity_Config",
            "Sensitivity Configuration",
            "Sensitivity analysis inputs currently configured in the workspace.",
            "Sensitivity Analysis Inputs",
        ),
        "goal_seek_config": (
            "27_Goal_Seek_Config",
            "Goal Seek Configuration",
            "Goal seek targets currently configured in the workspace.",
            "Goal Seek Inputs",
        ),
        "monte_carlo_drivers": (
            "28_MC_Driver_Config",
            "Monte Carlo Drivers",
            "Monte Carlo driver assumptions configured in the workspace.",
            "Monte Carlo Drivers",
        ),
        "monte_carlo_metrics": (
            "29_MC_Metric_Config",
            "Monte Carlo Metrics",
            "Selected Monte Carlo output metrics from the workspace.",
            "Monte Carlo Metrics",
        ),
        "break_even_inputs": (
            "30_Break_Even_Input",
            "Break-even Inputs",
            "Break-even assumptions configured in the workspace.",
            "Break-even Inputs",
        ),
    }

    for key, frame in (workbook_inputs or {}).items():
        if key not in support_sheet_map or frame is None or frame.empty:
            continue
        sheet_name, title, subtitle, section_title = support_sheet_map[key]
        _write_sheet_sections(
            wb,
            sheet_name,
            title,
            subtitle,
            [(section_title, frame.reset_index(drop=True))],
        )
