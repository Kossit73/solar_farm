"""Interactive Streamlit application for the Solar Farm Financial Model."""

from __future__ import annotations

import copy
import io
import html
import math
import os
import re
import tempfile
import uuid
from datetime import date
from pathlib import Path
from typing import Callable, Dict, List, Tuple
from urllib.parse import quote_plus
from urllib.request import Request, urlopen

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference

from solar_farm_financial_model.analytics_helpers import (
    SENSITIVITY_OPTIONS,
    apply_sensitivity_capex as _apply_sensitivity_capex_helper,
    apply_sensitivity_capacity_factor as _apply_sensitivity_capacity_factor_helper,
    apply_sensitivity_discount_rate as _apply_sensitivity_discount_rate_helper,
    apply_sensitivity_fixed_opex as _apply_sensitivity_fixed_opex_helper,
    apply_sensitivity_merchant_rate as _apply_sensitivity_merchant_rate_helper,
    apply_sensitivity_ppa_rate as _apply_sensitivity_ppa_rate_helper,
    apply_sensitivity_rec_rate as _apply_sensitivity_rec_rate_helper,
    apply_sensitivity_variable_opex as _apply_sensitivity_variable_opex_helper,
    simulate_metrics as _simulate_metrics_helper,
    simulate_outputs as _simulate_outputs_helper,
)
from solar_farm_financial_model.data_loader import load_assumptions
from solar_farm_financial_model.excel_reporting import (
    add_workbook_overview_sheet as _add_workbook_overview_sheet_helper,
    excel_title as _excel_title_helper,
    write_styled_table as _write_styled_table_helper,
)
from solar_farm_financial_model.input_parsing import (
    CALENDAR_HOURS_PER_YEAR,
    apply_energy_input_mode as _apply_energy_input_mode_helper,
    build_opex_items as _build_opex_items_helper,
    capex_item_from_initial as _capex_item_from_initial_helper,
    capex_item_from_row as _capex_item_from_row_helper,
    coerce_float as _coerce_float_helper,
    coerce_optional_float as _coerce_optional_float_helper,
    coerce_optional_int as _coerce_optional_int_helper,
    parse_spend_profile as _parse_spend_profile_helper,
    rows_from_tuple as _rows_from_tuple_helper,
    tupleize as _tupleize_helper,
)
from solar_farm_financial_model.presentation import (
    inject_app_theme as _inject_app_theme_helper,
    render_model_hero as _render_model_hero_helper,
)
from solar_farm_financial_model.ai import ConversationMemory, run_assistant_turn
from solar_farm_financial_model.ai.providers import (
    PROVIDER_SPECS,
    env_api_key,
    provider_capabilities,
    provider_default_base_url,
    provider_label,
    provider_models,
)
from solar_farm_financial_model.model import (
    ModelOutputs,
    SolarFarmFinancialModel,
    capex_item_schedule,
)
from solar_farm_financial_model.reporting import build_summary_report
from solar_farm_financial_model.schemas import (
    Assumptions,
    CapexItem,
    DebtFacility,
    FixedOpexItem,
    InventoryPayableSettings,
    ReceivableSettings,
    TaxRateSchedule,
    VariableOpexItem,
)


MetricLabels = {
    "project_npv": "Project NPV",
    "project_irr": "Project IRR",
    "equity_irr": "Equity IRR",
    "investor_irr": "Investor IRR",
    "owner_irr": "Owner IRR",
    "project_payback_months": "Payback (months)",
}

ENERGY_INPUT_MODE_OPTIONS: Dict[str, str] = {
    "capacity_factor": "Capacity factor (8760 hours)",
    "resource_hours": "Annual resource hours",
    "monthly_expected_mwh": "Monthly expected MWh",
}


def _render_llm_settings() -> Dict[str, object]:
    """Render provider-agnostic LLM settings and return active config."""
    with st.expander("LLM Settings", expanded=False):
        provider_options = list(PROVIDER_SPECS.keys())
        current_provider = str(st.session_state.get("llm_provider_name", "openai"))
        if current_provider not in provider_options:
            current_provider = "openai"

        provider_name = st.selectbox(
            "Provider",
            options=provider_options,
            index=provider_options.index(current_provider),
            format_func=provider_label,
            key="llm_provider_name",
        )

        provider_state = st.session_state.setdefault("llm_provider_configs", {})
        provider_config = dict(
            provider_state.get(
                provider_name,
                {
                    "model_name": provider_models(provider_name)[0],
                    "api_key": env_api_key(provider_name),
                    "base_url": provider_default_base_url(provider_name),
                    "temperature": 0.2,
                    "max_tokens": 1200,
                    "reasoning_mode": "medium",
                    "enable_tools": True,
                    "enable_web_search": True,
                },
            )
        )

        models = provider_models(provider_name)
        current_model = str(provider_config.get("model_name", models[0]))
        if current_model not in models:
            current_model = models[0]
        model_name = st.selectbox(
            "Model",
            options=models,
            index=models.index(current_model),
            key=f"llm_model_name_{provider_name}",
        )

        api_key_value = st.text_input(
            f"{provider_label(provider_name)} API Key",
            value=str(provider_config.get("api_key", "")),
            type="password",
            placeholder="Enter API key",
            help="Entered key is stored in current session environment only.",
            key=f"llm_api_key_{provider_name}",
        )
        base_url = st.text_input(
            "Base URL (optional)",
            value=str(provider_config.get("base_url", provider_default_base_url(provider_name))),
            placeholder="https://...",
            key=f"llm_base_url_{provider_name}",
        )
        temperature = st.slider(
            "Temperature",
            min_value=0.0,
            max_value=1.5,
            value=float(provider_config.get("temperature", 0.2)),
            step=0.1,
            key=f"llm_temperature_{provider_name}",
        )
        max_tokens = st.number_input(
            "Max Tokens",
            min_value=128,
            max_value=8192,
            value=int(provider_config.get("max_tokens", 1200)),
            step=128,
            key=f"llm_max_tokens_{provider_name}",
        )
        reasoning_mode = st.selectbox(
            "Reasoning Mode",
            options=["low", "medium", "high"],
            index=(
                ["low", "medium", "high"].index(str(provider_config.get("reasoning_mode", "medium")))
                if str(provider_config.get("reasoning_mode", "medium")) in {"low", "medium", "high"}
                else 1
            ),
            key=f"llm_reasoning_mode_{provider_name}",
        )
        enable_tools = st.checkbox(
            "Enable Tool Calls",
            value=bool(provider_config.get("enable_tools", True)),
            key=f"llm_enable_tools_{provider_name}",
        )
        enable_web_search = st.checkbox(
            "Enable Web Search (if supported)",
            value=bool(provider_config.get("enable_web_search", True)),
            key=f"llm_enable_web_search_{provider_name}",
        )

        caps = provider_capabilities(provider_name)
        st.caption(
            "Capabilities: "
            f"reasoning={'✅' if caps.reasoning else '—'}, "
            f"long_context={'✅' if caps.long_context else '—'}, "
            f"tool_use={'✅' if caps.tool_use else '—'}, "
            f"web_search={'✅' if caps.web_search else '—'}, "
            f"vision={'✅' if caps.vision else '—'}, "
            f"streaming={'✅' if caps.streaming else '—'}"
        )

    provider_state[provider_name] = {
        "model_name": model_name,
        "api_key": api_key_value.strip(),
        "base_url": base_url.strip(),
        "temperature": float(temperature),
        "max_tokens": int(max_tokens),
        "reasoning_mode": reasoning_mode,
        "enable_tools": bool(enable_tools),
        "enable_web_search": bool(enable_web_search),
    }
    st.session_state["llm_provider_configs"] = provider_state

    return {
        "provider_name": provider_name,
        **provider_state[provider_name],
    }


def _configure_llm_secrets() -> None:
    """Load provider API keys from Streamlit secrets when available."""
    secret_keys = [
        "OPENAI_API_KEY",
        "ANTHROPIC_API_KEY",
        "GOOGLE_API_KEY",
        "MISTRAL_API_KEY",
        "COHERE_API_KEY",
        "XAI_API_KEY",
        "DEEPSEEK_API_KEY",
        "LLAMA_API_KEY",
    ]
    for key in secret_keys:
        if os.environ.get(key):
            continue
        secret_value = st.secrets.get(key, "")
        if secret_value:
            os.environ[key] = str(secret_value)


@st.cache_data(show_spinner=False)
def _run_model(
    excel_bytes: bytes | None,
    override_items: Tuple[Tuple[str, float | bool], ...],
    monthly_generation_rows: Tuple[Tuple[str, float], ...],
    labour_rows: Tuple[Tuple[object, ...], ...],
    initial_investment_rows: Tuple[Tuple[object, ...], ...],
    cost_rows: Tuple[Tuple[object, ...], ...],
    receivable_rows: Tuple[Tuple[object, ...], ...],
    inventory_rows: Tuple[Tuple[object, ...], ...],
    fixed_asset_rows: Tuple[Tuple[object, ...], ...],
    loan_rows: Tuple[Tuple[object, ...], ...],
    tax_rows: Tuple[Tuple[object, ...], ...],
    inflation_rows: Tuple[Tuple[object, ...], ...],
    risk_rows: Tuple[Tuple[object, ...], ...],
) -> Tuple[ModelOutputs, Dict[str, pd.DataFrame], Assumptions]:
    """Execute the financial model with optional overrides and return outputs."""

    overrides = dict(override_items)
    monthly_generation_list = _rows_from_tuple(monthly_generation_rows, ("month", "expected_mwh"))
    labour_list = _rows_from_tuple(labour_rows, ("role", "annual_cost"))
    initial_investment_list = _rows_from_tuple(
        initial_investment_rows,
        (
            "id",
            "name",
            "amount",
            "depreciation_years",
            "method",
            "year",
            "month",
            "spend_profile",
            "opening_balance",
            "depreciation_rate",
            "service_month",
        ),
    )
    cost_list = _rows_from_tuple(cost_rows, ("name", "fixed_cost", "variable_cost", "inflation_rate"))
    receivable_list = _rows_from_tuple(
        receivable_rows,
        ("year", "days_in_year", "receivable_days", "prepaid_expense_days", "other_asset_days"),
    )
    inventory_list = _rows_from_tuple(
        inventory_rows,
        ("year", "days_in_year", "inventory_days", "accounts_payable_days"),
    )
    fixed_asset_list = _rows_from_tuple(
        fixed_asset_rows,
        (
            "asset_type",
            "method",
            "year",
            "acquisition",
            "asset_life",
            "net_book_value",
            "depreciation_rate",
            "total_asset_cost",
            "total_depreciation",
            "cumulative_depreciation",
            "ending_book_value",
        ),
    )
    loan_list = _rows_from_tuple(loan_rows, ("name", "year", "duration_years", "amount", "interest_rate"))
    tax_list = _rows_from_tuple(tax_rows, ("name", "year", "tax_rate"))
    inflation_list = _rows_from_tuple(inflation_rows, ("name", "year", "rate"))
    risk_list = _rows_from_tuple(
        risk_rows,
        ("name", "year", "inherent_risk", "climate_risk", "political_risk"),
    )
    temp_path: Path | None = None

    if excel_bytes:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(excel_bytes)
            temp_path = Path(tmp.name)

    try:
        assumptions = load_assumptions(temp_path)
    finally:
        if temp_path and temp_path.exists():
            temp_path.unlink()

    # Apply overrides captured from the UI
    project_name_override = str(overrides.get("project_name", "")).strip()
    if project_name_override:
        assumptions.global_assumptions.project_name = project_name_override

    assumptions.global_assumptions.discount_rate = float(overrides["discount_rate"])
    assumptions.global_assumptions.exit_multiple = float(overrides["exit_multiple"])
    assumptions.global_assumptions.include_terminal_value = bool(overrides["include_terminal"])
    assumptions.terminal_growth_rate = float(overrides["terminal_growth_rate"])

    energy = assumptions.energy
    energy.capacity_mw = float(overrides["capacity_mw"])
    energy.capacity_factor = float(overrides["capacity_factor"])
    energy.degradation_rate = float(overrides["degradation_rate"])
    energy.annual_hours = int(CALENDAR_HOURS_PER_YEAR)
    energy.panel_count = max(0.0, float(overrides["panel_count"]))
    energy.panel_watt_dc = max(0.0, float(overrides["panel_watt_dc"]))
    energy.panel_unit_cost = max(0.0, float(overrides["panel_unit_cost"]))
    energy.dc_ac_ratio = max(0.1, float(overrides["dc_ac_ratio"]))
    _apply_energy_input_mode_helper(
        energy,
        str(overrides.get("energy_input_mode", "capacity_factor")),
        capacity_factor=float(overrides["capacity_factor"]),
        annual_resource_hours=float(overrides.get("annual_resource_hours", 0.0)),
        monthly_generation_rows=monthly_generation_list,
    )
    setattr(assumptions, "energy_input_mode", str(overrides.get("energy_input_mode", "capacity_factor")))

    energy.annual_production_growth_rate = float(overrides["annual_production_growth_rate"])
    energy.monthly_min_mwh = float(overrides["monthly_min_mwh"])

    revenue = assumptions.revenue
    ppa_share = float(overrides["ppa_share"])
    merchant_share = max(0.0, min(1.0, 1.0 - ppa_share))
    revenue.ppa.share_of_output = ppa_share
    revenue.merchant.share_of_output = merchant_share

    revenue.ppa.rate_curve.initial = float(overrides["ppa_rate"])
    revenue.ppa.rate_curve.annual_escalation = float(overrides["ppa_escalation"])

    revenue.merchant.rate_curve.initial = float(overrides["merchant_rate"])
    revenue.merchant.rate_curve.annual_escalation = float(overrides["merchant_escalation"])

    revenue.rec.initial = float(overrides["rec_rate"])
    revenue.rec.annual_escalation = float(overrides["rec_escalation"])

    start_year = int(overrides.get("start_year", assumptions.global_assumptions.start_date.year))
    end_year = int(overrides.get("end_year", start_year + assumptions.global_assumptions.forecast_months // 12 - 1))
    if end_year < start_year:
        end_year = start_year

    assumptions.global_assumptions.start_date = date(start_year, 1, 1)
    assumptions.global_assumptions.forecast_months = max(12, (end_year - start_year + 1) * 12)

    assumptions.global_assumptions.tax.income_tax_rate = float(overrides["income_tax_rate"])
    assumptions.global_assumptions.tax.capital_gains_tax_rate = float(overrides["capital_gains_tax_rate"])
    assumptions.global_assumptions.distribution.investor_share = float(overrides["investor_share"])
    assumptions.global_assumptions.distribution.owner_share = float(overrides["owner_share"])

    inflation_rates = [
        _coerce_float(row.get("rate"))
        for row in inflation_list
        if row.get("rate") is not None
    ]
    inflation_default = float(np.mean(inflation_rates)) if inflation_rates else 0.02

    if cost_list:
        fixed_items, variable_items = _build_opex_items(cost_list, inflation_default)
    else:
        fixed_items = list(assumptions.fixed_opex)
        variable_items = list(assumptions.variable_opex)

    for row in labour_list:
        role = str(row.get("role", "")).strip()
        cost = _coerce_float(row.get("annual_cost"))
        if role and cost > 0:
            fixed_items.append(FixedOpexItem(name=role, annual_cost=cost, inflation_rate=inflation_default))

    assumptions.fixed_opex = tuple(fixed_items)
    assumptions.variable_opex = tuple(variable_items)

    receivable_settings = [
        ReceivableSettings(
            year=int(row.get("year", start_year)),
            days_in_year=int(row.get("days_in_year", 365)),
            receivable_days=_coerce_float(row.get("receivable_days")),
            prepaid_expense_days=_coerce_float(row.get("prepaid_expense_days")),
            other_asset_days=_coerce_float(row.get("other_asset_days")),
        )
        for row in receivable_list
    ]
    if receivable_settings:
        assumptions.receivable_settings = receivable_settings

    inventory_settings = [
        InventoryPayableSettings(
            year=int(row.get("year", start_year)),
            days_in_year=int(row.get("days_in_year", 365)),
            inventory_days=_coerce_float(row.get("inventory_days")),
            accounts_payable_days=_coerce_float(row.get("accounts_payable_days")),
        )
        for row in inventory_list
    ]
    if inventory_settings:
        assumptions.inventory_settings = inventory_settings

    tax_schedule = [
        TaxRateSchedule(year=int(row.get("year", start_year)), tax_rate=_coerce_float(row.get("tax_rate")))
        for row in tax_list
        if row.get("tax_rate") is not None
    ]
    if tax_schedule:
        assumptions.tax_schedule = tax_schedule

    capex_items = []
    for row in initial_investment_list:
        item = _capex_item_from_initial(row, start_year)
        if item is not None:
            capex_items.append(item)

    if not capex_items:
        for row in fixed_asset_list:
            item = _capex_item_from_row(row, start_year)
            if item:
                capex_items.append(item)

    if capex_items:
        capex_items.sort(key=lambda item: (item.service_month, item.name.lower()))
        assumptions.capex_items = tuple(capex_items)

    debt_facilities: List[DebtFacility] = []
    for row in loan_list:
        amount = _coerce_float(row.get("amount"))
        if amount <= 0:
            continue
        facility_year = int(row.get("year", start_year))
        duration_years = max(1, int(round(_coerce_float(row.get("duration_years"), 1.0))))
        interest_rate = max(0.0, _coerce_float(row.get("interest_rate")))
        start_month = max(1, (facility_year - start_year) * 12 + 1)
        debt_facilities.append(
            DebtFacility(
                name=str(row.get("name", "Facility")) or "Facility",
                principal=amount,
                interest_rate=interest_rate,
                term_months=duration_years * 12,
                interest_only_months=0,
                start_month=start_month,
            )
        )
    if debt_facilities:
        debt_facilities.sort(key=lambda facility: (facility.start_month, facility.name.lower()))
        assumptions.debt_facilities = tuple(debt_facilities)

    risk_totals = [
        _coerce_float(row.get("inherent_risk"))
        + _coerce_float(row.get("climate_risk"))
        + _coerce_float(row.get("political_risk"))
        for row in risk_list
    ]
    risk_premium = float(np.mean(risk_totals)) if risk_totals else 0.0
    assumptions.global_assumptions.discount_rate = min(
        0.99, float(overrides["discount_rate"]) + risk_premium
    )

    model = SolarFarmFinancialModel(assumptions)
    outputs = model.run()
    summary_tables = build_summary_report(outputs)
    return outputs, summary_tables, assumptions


def _format_currency(value: float) -> str:
    return f"$ {value:,.0f}"


def _format_percentage(value: float) -> str:
    return f"{value * 100:.1f}%"


def _format_metric(name: str, value: float) -> str:
    if name.endswith("irr") and pd.notna(value):
        return _format_percentage(value)
    if "payback" in name and pd.notna(value):
        return f"{value:.0f} months"
    if pd.notna(value):
        return _format_currency(value)
    return "N/A"


def _downloadable_csv(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=True).encode("utf-8")


QUESTION_TYPE_QUERIES: Dict[str, str] = {
    "valuation": "utility scale solar project valuation EV EBITDA IRR benchmark",
    "profitability": "utility scale solar project EBITDA margin benchmark",
    "liquidity": "project finance solar DSCR minimum benchmark",
    "leverage": "project finance debt service coverage ratio covenant utility scale solar",
    "growth": "utility scale solar revenue growth benchmark power price escalation",
    "pricing": "US utility scale solar PPA price benchmark",
    "efficiency": "utility scale solar capacity factor benchmark United States",
    "risk": "utility scale solar project risk benchmark debt sizing",
}


BENCHMARK_REFERENCE_RANGES: Dict[str, Dict[str, str]] = {
    "valuation": {"equity_irr": "~8%–12% (contracted) / 12%–18% (merchant-heavy)"},
    "profitability": {"ebitda_margin": "~55%–85% for mature generation assets"},
    "liquidity": {"dscr": ">=1.20x base case, >=1.00x downside"},
    "leverage": {"debt_to_ebitda": "~4.0x–7.0x project-finance range"},
    "growth": {"annual_escalation": "~1%–3% contracted escalation"},
    "pricing": {"ppa_price_usd_per_mwh": "region-dependent; commonly benchmarked vs local ISO/utility data"},
    "efficiency": {"capacity_factor": "~18%–30% utility-scale PV (resource-dependent)"},
    "risk": {"discount_rate": "~6%–10% contracted, higher for merchant exposure"},
}


def _classify_question_type(question: str) -> str:
    text = question.lower()
    keyword_map = {
        "valuation": ["valuation", "multiple", "ev", "irr", "npv", "moic"],
        "profitability": ["profit", "margin", "ebitda", "net income"],
        "liquidity": ["liquidity", "runway", "cash", "working capital"],
        "leverage": ["debt", "leverage", "dscr", "covenant", "coverage"],
        "growth": ["growth", "expansion", "increase", "trend"],
        "pricing": ["price", "ppa", "tariff", "merchant"],
        "efficiency": ["capacity factor", "efficiency", "utilization", "output"],
        "risk": ["risk", "stress", "downside", "volatility", "sensitivity"],
    }
    for question_type, words in keyword_map.items():
        if any(word in text for word in words):
            return question_type
    return "valuation"


@st.cache_data(show_spinner=False, ttl=86400)
def _web_search_benchmarks(query: str, max_results: int = 4) -> List[Dict[str, str]]:
    """Simple web search via DuckDuckGo HTML endpoint."""

    search_url = f"https://duckduckgo.com/html/?q={quote_plus(query)}"
    req = Request(search_url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(req, timeout=12) as response:
        html_text = response.read().decode("utf-8", errors="ignore")

    pattern = re.compile(r'class="result__a" href="(.*?)".*?>(.*?)</a>', re.DOTALL)
    results: List[Dict[str, str]] = []
    for match in pattern.finditer(html_text):
        if len(results) >= max_results:
            break
        url = html.unescape(match.group(1))
        title = re.sub(r"<.*?>", "", html.unescape(match.group(2))).strip()
        if url and title:
            results.append({"title": title, "url": url})
    return results


def _safe_ratio(numerator: float, denominator: float) -> float:
    if denominator is None or not np.isfinite(denominator) or abs(denominator) < 1e-12:
        return float("nan")
    return numerator / denominator


def _internal_model_summary(outputs: ModelOutputs, assumptions: Assumptions) -> Dict[str, float]:
    monthly = outputs.monthly_results
    annual = outputs.annual_summary
    latest_year = annual.iloc[-1] if not annual.empty else pd.Series(dtype=float)
    first_year = annual.iloc[0] if not annual.empty else pd.Series(dtype=float)

    revenue = float(latest_year.get("revenue_total", float("nan")))
    ebitda = float(latest_year.get("ebitda", float("nan")))
    net_income = float(latest_year.get("net_income", float("nan")))
    debt_balance = float(monthly.get("debt_balance", pd.Series([0.0])).iloc[-1])
    annual_debt_service = float(latest_year.get("debt_interest", 0.0) + latest_year.get("debt_principal", 0.0))
    if "dscr" in monthly.columns:
        dscr_series = monthly["dscr"].replace([np.inf, -np.inf], np.nan).dropna()
        dscr = float(dscr_series.mean()) if not dscr_series.empty else float("nan")
    else:
        dscr = _safe_ratio(float(latest_year.get("cfads", float("nan"))), annual_debt_service) if annual_debt_service else float("nan")

    annual_energy = float(monthly["energy_mwh"].sum()) / max(1.0, assumptions.global_assumptions.forecast_months / 12.0)
    total_capex = float(sum(item.amount for item in assumptions.capex_items))
    total_fixed_opex = float(sum(getattr(item, "annual_cost", 0.0) for item in assumptions.fixed_opex))
    weighted_variable_opex = float(sum(item.cost_per_mwh for item in assumptions.variable_opex))
    lcoe_proxy = _safe_ratio(total_fixed_opex + annual_energy * weighted_variable_opex, annual_energy) if annual_energy else float("nan")

    return {
        "project_npv": float(outputs.metrics.get("project_npv", float("nan"))),
        "project_irr": float(outputs.metrics.get("project_irr", float("nan"))),
        "equity_irr": float(outputs.metrics.get("equity_irr", float("nan"))),
        "investor_irr": float(outputs.metrics.get("investor_irr", float("nan"))),
        "revenue": revenue,
        "revenue_year_1": float(first_year.get("revenue_total", float("nan"))),
        "ebitda": ebitda,
        "net_income": net_income,
        "ebitda_margin": _safe_ratio(ebitda, revenue) if revenue else float("nan"),
        "debt_balance": debt_balance,
        "debt_to_ebitda": _safe_ratio(debt_balance, ebitda) if ebitda else float("nan"),
        "dscr_proxy": dscr,
        "capex_total": total_capex,
        "capex_per_mw": _safe_ratio(total_capex, assumptions.energy.capacity_mw),
        "opex_per_mw": _safe_ratio(total_fixed_opex, assumptions.energy.capacity_mw),
        "opex_per_mwh": weighted_variable_opex,
        "lcoe_proxy": lcoe_proxy,
        "ppa_rate": float(assumptions.revenue.ppa.rate_curve.initial),
        "merchant_rate": float(assumptions.revenue.merchant.rate_curve.initial),
        "capacity_factor": float(assumptions.energy.capacity_factor),
        "capacity_mw": float(assumptions.energy.capacity_mw),
        "discount_rate": float(assumptions.global_assumptions.discount_rate),
        "exit_multiple": float(assumptions.global_assumptions.exit_multiple),
    }


def _build_structured_ai_response(
    question: str,
    model_data: Dict[str, float],
    benchmark_range: Dict[str, str],
    sources: List[Dict[str, str]],
    question_type: str,
    prior_messages: List[Dict[str, str]],
) -> str:
    """Generate a structured, auditable response grounded in model + benchmark context."""

    status = "conservative"
    if pd.notna(model_data.get("project_npv")) and model_data.get("project_npv", 0.0) < 0:
        status = "high-risk"
    if pd.notna(model_data.get("equity_irr")) and model_data.get("equity_irr", 0.0) > 0.15:
        status = "aggressive upside"

    memory_note = (
        f"Using {len(prior_messages)} prior exchanges to maintain continuity."
        if prior_messages
        else "No prior chat context yet in this session."
    )

    sources_md = "\n".join([f"- {s['title']}: {s['url']}" for s in sources]) or "- No live benchmark sources retrieved."
    benchmark_md = "\n".join([f"- {k.replace('_', ' ').title()}: {v}" for k, v in benchmark_range.items()]) or "- No benchmark range mapping for this question type."

    return (
        "### 1) Direct answer\n"
        f"This is a **{question_type}** question; the current model looks **{status}** after combining internal outputs with benchmark context.\n\n"
        "### 2) Internal model insight\n"
        f"- Project NPV: `{model_data['project_npv']:,.0f}`\n"
        f"- Project IRR / Equity IRR / Investor IRR: `{model_data['project_irr']:.2%}` / `{model_data['equity_irr']:.2%}` / `{model_data['investor_irr']:.2%}`\n"
        f"- Revenue (Y1 → latest): `{model_data['revenue_year_1']:,.0f}` → `{model_data['revenue']:,.0f}`\n"
        f"- EBITDA margin: `{model_data['ebitda_margin']:.2%}`\n"
        f"- Debt/EBITDA (proxy): `{model_data['debt_to_ebitda']:.2f}x` and DSCR proxy `{model_data['dscr_proxy']:.2f}x`\n"
        f"- CAPEX total / CAPEX per MW: `{model_data['capex_total']:,.0f}` / `{model_data['capex_per_mw']:,.0f}`\n"
        f"- OPEX per MW / OPEX per MWh: `{model_data['opex_per_mw']:,.0f}` / `{model_data['opex_per_mwh']:,.2f}`\n"
        f"- LCOE proxy: `{model_data['lcoe_proxy']:,.2f}` per MWh\n"
        f"- Capacity factor, PPA rate, merchant rate: `{model_data['capacity_factor']:.2%}`, `{model_data['ppa_rate']:,.2f}`, `{model_data['merchant_rate']:,.2f}`\n\n"
        "### 3) External benchmark comparison\n"
        f"{benchmark_md}\n\n"
        "### 4) Interpretation\n"
        f"- Relative status assessment: **{status}**.\n"
        "- Guardrails: distinguish model facts above from benchmark assumptions below.\n"
        f"- Context continuity: {memory_note}\n"
        "- If benchmark evidence is weak, treat conclusions as directional and validate with transaction-specific comps.\n\n"
        "### 5) Recommendation\n"
        "- Stress-test the highest-sensitivity assumptions for this topic (pricing, capacity factor, opex, debt terms).\n"
        "- Validate CAPEX/OPEX and return assumptions against one contracted case and one merchant-downside case.\n"
        "- If DSCR proxy or return metrics fall outside target ranges, rebalance leverage and commercial structure before IC.\n\n"
        "### 6) Sources\n"
        "- Internal model: current run outputs and assumptions from this session.\n"
        "- External benchmark references:\n"
        f"{sources_md}"
    )


def _render_ai_benchmark_assistant(outputs: ModelOutputs, assumptions: Assumptions) -> None:
    st.header("AI Reasoning Chatbot")
    st.caption(
        "Model-grounded and benchmark-aware copilot for valuation, profitability, leverage, pricing, efficiency, and risk."
    )
    llm_config = _render_llm_settings()
    if llm_config.get("api_key"):
        provider = provider_label(str(llm_config.get("provider_name", "openai")))
        model = str(llm_config.get("model_name", ""))
        st.success(f"{provider} configured (`{model}`).")
    else:
        st.warning(
            "No provider API key detected. Chatbot is running in local fallback mode. "
            "Set a provider key in LLM Settings to enable hosted reasoning."
        )

    if "ai_chat_history" not in st.session_state:
        st.session_state["ai_chat_history"] = []
    if "ai_memory" not in st.session_state or not isinstance(st.session_state["ai_memory"], ConversationMemory):
        st.session_state["ai_memory"] = ConversationMemory()

    for item in st.session_state["ai_chat_history"]:
        with st.chat_message("user"):
            st.write(item["question"])
        with st.chat_message("assistant"):
            st.markdown(item["answer"])

    prompt = st.chat_input("Ask about IRR, DSCR, CAPEX/MW, OPEX/MWh, valuation realism, feasibility, etc.")
    if not prompt:
        return

    turn, memory = run_assistant_turn(
        question=prompt,
        outputs=outputs,
        assumptions=assumptions,
        memory=st.session_state["ai_memory"],
        llm_config=llm_config,
    )
    st.session_state["ai_memory"] = memory
    st.session_state["ai_chat_history"].append({"question": prompt, "answer": turn.answer_markdown})

    with st.chat_message("user"):
        st.write(prompt)
    with st.chat_message("assistant"):
        st.markdown(turn.answer_markdown)


def _excel_title(ws, title: str, subtitle: str) -> None:
    _excel_title_helper(ws, title, subtitle)


def _write_styled_table(
    ws,
    df: pd.DataFrame,
    title: str,
    start_row: int,
    start_col: int = 1,
) -> int:
    return _write_styled_table_helper(ws, df, title, start_row, start_col)


@st.cache_data(show_spinner=False)
def _downloadable_excel(
    outputs: ModelOutputs,
    summary_tables: Dict[str, pd.DataFrame],
    assumptions: Assumptions,
) -> bytes:
    """Create an investor-ready workbook aligned to Key Metrics, Financials, and Key Analytics tabs."""

    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Key Metrics Dashboard
    ws_metrics = wb.active
    ws_metrics.title = "Key Metrics Dashboard"
    _excel_title(
        ws_metrics,
        f"{assumptions.global_assumptions.project_name} | Key Metrics Dashboard",
        "KPIs, drivers, and trend visuals",
    )

    metrics_df = summary_tables["metrics"].copy()
    metrics_df["metric"] = metrics_df["metric"].map(
        lambda m: MetricLabels.get(m, str(m).replace("_", " ").title())
    )
    metrics_df = metrics_df.rename(columns={"metric": "Metric", "value": "Value"})
    row = _write_styled_table(ws_metrics, metrics_df, "KPI Snapshot", start_row=4)

    kpi_chart = BarChart()
    kpi_chart.title = "KPI Comparison"
    kpi_chart.y_axis.title = "Value"
    kpi_chart.width = 11
    kpi_chart.height = 5.5
    kpi_chart.add_data(Reference(ws_metrics, min_col=2, min_row=5, max_row=5 + len(metrics_df)), titles_from_data=True)
    kpi_chart.set_categories(Reference(ws_metrics, min_col=1, min_row=6, max_row=5 + len(metrics_df)))
    ws_metrics.add_chart(kpi_chart, f"A{row}")

    row += 15
    key_drivers_df = summary_tables["key_drivers"].copy()
    row = _write_styled_table(ws_metrics, key_drivers_df, "Key Drivers", start_row=row)

    annual_core = outputs.annual_summary.reset_index().rename(columns={outputs.annual_summary.index.name or "index": "Year"})
    row = _write_styled_table(
        ws_metrics,
        annual_core[["Year", "revenue_total", "ebitda", "fcff", "equity_cash_flow"]],
        "Annual Performance Trend",
        start_row=row,
    )
    trend_chart = LineChart()
    trend_chart.title = "Revenue / EBITDA / FCFF"
    trend_chart.width = 13
    trend_chart.height = 6
    for col in (2, 3, 4):
        trend_chart.add_data(
            Reference(ws_metrics, min_col=col, min_row=row - len(annual_core) - 1, max_row=row - 2),
            titles_from_data=True,
        )
    trend_chart.set_categories(
        Reference(ws_metrics, min_col=1, min_row=row - len(annual_core), max_row=row - 2)
    )
    ws_metrics.add_chart(trend_chart, f"A{row}")

    row += 15
    revenue_schedule = outputs.monthly_results.filter(like="revenue_").resample("YE").sum()
    revenue_schedule.index = revenue_schedule.index.year
    revenue_schedule = revenue_schedule.reset_index().rename(columns={"index": "Year", "month_start": "Year"})
    row = _write_styled_table(ws_metrics, revenue_schedule, "Gross Revenue Schedule", start_row=row)
    revenue_chart = BarChart()
    revenue_chart.type = "col"
    revenue_chart.grouping = "stacked"
    revenue_chart.title = "Revenue Composition"
    revenue_chart.width = 13
    revenue_chart.height = 5.5
    revenue_chart.add_data(
        Reference(ws_metrics, min_col=2, min_row=row - len(revenue_schedule) - 1, max_col=revenue_schedule.shape[1], max_row=row - 2),
        titles_from_data=True,
    )
    revenue_chart.set_categories(Reference(ws_metrics, min_col=1, min_row=row - len(revenue_schedule), max_row=row - 2))
    ws_metrics.add_chart(revenue_chart, f"A{row}")

    row += 15
    expense_schedule = outputs.monthly_results.filter(like="opex_").resample("YE").sum()
    expense_schedule.index = expense_schedule.index.year
    expense_schedule["total_opex"] = expense_schedule.sum(axis=1)
    expense_schedule = expense_schedule.reset_index().rename(columns={"index": "Year", "month_start": "Year"})
    row = _write_styled_table(ws_metrics, expense_schedule, "Operating Expense Schedule", start_row=row)
    opex_chart = BarChart()
    opex_chart.type = "col"
    opex_chart.grouping = "stacked"
    opex_chart.title = "Operating Cost Composition"
    opex_chart.width = 13
    opex_chart.height = 5.5
    opex_chart.add_data(
        Reference(ws_metrics, min_col=2, min_row=row - len(expense_schedule) - 1, max_col=expense_schedule.shape[1], max_row=row - 2),
        titles_from_data=True,
    )
    opex_chart.set_categories(Reference(ws_metrics, min_col=1, min_row=row - len(expense_schedule), max_row=row - 2))
    ws_metrics.add_chart(opex_chart, f"A{row}")

    row += 15
    debt_schedule = outputs.annual_summary.reset_index().rename(columns={outputs.annual_summary.index.name or "index": "Year"})[
        ["Year", "debt_draw", "debt_principal", "debt_interest"]
    ]
    row = _write_styled_table(ws_metrics, debt_schedule, "Debt Service Schedule", start_row=row)
    debt_chart = LineChart()
    debt_chart.title = "Debt Draw / Principal / Interest"
    debt_chart.width = 13
    debt_chart.height = 5.5
    for col in (2, 3, 4):
        debt_chart.add_data(
            Reference(ws_metrics, min_col=col, min_row=row - len(debt_schedule) - 1, max_row=row - 2),
            titles_from_data=True,
        )
    debt_chart.set_categories(Reference(ws_metrics, min_col=1, min_row=row - len(debt_schedule), max_row=row - 2))
    ws_metrics.add_chart(debt_chart, f"A{row}")

    row += 15
    cash_schedule = outputs.annual_summary.reset_index().rename(columns={outputs.annual_summary.index.name or "index": "Year"})[
        ["Year", "fcff", "equity_cash_flow", "capex", "delta_working_capital"]
    ]
    row = _write_styled_table(ws_metrics, cash_schedule, "Cash Flow Schedule", start_row=row)
    cash_sched_chart = LineChart()
    cash_sched_chart.title = "FCFF vs Equity Cash Flow"
    cash_sched_chart.width = 13
    cash_sched_chart.height = 5.5
    for col in (2, 3):
        cash_sched_chart.add_data(
            Reference(ws_metrics, min_col=col, min_row=row - len(cash_schedule) - 1, max_row=row - 2),
            titles_from_data=True,
        )
    cash_sched_chart.set_categories(Reference(ws_metrics, min_col=1, min_row=row - len(cash_schedule), max_row=row - 2))
    ws_metrics.add_chart(cash_sched_chart, f"A{row}")

    row += 15
    monthly_energy = outputs.monthly_results.reset_index()[["month_start", "energy_mwh"]].rename(
        columns={"month_start": "Month", "energy_mwh": "Monthly Energy Production"}
    )
    row = _write_styled_table(ws_metrics, monthly_energy, "Monthly energy production", start_row=row)
    monthly_energy_chart = LineChart()
    monthly_energy_chart.title = "Monthly energy production"
    monthly_energy_chart.width = 13
    monthly_energy_chart.height = 5.5
    monthly_energy_chart.add_data(
        Reference(ws_metrics, min_col=2, min_row=row - len(monthly_energy) - 1, max_row=row - 2),
        titles_from_data=True,
    )
    monthly_energy_chart.set_categories(
        Reference(ws_metrics, min_col=1, min_row=row - len(monthly_energy), max_row=row - 2)
    )
    ws_metrics.add_chart(monthly_energy_chart, f"A{row}")

    row += 15
    annual_energy = outputs.monthly_results["energy_mwh"].resample("YE").sum().reset_index()
    annual_energy["Year"] = pd.to_datetime(annual_energy["month_start"]).dt.year
    annual_energy = annual_energy.drop(columns=["month_start"]).rename(columns={"energy_mwh": "Energy production"})
    row = _write_styled_table(ws_metrics, annual_energy, "Energy production", start_row=row)
    energy_chart = BarChart()
    energy_chart.title = "Energy production"
    energy_chart.width = 12
    energy_chart.height = 5.5
    energy_chart.add_data(
        Reference(ws_metrics, min_col=2, min_row=row - len(annual_energy) - 1, max_row=row - 2),
        titles_from_data=True,
    )
    energy_chart.set_categories(Reference(ws_metrics, min_col=1, min_row=row - len(annual_energy), max_row=row - 2))
    ws_metrics.add_chart(energy_chart, f"A{row}")

    row += 15
    equity_cf = outputs.monthly_results.reset_index()[["month_start", "equity_cash_flow"]].rename(
        columns={"month_start": "Month", "equity_cash_flow": "Equity cash flow"}
    )
    row = _write_styled_table(ws_metrics, equity_cf, "Equity cash flow", start_row=row)
    equity_chart = LineChart()
    equity_chart.title = "Equity cash flow"
    equity_chart.width = 13
    equity_chart.height = 5.5
    equity_chart.add_data(
        Reference(ws_metrics, min_col=2, min_row=row - len(equity_cf) - 1, max_row=row - 2),
        titles_from_data=True,
    )
    equity_chart.set_categories(Reference(ws_metrics, min_col=1, min_row=row - len(equity_cf), max_row=row - 2))
    ws_metrics.add_chart(equity_chart, f"A{row}")

    row += 15
    total_opex_monthly = outputs.monthly_results.reset_index()[["month_start", "total_opex"]].rename(
        columns={"month_start": "Month", "total_opex": "Total operating cost"}
    )
    row = _write_styled_table(ws_metrics, total_opex_monthly, "Total operating cost", start_row=row)
    total_opex_chart = LineChart()
    total_opex_chart.title = "Total operating cost"
    total_opex_chart.width = 13
    total_opex_chart.height = 5.5
    total_opex_chart.add_data(
        Reference(ws_metrics, min_col=2, min_row=row - len(total_opex_monthly) - 1, max_row=row - 2),
        titles_from_data=True,
    )
    total_opex_chart.set_categories(
        Reference(ws_metrics, min_col=1, min_row=row - len(total_opex_monthly), max_row=row - 2)
    )
    ws_metrics.add_chart(total_opex_chart, f"A{row}")

    row += 15
    cost_breakdown = outputs.monthly_results.filter(like="opex_").resample("YE").sum().reset_index()
    cost_breakdown["Year"] = pd.to_datetime(cost_breakdown["month_start"]).dt.year
    cost_breakdown = cost_breakdown.drop(columns=["month_start"])
    cost_breakdown = cost_breakdown[["Year"] + [c for c in cost_breakdown.columns if c != "Year"]]
    row = _write_styled_table(ws_metrics, cost_breakdown, "Cost breakdown", start_row=row)
    cost_breakdown_chart = BarChart()
    cost_breakdown_chart.type = "col"
    cost_breakdown_chart.grouping = "stacked"
    cost_breakdown_chart.title = "Cost breakdown"
    cost_breakdown_chart.width = 13
    cost_breakdown_chart.height = 5.5
    cost_breakdown_chart.add_data(
        Reference(ws_metrics, min_col=2, min_row=row - len(cost_breakdown) - 1, max_col=cost_breakdown.shape[1], max_row=row - 2),
        titles_from_data=True,
    )
    cost_breakdown_chart.set_categories(
        Reference(ws_metrics, min_col=1, min_row=row - len(cost_breakdown), max_row=row - 2)
    )
    ws_metrics.add_chart(cost_breakdown_chart, f"A{row}")

    row += 15
    capex_debt = outputs.monthly_results.reset_index()[
        ["month_start", "capex", "debt_draw", "debt_principal", "debt_balance"]
    ].rename(columns={"month_start": "Month"})
    row = _write_styled_table(ws_metrics, capex_debt, "Capital expenditure and debt", start_row=row)
    capex_debt_chart = LineChart()
    capex_debt_chart.title = "Capital expenditure and debt"
    capex_debt_chart.width = 13
    capex_debt_chart.height = 5.5
    for col in (2, 3, 4, 5):
        capex_debt_chart.add_data(
            Reference(ws_metrics, min_col=col, min_row=row - len(capex_debt) - 1, max_row=row - 2),
            titles_from_data=True,
        )
    capex_debt_chart.set_categories(
        Reference(ws_metrics, min_col=1, min_row=row - len(capex_debt), max_row=row - 2)
    )
    ws_metrics.add_chart(capex_debt_chart, f"A{row}")

    row += 15
    cash_returns = outputs.monthly_results.reset_index()[
        ["month_start", "fcff", "equity_cash_flow", "investor_cash_flow", "owner_cash_flow"]
    ].rename(columns={"month_start": "Month", "fcff": "FCFF"})
    row = _write_styled_table(ws_metrics, cash_returns, "Cash Flow & Returns", start_row=row)
    cash_returns_chart = LineChart()
    cash_returns_chart.title = "Cash Flow & Returns"
    cash_returns_chart.width = 13
    cash_returns_chart.height = 5.5
    for col in (2, 3, 4, 5):
        cash_returns_chart.add_data(
            Reference(ws_metrics, min_col=col, min_row=row - len(cash_returns) - 1, max_row=row - 2),
            titles_from_data=True,
        )
    cash_returns_chart.set_categories(
        Reference(ws_metrics, min_col=1, min_row=row - len(cash_returns), max_row=row - 2)
    )
    ws_metrics.add_chart(cash_returns_chart, f"A{row}")

    row += 15
    cumulative_cash = outputs.monthly_results.reset_index()[
        ["month_start", "fcff", "equity_cash_flow", "investor_cash_flow", "owner_cash_flow"]
    ].rename(
        columns={
            "month_start": "Month",
            "fcff": "FCFF",
            "equity_cash_flow": "Equity Cash Flow",
            "investor_cash_flow": "Investor Cash Flow",
            "owner_cash_flow": "Owner Cash Flow",
        }
    )
    for col in ["FCFF", "Equity Cash Flow", "Investor Cash Flow", "Owner Cash Flow"]:
        cumulative_cash[f"Cumulative {col}"] = cumulative_cash[col].cumsum()
    cumulative_cash = cumulative_cash[
        [
            "Month",
            "Cumulative FCFF",
            "Cumulative Equity Cash Flow",
            "Cumulative Investor Cash Flow",
            "Cumulative Owner Cash Flow",
        ]
    ]
    row = _write_styled_table(ws_metrics, cumulative_cash, "Cumulative cash flows", start_row=row)
    cumulative_chart = LineChart()
    cumulative_chart.title = "Cumulative cash flows"
    cumulative_chart.width = 13
    cumulative_chart.height = 5.5
    for col in (2, 3, 4, 5):
        cumulative_chart.add_data(
            Reference(ws_metrics, min_col=col, min_row=row - len(cumulative_cash) - 1, max_row=row - 2),
            titles_from_data=True,
        )
    cumulative_chart.set_categories(
        Reference(ws_metrics, min_col=1, min_row=row - len(cumulative_cash), max_row=row - 2)
    )
    ws_metrics.add_chart(cumulative_chart, f"A{row}")

    # ------------------------------------------------------------------
    # Sheet 2: Financials
    ws_fin = wb.create_sheet("Financials")
    _excel_title(ws_fin, "Financials", "Financial Performance, Position, and Cash Flow Statement")

    monthly = outputs.monthly_results
    income = outputs.annual_summary[
        ["revenue_total", "total_opex", "ebitda", "ebit", "tax_payment", "net_income"]
    ].reset_index().rename(columns={outputs.annual_summary.index.name or "index": "Year"})
    row = _write_styled_table(ws_fin, income, "Financial Performance", start_row=4)

    income_chart = BarChart()
    income_chart.title = "EBITDA and Net Income by Year"
    income_chart.width = 12
    income_chart.height = 5.5
    for col in (4, 7):
        income_chart.add_data(
            Reference(ws_fin, min_col=col, min_row=5, max_row=5 + len(income)),
            titles_from_data=True,
        )
    income_chart.set_categories(Reference(ws_fin, min_col=1, min_row=6, max_row=5 + len(income)))
    ws_fin.add_chart(income_chart, f"A{row}")

    row += 15
    gross_revenue = outputs.monthly_results.filter(like="revenue_").resample("YE").sum()
    gross_revenue.index = gross_revenue.index.year
    gross_revenue = gross_revenue.reset_index().rename(columns={"index": "Year", "month_start": "Year"})
    row = _write_styled_table(ws_fin, gross_revenue, "Gross Revenue Schedule", start_row=row)
    gross_chart = BarChart()
    gross_chart.type = "col"
    gross_chart.grouping = "stacked"
    gross_chart.title = "Gross Revenue Schedule"
    gross_chart.width = 12
    gross_chart.height = 5.5
    gross_chart.add_data(
        Reference(ws_fin, min_col=2, min_row=row - len(gross_revenue) - 1, max_col=gross_revenue.shape[1], max_row=row - 2),
        titles_from_data=True,
    )
    gross_chart.set_categories(Reference(ws_fin, min_col=1, min_row=row - len(gross_revenue), max_row=row - 2))
    ws_fin.add_chart(gross_chart, f"A{row}")

    row += 15
    total_expense = outputs.monthly_results.filter(like="opex_").resample("YE").sum()
    total_expense.index = total_expense.index.year
    total_expense["Total"] = total_expense.sum(axis=1)
    total_expense = total_expense.reset_index().rename(columns={"index": "Year", "month_start": "Year"})
    row = _write_styled_table(ws_fin, total_expense, "Total Expense Schedule", start_row=row)
    expense_chart = BarChart()
    expense_chart.type = "col"
    expense_chart.grouping = "stacked"
    expense_chart.title = "Total Expense Composition"
    expense_chart.width = 12
    expense_chart.height = 5.5
    expense_chart.add_data(
        Reference(ws_fin, min_col=2, min_row=row - len(total_expense) - 1, max_col=total_expense.shape[1], max_row=row - 2),
        titles_from_data=True,
    )
    expense_chart.set_categories(Reference(ws_fin, min_col=1, min_row=row - len(total_expense), max_row=row - 2))
    ws_fin.add_chart(expense_chart, f"A{row}")

    opening_ppe = monthly.get("ppe_opening_balance", pd.Series(0.0, index=monthly.index)).cumsum()
    net_ppe = (opening_ppe + monthly["capex"].cumsum() - monthly["depreciation"].cumsum()).clip(lower=0)
    cash_balance = monthly["equity_cash_flow"].cumsum()
    debt_balance = monthly.get("debt_balance", pd.Series(0.0, index=monthly.index))
    total_assets = cash_balance + monthly.get("accounts_receivable", 0) + net_ppe
    total_liabilities = debt_balance + monthly.get("accounts_payable", 0)
    balance = pd.DataFrame(
        {
            "Cash": cash_balance,
            "Net PP&E": net_ppe,
            "Total Assets": total_assets,
            "Debt Outstanding": debt_balance,
            "Total Liabilities": total_liabilities,
            "Equity": total_assets - total_liabilities,
        },
        index=monthly.index,
    ).resample("YE").last()
    balance.index = balance.index.year
    balance_df = balance.reset_index().rename(columns={"index": "Year"})

    row += 15
    row = _write_styled_table(ws_fin, balance_df, "Financial Position", start_row=row)
    position_chart = LineChart()
    position_chart.title = "Assets vs Liabilities vs Equity"
    position_chart.width = 12
    position_chart.height = 5.5
    for col in (4, 6, 7):
        position_chart.add_data(
            Reference(ws_fin, min_col=col, min_row=row - len(balance_df) - 1, max_row=row - 2),
            titles_from_data=True,
        )
    position_chart.set_categories(
        Reference(ws_fin, min_col=1, min_row=row - len(balance_df), max_row=row - 2)
    )
    ws_fin.add_chart(position_chart, f"A{row}")

    ebitda = monthly["ebitda"].resample("YE").sum()
    taxes = monthly["tax_payment"].resample("YE").sum()
    interest = monthly["debt_interest"].resample("YE").sum()
    wc = monthly.get("delta_working_capital", pd.Series(0.0, index=monthly.index)).resample("YE").sum()
    investing_cf = (-monthly["capex"]).resample("YE").sum()
    financing_cf = (monthly["debt_draw"] - monthly["debt_principal"]).resample("YE").sum()
    net_cf = ebitda - taxes - interest - wc + investing_cf + financing_cf
    cash_flow_df = pd.DataFrame(
        {
            "Year": ebitda.index.year,
            "Operating CF": ebitda - taxes - interest - wc,
            "Investing CF": investing_cf,
            "Financing CF": financing_cf,
            "Net Change in Cash": net_cf,
        }
    )

    row += 15
    row = _write_styled_table(ws_fin, cash_flow_df, "Cash Flow Statement", start_row=row)
    cf_chart = LineChart()
    cf_chart.title = "Operating / Investing / Financing Cash Flows"
    cf_chart.width = 12
    cf_chart.height = 5.5
    for col in (2, 3, 4):
        cf_chart.add_data(
            Reference(ws_fin, min_col=col, min_row=row - len(cash_flow_df) - 1, max_row=row - 2),
            titles_from_data=True,
        )
    cf_chart.set_categories(
        Reference(ws_fin, min_col=1, min_row=row - len(cash_flow_df), max_row=row - 2)
    )
    ws_fin.add_chart(cf_chart, f"A{row}")

    # ------------------------------------------------------------------
    # Sheet 3: Key Analytics
    ws_analytics = wb.create_sheet("Key Analytics")
    _excel_title(ws_analytics, "Key Analytics", "Sensitivity, scenario, simulation, and break-even diagnostics")

    sensitivity_records: List[Dict[str, object]] = []
    for variable_key in list(SENSITIVITY_OPTIONS.keys())[:4]:
        label, apply_fn = SENSITIVITY_OPTIONS[variable_key]
        for multiplier in (0.90, 1.00, 1.10):
            metrics = outputs.metrics if math.isclose(multiplier, 1.0) else _simulate_metrics(
                assumptions, lambda a, fn=apply_fn, m=multiplier: fn(a, m)
            )
            sensitivity_records.append(
                {
                    "Variable": label,
                    "Multiplier": multiplier,
                    "Project NPV": metrics.get("project_npv", float("nan")),
                    "Project IRR": metrics.get("project_irr", float("nan")),
                    "Payback (months)": metrics.get("project_payback_months", float("nan")),
                }
            )
    sensitivity_df = pd.DataFrame(sensitivity_records)
    row = _write_styled_table(ws_analytics, sensitivity_df, "Sensitivity Analysis (0.9x / 1.0x / 1.1x)", start_row=4)
    sens_chart = BarChart()
    sens_chart.title = "Sensitivity: Project NPV by Multiplier"
    sens_chart.width = 12
    sens_chart.height = 5.5
    sens_chart.add_data(
        Reference(ws_analytics, min_col=3, min_row=5, max_row=5 + len(sensitivity_df)),
        titles_from_data=True,
    )
    sens_chart.set_categories(Reference(ws_analytics, min_col=2, min_row=6, max_row=5 + len(sensitivity_df)))
    ws_analytics.add_chart(sens_chart, f"A{row}")

    goal_seek_records: List[Dict[str, object]] = []
    target_npv = float(outputs.metrics.get("project_npv", float("nan"))) * 1.1
    for variable_key, (label, apply_fn) in list(SENSITIVITY_OPTIONS.items())[:4]:
        best_multiplier = 1.0
        best_value = float(outputs.metrics.get("project_npv", float("nan")))
        best_gap = abs(best_value - target_npv)
        for multiplier in GOAL_SEEK_MULTIPLIERS:
            metrics = outputs.metrics if math.isclose(multiplier, 1.0) else _simulate_metrics(
                assumptions,
                lambda a, fn=apply_fn, m=multiplier: fn(a, m),
            )
            candidate = float(metrics.get("project_npv", float("nan")))
            gap = abs(candidate - target_npv)
            if gap < best_gap:
                best_gap = gap
                best_multiplier = multiplier
                best_value = candidate
        goal_seek_records.append(
            {
                "Variable": label,
                "Target Project NPV": target_npv,
                "Recommended Multiplier": best_multiplier,
                "Projected Project NPV": best_value,
                "Gap vs Target": best_value - target_npv,
            }
        )
    goal_seek_df = pd.DataFrame(goal_seek_records)
    row += 15
    row = _write_styled_table(ws_analytics, goal_seek_df, "Goal Seek Schedule", start_row=row)
    goal_chart = BarChart()
    goal_chart.title = "Goal Seek: Projected NPV by Variable"
    goal_chart.width = 12
    goal_chart.height = 5.5
    goal_chart.add_data(
        Reference(ws_analytics, min_col=4, min_row=row - len(goal_seek_df) - 1, max_row=row - 2),
        titles_from_data=True,
    )
    goal_chart.set_categories(Reference(ws_analytics, min_col=1, min_row=row - len(goal_seek_df), max_row=row - 2))
    ws_analytics.add_chart(goal_chart, f"A{row}")

    scenario_rows = [
        {"Scenario": "Base", "Revenue Multiplier": 1.00, "Opex Multiplier": 1.00, "Capex Multiplier": 1.00},
        {"Scenario": "Upside", "Revenue Multiplier": 1.05, "Opex Multiplier": 0.95, "Capex Multiplier": 0.95},
        {"Scenario": "Downside", "Revenue Multiplier": 0.95, "Opex Multiplier": 1.05, "Capex Multiplier": 1.05},
    ]

    def _apply_combo_multipliers(
        target: Assumptions,
        revenue_multiplier: float,
        opex_multiplier: float,
        capex_multiplier: float,
    ) -> None:
        _apply_sensitivity_ppa_rate(target, revenue_multiplier)
        _apply_sensitivity_merchant_rate(target, revenue_multiplier)
        _apply_sensitivity_rec_rate(target, revenue_multiplier)
        _apply_sensitivity_fixed_opex(target, opex_multiplier)
        _apply_sensitivity_variable_opex(target, opex_multiplier)
        _apply_sensitivity_capex(target, capex_multiplier)

    scenario_records: List[Dict[str, object]] = []
    for row_cfg in scenario_rows:
        if row_cfg["Scenario"] == "Base":
            metrics = outputs.metrics
        else:
            metrics = _simulate_metrics(
                assumptions,
                lambda a, cfg=row_cfg: _apply_combo_multipliers(
                    a,
                    cfg["Revenue Multiplier"],
                    cfg["Opex Multiplier"],
                    cfg["Capex Multiplier"],
                ),
            )
        scenario_records.append(
            {
                **row_cfg,
                "Project NPV": metrics.get("project_npv", float("nan")),
                "Project IRR": metrics.get("project_irr", float("nan")),
            }
        )
    scenario_df = pd.DataFrame(scenario_records)
    row += 15
    row = _write_styled_table(ws_analytics, scenario_df, "Scenario / IFs Analysis", start_row=row)
    scen_chart = BarChart()
    scen_chart.title = "Scenario Comparison: Project NPV"
    scen_chart.width = 11
    scen_chart.height = 5.5
    scen_chart.add_data(
        Reference(ws_analytics, min_col=5, min_row=row - len(scenario_df) - 1, max_row=row - 2),
        titles_from_data=True,
    )
    scen_chart.set_categories(
        Reference(ws_analytics, min_col=1, min_row=row - len(scenario_df), max_row=row - 2)
    )
    ws_analytics.add_chart(scen_chart, f"A{row}")

    rng = np.random.default_rng(42)
    mc_values = []
    for _ in range(100):
        revenue_mult = float(rng.normal(1.0, 0.05))
        opex_mult = float(rng.normal(1.0, 0.04))
        metrics = _simulate_metrics(
            assumptions,
            lambda a, rm=revenue_mult, om=opex_mult: _apply_combo_multipliers(
                a,
                rm,
                om,
                1.0,
            ),
        )
        mc_values.append(
            {
                "Revenue Multiplier": revenue_mult,
                "Opex Multiplier": opex_mult,
                "Project NPV": metrics.get("project_npv", float("nan")),
            }
        )
    mc_df = pd.DataFrame(mc_values)
    mc_summary = mc_df["Project NPV"].agg(["mean", "std", "min", "max"]).to_frame(name="Value").reset_index()
    mc_summary = mc_summary.rename(columns={"index": "Statistic"})
    row += 15
    row = _write_styled_table(ws_analytics, mc_summary, "Monte Carlo Summary (Project NPV)", start_row=row)
    mc_chart = LineChart()
    mc_chart.title = "Monte Carlo Project NPV Trend (sample order)"
    mc_chart.width = 12
    mc_chart.height = 5.5
    mc_export = mc_df[["Project NPV"]].reset_index().rename(columns={"index": "Run"})
    row = _write_styled_table(ws_analytics, mc_export.head(80), "Monte Carlo Simulation", start_row=row + 14)
    mc_chart.add_data(
        Reference(ws_analytics, min_col=2, min_row=row - 121, max_row=row - 2),
        titles_from_data=True,
    )
    mc_chart.set_categories(Reference(ws_analytics, min_col=1, min_row=row - 120, max_row=row - 2))
    ws_analytics.add_chart(mc_chart, f"A{row}")

    break_even_inputs = pd.DataFrame(st.session_state.get(BREAK_EVEN_STATE_KEY, BREAK_EVEN_DEFAULTS)).copy()
    if not break_even_inputs.empty:
        selected_columns = ["product", "fixed_cost", "variable_cost", "selling_price", "target_profit", "expected_volume"]
        break_even_inputs = break_even_inputs[[c for c in selected_columns if c in break_even_inputs.columns]]
        break_even_inputs = break_even_inputs.rename(
            columns={
                "product": "Product",
                "fixed_cost": "Fixed Cost",
                "variable_cost": "Variable Cost",
                "selling_price": "Selling Price",
                "target_profit": "Target Profit",
                "expected_volume": "Expected Volume",
            }
        )
    row += 15
    row = _write_styled_table(ws_analytics, break_even_inputs, "Break-even Analysis Inputs", start_row=row)

    break_even_results_records: List[Dict[str, object]] = []
    for _, be_input in break_even_inputs.iterrows():
        product = str(be_input.get("Product", "Unlabelled"))
        fixed_cost = float(be_input.get("Fixed Cost", 0.0))
        variable_cost = float(be_input.get("Variable Cost", 0.0))
        selling_price = float(be_input.get("Selling Price", 0.0))
        target_profit = float(be_input.get("Target Profit", 0.0))
        contribution = selling_price - variable_cost
        if contribution > 0:
            units = (fixed_cost + target_profit) / contribution
            revenue = units * selling_price
        else:
            units = float("nan")
            revenue = float("nan")
        break_even_results_records.append(
            {
                "Product": product,
                "Contribution Margin": contribution,
                "Break-even Units": units,
                "Break-even Revenue": revenue,
            }
        )
    break_even_results = pd.DataFrame(break_even_results_records)
    row += 15
    row = _write_styled_table(ws_analytics, break_even_results, "Break-even Results", start_row=row)
    break_even_chart = BarChart()
    break_even_chart.title = "Break-even Results"
    break_even_chart.width = 12
    break_even_chart.height = 5.5
    break_even_chart.add_data(
        Reference(ws_analytics, min_col=3, min_row=row - len(break_even_results) - 1, max_row=row - 2),
        titles_from_data=True,
    )
    break_even_chart.set_categories(
        Reference(ws_analytics, min_col=1, min_row=row - len(break_even_results), max_row=row - 2)
    )
    ws_analytics.add_chart(break_even_chart, f"A{row}")

    be_df = pd.DataFrame(
        {
            "Month": monthly.index.astype(str),
            "Equity Cash Flow": monthly["equity_cash_flow"].values,
            "Cumulative Equity Cash Flow": monthly["equity_cash_flow"].cumsum().values,
            "FCFF": monthly["fcff"].values,
        }
    )
    row += 15
    row = _write_styled_table(ws_analytics, be_df, "Break-Even & Payback", start_row=row)
    be_chart = LineChart()
    be_chart.title = "Cumulative Equity Cash Flow (Break-even Tracking)"
    be_chart.width = 12
    be_chart.height = 5.5
    be_chart.add_data(
        Reference(ws_analytics, min_col=3, min_row=row - len(be_df) - 1, max_row=row - 2),
        titles_from_data=True,
    )
    be_chart.set_categories(
        Reference(ws_analytics, min_col=1, min_row=row - len(be_df), max_row=row - 2)
    )
    ws_analytics.add_chart(be_chart, f"A{row}")

    _add_workbook_overview_sheet(wb, outputs, summary_tables, assumptions)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _coerce_float(value: object, default: float = 0.0) -> float:
    return _coerce_float_helper(value, default)


def _build_opex_items(
    cost_rows: List[Dict[str, object]],
    default_inflation: float,
) -> Tuple[List[FixedOpexItem], List[VariableOpexItem]]:
    return _build_opex_items_helper(cost_rows, default_inflation)


def _coerce_optional_float(value: object) -> float | None:
    return _coerce_optional_float_helper(value)


def _coerce_optional_int(value: object) -> int | None:
    return _coerce_optional_int_helper(value)


def _rows_from_tuple(data: Tuple[Tuple[object, ...], ...], fields: Tuple[str, ...]) -> List[Dict[str, object]]:
    return _rows_from_tuple_helper(data, fields)


def _tupleize(rows: List[Dict[str, object]], fields: Tuple[str, ...]) -> Tuple[Tuple[object, ...], ...]:
    return _tupleize_helper(rows, fields)


def _projection_year_bounds() -> Tuple[int, int]:
    """Return the inclusive start/end years configured for the projection horizon."""

    start_year = int(st.session_state.get("projection_start_year", 2025))
    end_year = int(st.session_state.get("projection_end_year", start_year))
    if end_year < start_year:
        end_year = start_year
    return start_year, end_year


def _projection_year_options() -> List[int]:
    """Return the selectable year options constrained to the projection horizon."""

    start_year, end_year = _projection_year_bounds()
    return list(range(start_year, end_year + 1)) or [start_year]


def _projection_timeline_index() -> pd.DatetimeIndex:
    """Return a monthly timeline covering the configured projection horizon."""

    start_year, end_year = _projection_year_bounds()
    months = max(1, (end_year - start_year + 1) * 12)
    start = pd.Timestamp(year=start_year, month=1, day=1)
    return pd.date_range(start=start, periods=months, freq="MS")


def _ensure_initial_investment_state() -> None:
    """Initialize the initial investment session state if missing."""

    if "initial_investment" not in st.session_state:
        defaults = copy.deepcopy(INITIAL_INVESTMENT_DEFAULTS)
        st.session_state["initial_investment"] = defaults
        _sync_initial_investment_to_fixed_assets()


def _parse_spend_profile(value: object) -> List[float]:
    return _parse_spend_profile_helper(value)


def _capex_item_from_row(row: Dict[str, object], start_year: int) -> CapexItem | None:
    return _capex_item_from_row_helper(row, start_year)


def _capex_item_from_initial(row: Dict[str, object], start_year: int) -> CapexItem | None:
    return _capex_item_from_initial_helper(row, start_year)


def _sync_initial_investment_to_fixed_assets() -> None:
    """Refresh the fixed asset schedule state from the initial investment inputs."""

    if "initial_investment" not in st.session_state:
        return

    start_year, _ = _projection_year_bounds()
    timeline = _projection_timeline_index()
    fixed_rows: List[Dict[str, object]] = []
    for row in st.session_state["initial_investment"]:
        item = _capex_item_from_initial(row, start_year)
        if item is None:
            continue
        _, depreciation_series, _, summary = capex_item_schedule(item, timeline)
        service_month = summary.get("service_month", 1)
        service_year = start_year + max(0, service_month - 1) // 12
        fixed_rows.append(
            {
                "id": row.get("id"),
                "asset_type": summary.get("asset_type", item.name),
                "method": summary.get("method", item.method),
                "year": service_year,
                "acquisition": summary.get("acquisition", item.amount),
                "asset_life": summary.get("depreciation_years", item.depreciation_years),
                "net_book_value": summary.get("opening_balance", item.opening_balance),
                "depreciation_rate": summary.get("depreciation_rate", item.depreciation_rate),
                "total_asset_cost": summary.get("total_asset_cost", item.amount + item.opening_balance),
                "total_depreciation": summary.get("total_depreciation", float(depreciation_series.sum())),
                "cumulative_depreciation": summary.get("cumulative_depreciation", float(depreciation_series.sum())),
                "ending_book_value": summary.get("ending_book_value", 0.0),
                "spend_profile": row.get("spend_profile", "1.0"),
                "service_month": service_month,
                "opening_balance": summary.get("opening_balance", item.opening_balance),
            }
        )

    st.session_state["fixed_assets_schedule"] = fixed_rows


def _format_projection_caption(assumptions: Assumptions) -> str:
    """Return a human-readable summary of the configured projection horizon."""

    start = pd.Timestamp(assumptions.global_assumptions.start_date)
    months = max(1, int(assumptions.global_assumptions.forecast_months))
    end = (start + pd.DateOffset(months=months - 1)).date()
    years = months / 12.0
    return f"Projection horizon: {start.strftime('%Y-%m')} – {end.strftime('%Y-%m')} ({years:.1f} years)"


st.set_page_config(page_title="Solar Farm Financial Model", layout="wide")
st.markdown(
    """
    <style>
    [data-testid="stSidebar"],
    [data-testid="stSidebarNav"],
    [data-testid="collapsedControl"] {
        display: none;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def _inject_app_theme() -> None:
    _inject_app_theme_helper()


def _render_model_hero() -> None:
    _render_model_hero_helper()


def _add_workbook_overview_sheet(
    wb: Workbook,
    outputs: ModelOutputs,
    summary_tables: Dict[str, pd.DataFrame],
    assumptions: Assumptions,
) -> None:
    return _add_workbook_overview_sheet_helper(wb, outputs, summary_tables, assumptions, MetricLabels)
    if "Overview" in wb.sheetnames:
        del wb["Overview"]
    ws = wb.create_sheet("Overview", 0)
    accent = "B45309"
    ws["A1"] = assumptions.global_assumptions.project_name or "Solar Farm Financial Model"
    ws["A1"].font = Font(size=20, bold=True, color="0F172A")
    ws["A2"] = "Executive overview covering key metrics, debt profile, cash flows, and analytics output."
    ws["A2"].font = Font(size=11, color="475569")
    ws["A4"] = "Executive Snapshot"
    ws["A4"].font = Font(size=12, bold=True, color=accent)
    ws["A5"] = "Metric"
    ws["B5"] = "Value"
    for cell in ws[5]:
        cell.fill = PatternFill("solid", fgColor=accent)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    metrics_df = summary_tables.get("metrics", pd.DataFrame()).copy()
    rows: List[Tuple[str, object]] = []
    if not metrics_df.empty:
        for _, metric_row in metrics_df.head(6).iterrows():
            label = MetricLabels.get(metric_row.get("metric"), str(metric_row.get("metric")).replace("_", " ").title())
            rows.append((label, metric_row.get("value")))
    rows.append(("Projection Start", assumptions.global_assumptions.start_date))
    rows.append(("Forecast Months", assumptions.global_assumptions.forecast_months))
    for row_idx, (label, value) in enumerate(rows[:8], start=6):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
    ws["D4"] = "Workbook Notes"
    ws["D4"].font = Font(size=12, bold=True, color=accent)
    notes = [
        "The detailed dashboard, statements, and analytics tabs remain intact in this export.",
        "Use this overview as the executive cover sheet for sponsor and lender circulation.",
        "Key metrics and project assumptions stay aligned to the live model run.",
    ]
    for row_idx, note in enumerate(notes, start=5):
        ws.cell(row=row_idx, column=4, value=f"• {note}")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["D"].width = 60
    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False


GenericTableRow = Dict[str, object]


_TABLE_TYPE_LABELS = {"number": "Number", "percent": "Percent", "boolean": "Boolean"}


CORE_ASSUMPTION_DEFAULTS: List[GenericTableRow] = [
    {
        "id": "discount_rate",
        "label": "Discount Rate",
        "value": 0.10,
        "input_type": "percent",
        "min": 0.0,
        "max": 1.0,
        "step": 0.01,
    },
    {
        "id": "exit_multiple",
        "label": "Exit EBITDA Multiple",
        "value": 5.0,
        "input_type": "number",
        "min": 0.0,
        "step": 0.5,
    },
    {
        "id": "include_terminal",
        "label": "Include Terminal Value",
        "value": True,
        "input_type": "boolean",
    },
    {
        "id": "terminal_growth_rate",
        "label": "Terminal Growth Rate",
        "value": 0.02,
        "input_type": "percent",
        "min": 0.0,
        "max": 0.25,
        "step": 0.005,
    },
]


GLOBAL_DEFAULTS: List[GenericTableRow] = [
    {
        "id": "income_tax_rate",
        "label": "Income Tax Rate",
        "value": 0.25,
        "input_type": "percent",
        "min": 0.0,
        "max": 1.0,
        "step": 0.01,
    },
    {
        "id": "capital_gains_tax_rate",
        "label": "Capital Gains Tax Rate",
        "value": 0.10,
        "input_type": "percent",
        "min": 0.0,
        "max": 1.0,
        "step": 0.01,
    },
    {
        "id": "investor_share",
        "label": "Investor Share",
        "value": 0.95,
        "input_type": "percent",
        "min": 0.0,
        "max": 1.0,
        "step": 0.01,
    },
    {
        "id": "owner_share",
        "label": "Owner Share",
        "value": 0.05,
        "input_type": "percent",
        "min": 0.0,
        "max": 1.0,
        "step": 0.01,
    },
]


ENERGY_COMMON_DEFAULTS: List[GenericTableRow] = [
    {
        "id": "capacity_mw",
        "label": "Capacity (MW)",
        "value": 10.0,
        "input_type": "number",
        "min": 0.0,
        "step": 0.5,
    },
    {
        "id": "degradation_rate",
        "label": "Annual Degradation",
        "value": 0.005,
        "input_type": "percent",
        "min": 0.0,
        "max": 0.10,
        "step": 0.001,
    },
    {
        "id": "panel_watt_dc",
        "label": "Panel Watt (DC)",
        "value": 550.0,
        "input_type": "number",
        "min": 0.0,
        "step": 10.0,
    },
    {
        "id": "dc_ac_ratio",
        "label": "DC/AC Ratio",
        "value": 1.25,
        "input_type": "number",
        "min": 0.1,
        "step": 0.01,
    },
    {
        "id": "annual_production_growth_rate",
        "label": "Annual Production Growth",
        "value": 0.0,
        "input_type": "percent",
        "min": -0.5,
        "max": 1.0,
        "step": 0.005,
    },
    {
        "id": "monthly_min_mwh",
        "label": "Minimum Monthly Generation (MWh)",
        "value": 10.0,
        "input_type": "number",
        "min": 0.0,
        "step": 1.0,
    },
]


ENERGY_CAPACITY_FACTOR_DEFAULTS: List[GenericTableRow] = [
    {
        "id": "capacity_factor",
        "label": "Capacity Factor",
        "value": 0.145,
        "input_type": "percent",
        "min": 0.0,
        "max": 1.0,
        "step": 0.005,
    },
]


ENERGY_RESOURCE_HOURS_DEFAULTS: List[GenericTableRow] = [
    {
        "id": "annual_resource_hours",
        "label": "Annual Resource Hours",
        "value": round(0.145 * CALENDAR_HOURS_PER_YEAR, 1),
        "input_type": "number",
        "min": 0.0,
        "step": 25.0,
    },
]


REVENUE_DEFAULTS: List[GenericTableRow] = [
    {
        "id": "ppa_share",
        "label": "Share of Output via PPA",
        "value": 0.90,
        "input_type": "percent",
        "min": 0.0,
        "max": 1.0,
        "step": 0.05,
    },
    {
        "id": "ppa_rate",
        "label": "Year 1 PPA Rate ($/MWh)",
        "value": 160.0,
        "input_type": "number",
        "min": 0.0,
        "step": 5.0,
    },
    {
        "id": "ppa_escalation",
        "label": "PPA Annual Escalation",
        "value": 0.015,
        "input_type": "number",
        "min": 0.0,
        "max": 0.10,
        "step": 0.005,
        "format": "%.3f",
    },
    {
        "id": "merchant_rate",
        "label": "Year 1 Merchant Rate ($/MWh)",
        "value": 56.58,
        "input_type": "number",
        "min": 0.0,
        "step": 1.0,
    },
    {
        "id": "merchant_escalation",
        "label": "Merchant Annual Escalation",
        "value": 0.015,
        "input_type": "number",
        "min": 0.0,
        "max": 0.10,
        "step": 0.005,
        "format": "%.3f",
    },
    {
        "id": "rec_rate",
        "label": "Year 1 REC Price ($/MWh)",
        "value": 40.0,
        "input_type": "number",
        "min": 0.0,
        "step": 1.0,
    },
    {
        "id": "rec_escalation",
        "label": "REC Annual Escalation",
        "value": 0.02,
        "input_type": "number",
        "min": 0.0,
        "max": 0.10,
        "step": 0.005,
        "format": "%.3f",
    },
]


def _apply_sensitivity_ppa_rate(assumptions: Assumptions, multiplier: float) -> None:
    _apply_sensitivity_ppa_rate_helper(assumptions, multiplier)


def _apply_sensitivity_merchant_rate(assumptions: Assumptions, multiplier: float) -> None:
    _apply_sensitivity_merchant_rate_helper(assumptions, multiplier)


def _apply_sensitivity_rec_rate(assumptions: Assumptions, multiplier: float) -> None:
    _apply_sensitivity_rec_rate_helper(assumptions, multiplier)


def _apply_sensitivity_capacity_factor(assumptions: Assumptions, multiplier: float) -> None:
    _apply_sensitivity_capacity_factor_helper(assumptions, multiplier)


def _apply_sensitivity_capex(assumptions: Assumptions, multiplier: float) -> None:
    _apply_sensitivity_capex_helper(assumptions, multiplier)


def _apply_sensitivity_fixed_opex(assumptions: Assumptions, multiplier: float) -> None:
    _apply_sensitivity_fixed_opex_helper(assumptions, multiplier)


def _apply_sensitivity_variable_opex(assumptions: Assumptions, multiplier: float) -> None:
    _apply_sensitivity_variable_opex_helper(assumptions, multiplier)


def _apply_sensitivity_discount_rate(assumptions: Assumptions, multiplier: float) -> None:
    _apply_sensitivity_discount_rate_helper(assumptions, multiplier)


SENSITIVITY_OPTIONS: Dict[str, Tuple[str, Callable[[Assumptions, float], None]]] = {
    "ppa_rate": ("PPA Rate", _apply_sensitivity_ppa_rate),
    "merchant_rate": ("Merchant Rate", _apply_sensitivity_merchant_rate),
    "rec_rate": ("REC Price", _apply_sensitivity_rec_rate),
    "capacity_factor": ("Capacity Factor", _apply_sensitivity_capacity_factor),
    "capex_total": ("Total Capex", _apply_sensitivity_capex),
    "fixed_opex": ("Fixed Opex", _apply_sensitivity_fixed_opex),
    "variable_opex": ("Variable Opex", _apply_sensitivity_variable_opex),
    "discount_rate": ("Discount Rate", _apply_sensitivity_discount_rate),
}


SENSITIVITY_DEFAULTS: List[Dict[str, object]] = [
    {
        "id": "sensitivity_ppa",
        "variable": "ppa_rate",
        "multipliers": "0.90, 1.00, 1.10",
    },
    {
        "id": "sensitivity_capacity",
        "variable": "capacity_factor",
        "multipliers": "0.95, 1.00, 1.05",
    },
    {
        "id": "sensitivity_capex",
        "variable": "capex_total",
        "multipliers": "0.90, 1.00, 1.10",
    },
]


SENSITIVITY_STATE_KEY = "sensitivity_config"


GOAL_SEEK_SOURCE_LABELS: Dict[str, str] = {
    "metrics": "Model Metrics",
    "income_statement": "Statement of Financial Performance",
    "cash_flow": "Statement of Cash Flows",
}


GOAL_SEEK_METRIC_OPTIONS: Dict[str, Dict[str, str]] = {
    "metrics": MetricLabels,
    "income_statement": {
        "revenue_total": "Gross Revenue",
        "total_opex": "Total Operating Expenses",
        "ebitda": "EBITDA",
        "ebit": "EBIT",
        "net_income": "Net Income",
        "depreciation": "Depreciation",
    },
    "cash_flow": {
        "fcff": "Free Cash Flow to Firm",
        "equity_cash_flow": "Equity Cash Flow",
        "debt_draw": "Debt Draws",
        "debt_principal": "Debt Principal Payments",
        "debt_interest": "Debt Interest",
        "capex": "Capital Expenditure",
        "delta_working_capital": "Change in Working Capital",
    },
}


GOAL_SEEK_DEFAULTS: List[Dict[str, object]] = [
    {
        "id": "goal_seek_npv",
        "source": "metrics",
        "metric": "project_npv",
        "variable": "ppa_rate",
        "target": "2500000",
        "year": "2025",
    },
    {
        "id": "goal_seek_income",
        "source": "income_statement",
        "metric": "net_income",
        "variable": "capacity_factor",
        "target": "1500000",
        "year": "2026",
    },
]


GOAL_SEEK_STATE_KEY = "goal_seek_config"


GOAL_SEEK_MULTIPLIERS = tuple(float(x) for x in np.linspace(0.5, 1.5, 41))


MONTHLY_GENERATION_DEFAULTS = [
    {"month": "January", "expected_mwh": 635.1},
    {"month": "February", "expected_mwh": 635.1},
    {"month": "March", "expected_mwh": 635.1},
    {"month": "April", "expected_mwh": 1270.2},
    {"month": "May", "expected_mwh": 1524.2},
    {"month": "June", "expected_mwh": 2159.3},
    {"month": "July", "expected_mwh": 2159.3},
    {"month": "August", "expected_mwh": 1270.2},
    {"month": "September", "expected_mwh": 635.1},
    {"month": "October", "expected_mwh": 508.1},
    {"month": "November", "expected_mwh": 635.1},
    {"month": "December", "expected_mwh": 635.1},
]


LABOUR_DEFAULTS = [
    {"role": "Plant Manager", "annual_cost": 95_000.0},
    {"role": "Field Technicians", "annual_cost": 180_000.0},
    {"role": "Control Room Operator", "annual_cost": 80_000.0},
    {"role": "Maintenance Crew", "annual_cost": 150_000.0},
]


INITIAL_INVESTMENT_DEFAULTS = [
    {
        "id": "capex_panels",
        "name": "Solar Panels",
        "amount": 5_800_000.0,
        "depreciation_years": 20,
        "method": "Straight-Line",
        "year": 2025,
        "month": 1,
        "spend_profile": "0.5, 0.5",
        "opening_balance": 0.0,
        "depreciation_rate": 0.0,
        "service_month": 1,
    },
    {
        "id": "capex_mounting",
        "name": "Mounting System",
        "amount": 1_600_000.0,
        "depreciation_years": 20,
        "method": "Straight-Line",
        "year": 2025,
        "month": 1,
        "spend_profile": "0.5, 0.5",
        "opening_balance": 0.0,
        "depreciation_rate": 0.0,
        "service_month": 1,
    },
    {
        "id": "capex_inverters",
        "name": "Inverters",
        "amount": 1_776_000.0,
        "depreciation_years": 15,
        "method": "Straight-Line",
        "year": 2025,
        "month": 1,
        "spend_profile": "0.5, 0.5",
        "opening_balance": 0.0,
        "depreciation_rate": 0.0,
        "service_month": 1,
    },
    {
        "id": "capex_electrical",
        "name": "Electrical & Wiring",
        "amount": 1_800_000.0,
        "depreciation_years": 20,
        "method": "Straight-Line",
        "year": 2025,
        "month": 1,
        "spend_profile": "0.5, 0.5",
        "opening_balance": 0.0,
        "depreciation_rate": 0.0,
        "service_month": 1,
    },
    {
        "id": "capex_land",
        "name": "Land Acquisition",
        "amount": 200_000.0,
        "depreciation_years": 20,
        "method": "Straight-Line",
        "year": 2025,
        "month": 1,
        "spend_profile": "0.5, 0.5",
        "opening_balance": 0.0,
        "depreciation_rate": 0.0,
        "service_month": 1,
    },
    {
        "id": "capex_permitting",
        "name": "Permitting & Compliance",
        "amount": 100_000.0,
        "depreciation_years": 10,
        "method": "Straight-Line",
        "year": 2025,
        "month": 1,
        "spend_profile": "0.5, 0.5",
        "opening_balance": 0.0,
        "depreciation_rate": 0.0,
        "service_month": 1,
    },
    {
        "id": "capex_labour",
        "name": "Construction Labour",
        "amount": 500_000.0,
        "depreciation_years": 5,
        "method": "Straight-Line",
        "year": 2025,
        "month": 1,
        "spend_profile": "0.5, 0.5",
        "opening_balance": 0.0,
        "depreciation_rate": 0.0,
        "service_month": 1,
    },
    {
        "id": "capex_contingency",
        "name": "Contingency",
        "amount": 1_177_608.0,
        "depreciation_years": 5,
        "method": "Straight-Line",
        "year": 2025,
        "month": 1,
        "spend_profile": "0.5, 0.5",
        "opening_balance": 0.0,
        "depreciation_rate": 0.0,
        "service_month": 1,
    },
]


OPERATING_EXPENSE_DEFAULTS = [
    {
        "id": "insurance",
        "name": "Insurance",
        "fixed_cost": 20_000.0,
        "variable_cost": 0.0,
        "inflation_rate": 0.02,
    },
    {
        "id": "om_contract",
        "name": "O&M Service Contract",
        "fixed_cost": 6_000.0,
        "variable_cost": 0.0,
        "inflation_rate": 0.02,
    },
    {
        "id": "vegetation",
        "name": "Vegetation Management",
        "fixed_cost": 10_000.0,
        "variable_cost": 0.0,
        "inflation_rate": 0.02,
    },
    {
        "id": "g_and_a",
        "name": "General & Administrative",
        "fixed_cost": 100_000.0,
        "variable_cost": 0.0,
        "inflation_rate": 0.03,
    },
    {
        "id": "sales_marketing",
        "name": "Sales & Marketing",
        "fixed_cost": 8_500.0,
        "variable_cost": 0.0,
        "inflation_rate": 0.02,
    },
    {
        "id": "research_development",
        "name": "Research & Development",
        "fixed_cost": 7_650.0,
        "variable_cost": 0.0,
        "inflation_rate": 0.02,
    },
]


ACCOUNTS_RECEIVABLE_DEFAULTS = [
    {"year": 2025, "days_in_year": 365, "receivable_days": 45, "prepaid_expense_days": 30, "other_asset_days": 5},
    {"year": 2026, "days_in_year": 365, "receivable_days": 45, "prepaid_expense_days": 30, "other_asset_days": 5},
]


INVENTORY_PAYABLE_DEFAULTS = [
    {"year": 2025, "days_in_year": 365, "inventory_days": 50, "accounts_payable_days": 45},
    {"year": 2026, "days_in_year": 365, "inventory_days": 50, "accounts_payable_days": 45},
]


SCENARIO_DEFAULTS: List[Dict[str, object]] = [
    {
        "id": "scenario_base",
        "name": "Scenario - base",
        "inflation_rate": 0.020,
        "interest_rate": 0.050,
        "revenue_multiplier": 1.00,
        "opex_multiplier": 1.00,
        "capex_multiplier": 1.00,
        "capacity_multiplier": 1.00,
    },
    {
        "id": "scenario_best",
        "name": "Scenario - best",
        "inflation_rate": 0.018,
        "interest_rate": 0.045,
        "revenue_multiplier": 1.05,
        "opex_multiplier": 0.95,
        "capex_multiplier": 0.95,
        "capacity_multiplier": 1.05,
    },
    {
        "id": "scenario_worst",
        "name": "Scenario - worst",
        "inflation_rate": 0.024,
        "interest_rate": 0.055,
        "revenue_multiplier": 0.95,
        "opex_multiplier": 1.05,
        "capex_multiplier": 1.05,
        "capacity_multiplier": 0.95,
    },
]


SCENARIO_STATE_KEY = "scenario_if_config"


SCENARIO_TOOL_SECTIONS = [
    ("decision_tree", {"label": "Decision Tree Tools", "show_iterate": True}),
    ("stress_testing", {"label": "Stress Testing", "show_iterate": False}),
    ("backtesting", {"label": "Backtesting", "show_iterate": True}),
    ("walk_forward", {"label": "Walk-forward Testing", "show_iterate": True}),
    ("driver_model", {"label": "Driver-based Modeling", "show_iterate": True}),
    ("real_options", {"label": "Real Options Analysis (ROA)", "show_iterate": False}),
]


SCENARIO_TOOL_DEFAULTS: Dict[str, Dict[str, object]] = {
    "decision_tree": {
        "variables": [
            {"id": "decision_tree_rev", "variable": "Net Revenue"},
            {"id": "decision_tree_income", "variable": "Net Income"},
        ],
        "iterate": "Capacity factor +/- 5%",
    },
    "stress_testing": {
        "variables": [
            {"id": "stress_revenue", "variable": "Net Revenue"},
            {"id": "stress_ebitda", "variable": "EBITDA"},
        ],
        "iterate": "",
    },
    "backtesting": {
        "variables": [
            {"id": "backtesting_revenue", "variable": "Net Revenue"},
        ],
        "iterate": "Historic performance window",
    },
    "walk_forward": {
        "variables": [
            {"id": "walkforward_revenue", "variable": "Net Revenue"},
        ],
        "iterate": "Rolling 12-month recalibration",
    },
    "driver_model": {
        "variables": [
            {"id": "driver_ebitda", "variable": "EBITDA"},
            {"id": "driver_ebit", "variable": "EBIT"},
        ],
        "iterate": "Update quarterly drivers",
    },
    "real_options": {
        "variables": [
            {"id": "roa_income", "variable": "Net Income"},
        ],
        "iterate": "",
    },
}


SCENARIO_TOOL_STATE_KEY = "scenario_tool_config"


LOAN_SCHEDULE_DEFAULTS = [
    {
        "name": "Construction Loan",
        "year": 2025,
        "duration_years": 5,
        "amount": 2_000_000.0,
        "interest_rate": 0.06,
    },
]


TAX_SCHEDULE_DEFAULTS = [
    {"name": "Federal Tax", "year": 2025, "tax_rate": 0.25},
    {"name": "Federal Tax", "year": 2026, "tax_rate": 0.25},
]


INFLATION_SCHEDULE_DEFAULTS = [
    {"name": "Base Inflation", "year": 2025, "rate": 0.025},
    {"name": "Base Inflation", "year": 2026, "rate": 0.025},
]


RISK_SCHEDULE_DEFAULTS = [
    {
        "name": "Baseline",
        "year": 2025,
        "inherent_risk": 0.05,
        "climate_risk": 0.02,
        "political_risk": 0.03,
    },
    {
        "name": "Baseline",
        "year": 2026,
        "inherent_risk": 0.05,
        "climate_risk": 0.02,
        "political_risk": 0.03,
    },
]


MONTE_CARLO_DISTRIBUTION_OPTIONS: Dict[str, str] = {
    "normal": "Normal (mean & std)",
    "uniform": "Uniform (min & max)",
    "triangular": "Triangular (min, mode, max)",
}


MONTE_CARLO_DRIVER_DEFAULTS: List[Dict[str, object]] = [
    {
        "id": "mc_capacity",
        "variable": "capacity_factor",
        "distribution": "normal",
        "mean": 1.0,
        "std_dev": 0.04,
        "min": 0.85,
        "max": 1.10,
    },
    {
        "id": "mc_ppa",
        "variable": "ppa_rate",
        "distribution": "normal",
        "mean": 1.0,
        "std_dev": 0.05,
        "min": 0.80,
        "max": 1.20,
    },
    {
        "id": "mc_capex",
        "variable": "capex_total",
        "distribution": "triangular",
        "min": 0.90,
        "mode": 1.00,
        "max": 1.10,
    },
]


MONTE_CARLO_METRIC_DEFAULTS: List[str] = [
    "project_npv",
    "project_irr",
    "equity_irr",
    "project_payback_months",
]


MONTE_CARLO_CONFIG_DEFAULT: Dict[str, object] = {
    "iterations": 500,
    "seed": 42,
    "drivers": copy.deepcopy(MONTE_CARLO_DRIVER_DEFAULTS),
    "metrics": MONTE_CARLO_METRIC_DEFAULTS.copy(),
}


MONTE_CARLO_STATE_KEY = "monte_carlo_config"


BREAK_EVEN_STATE_KEY = "break_even_inputs"


BREAK_EVEN_DEFAULTS: List[Dict[str, object]] = [
    {
        "id": "be_ppa",
        "product": "Utility PPA",
        "fixed_cost": 650_000.0,
        "variable_cost": 18.0,
        "selling_price": 45.0,
        "target_profit": 150_000.0,
        "expected_volume": 120_000.0,
    },
    {
        "id": "be_merchant",
        "product": "Merchant Sales",
        "fixed_cost": 420_000.0,
        "variable_cost": 20.0,
        "selling_price": 55.0,
        "target_profit": 100_000.0,
        "expected_volume": 95_000.0,
    },
]


def _ensure_table_state(state_key: str, defaults: List[GenericTableRow]) -> None:
    if state_key not in st.session_state:
        st.session_state[state_key] = copy.deepcopy(defaults)
    else:
        existing_rows = st.session_state[state_key]
        existing_ids = {str(row.get("id")) for row in existing_rows if row.get("id")}
        for default_row in defaults:
            default_id = str(default_row.get("id")) if default_row.get("id") is not None else None
            if default_id and default_id not in existing_ids:
                new_row = copy.deepcopy(default_row)
                existing_rows.append(new_row)
                existing_ids.add(default_id)

    if defaults and "id" in defaults[0]:
        for row in st.session_state[state_key]:
            row.setdefault("id", uuid.uuid4().hex)


def _ensure_schedule_row_ids(state_key: str) -> None:
    rows = st.session_state.get(state_key, [])
    for row in rows:
        row.setdefault("id", uuid.uuid4().hex)


def _schedule_edit_state_key(state_key: str) -> str:
    return f"{state_key}_editing_row"


def _schedule_increment_state_key(state_key: str) -> str:
    return f"{state_key}_annual_increment_percent"


def _clamp_numeric(value: float, minimum: float | None = None, maximum: float | None = None) -> float:
    if minimum is not None:
        value = max(minimum, value)
    if maximum is not None:
        value = min(maximum, value)
    return value


def _apply_flat_increment(
    rows: List[Dict[str, object]],
    fields: Tuple[str, ...],
    increment_percent: float,
    *,
    minimum: float = 0.0,
    maximums: Dict[str, float] | None = None,
) -> List[Dict[str, object]]:
    multiplier = 1.0 + (increment_percent / 100.0)
    updated_rows = copy.deepcopy(rows)
    for row in updated_rows:
        for field in fields:
            base_value = _coerce_float(row.get(field))
            max_value = maximums.get(field) if maximums else None
            row[field] = _clamp_numeric(base_value * multiplier, minimum, max_value)
    return updated_rows


def _apply_sequential_annual_increment(
    rows: List[Dict[str, object]],
    fields: Tuple[str, ...],
    increment_percent: float,
    *,
    year_field: str = "year",
    minimum: float = 0.0,
    maximums: Dict[str, float] | None = None,
) -> List[Dict[str, object]]:
    if len(rows) < 2:
        return copy.deepcopy(rows)

    multiplier = 1.0 + (increment_percent / 100.0)
    indexed_rows = list(enumerate(copy.deepcopy(rows)))
    indexed_rows.sort(key=lambda item: (int(item[1].get(year_field, 0)), item[0]))

    for idx in range(1, len(indexed_rows)):
        prior_row = indexed_rows[idx - 1][1]
        current_row = indexed_rows[idx][1]
        for field in fields:
            prior_value = _coerce_float(prior_row.get(field))
            max_value = maximums.get(field) if maximums else None
            current_row[field] = _clamp_numeric(prior_value * multiplier, minimum, max_value)

    restored_rows = [None] * len(indexed_rows)
    for original_index, row in indexed_rows:
        restored_rows[original_index] = row
    return [row for row in restored_rows if row is not None]


def _next_projection_year_for_rows(rows: List[Dict[str, object]], year_options: List[int], start_year: int) -> int:
    next_year = max((int(row.get("year", start_year)) for row in rows), default=start_year) + 1
    if year_options:
        next_year = min(next_year, year_options[-1])
        next_year = max(next_year, year_options[0])
    return next_year


def _render_label_value_table(title: str, state_key: str, defaults: List[GenericTableRow]) -> None:
    _ensure_table_state(state_key, defaults)
    st.markdown(f"### {title}")
    rows = st.session_state[state_key]
    updated_rows: List[GenericTableRow] = []

    for idx, row in enumerate(rows):
        with st.container(border=True):
            col_label, col_type, col_value, col_remove = st.columns([3, 1.5, 2, 1])
            row_id = row.get("id") or uuid.uuid4().hex
            label = col_label.text_input(
                "Label",
                value=str(row.get("label", "")),
                key=f"{state_key}_label_{idx}",
            )

            type_options = list(_TABLE_TYPE_LABELS.keys())
            current_type = str(row.get("input_type", "number"))
            if current_type not in type_options:
                current_type = "number"
            type_index = type_options.index(current_type)
            input_type = col_type.selectbox(
                "Type",
                type_options,
                index=type_index,
                key=f"{state_key}_type_{idx}",
                format_func=lambda opt: _TABLE_TYPE_LABELS[opt],
            )

            value_key = f"{state_key}_value_{idx}"
            if input_type == "boolean":
                value = col_value.checkbox(
                    "Value",
                    value=bool(row.get("value", False)),
                    key=value_key,
                )
            else:
                number_kwargs: Dict[str, float | str] = {}
                if row.get("min") is not None:
                    number_kwargs["min_value"] = float(row["min"])
                if row.get("max") is not None:
                    number_kwargs["max_value"] = float(row["max"])
                step = float(row.get("step", 0.01 if input_type == "percent" else 1.0))
                number_kwargs["step"] = step
                if row.get("format"):
                    number_kwargs["format"] = str(row["format"])
                value = col_value.number_input(
                    "Value",
                    value=float(row.get("value", 0.0)),
                    key=value_key,
                    **number_kwargs,
                )

            remove_clicked = col_remove.button("Remove", key=f"{state_key}_remove_{idx}")
        if remove_clicked:
            continue
        updated_rows.append(
            {
                "id": row_id,
                "label": label,
                "input_type": input_type,
                "value": value,
                **{k: v for k, v in row.items() if k not in {"id", "label", "input_type", "value"}},
            }
        )

    st.session_state[state_key] = updated_rows

    if st.button(f"Add {title} Row", key=f"{state_key}_add"):
        st.session_state[state_key].append(
            {
                "id": uuid.uuid4().hex,
                "label": "New Item",
                "input_type": "number",
                "value": 0.0,
                "step": 0.1,
            }
        )


def _render_monthly_generation_table() -> List[Dict[str, object]]:
    state_key = "monthly_generation_table"
    if state_key not in st.session_state:
        st.session_state[state_key] = copy.deepcopy(MONTHLY_GENERATION_DEFAULTS)
    _ensure_schedule_row_ids(state_key)

    st.markdown("### Expected Monthly Production (MWh)")
    action_cols = st.columns([1.4, 1, 1, 3.6])
    increment_key = _schedule_increment_state_key(state_key)
    st.session_state.setdefault(increment_key, 0.0)
    increment_percent = action_cols[0].number_input(
        "Annual Increment (%)",
        key=increment_key,
        min_value=-100.0,
        step=0.5,
        format="%.2f",
        help="Apply the same percentage change to each monthly production row.",
    )
    if action_cols[1].button("Apply Increment", key=f"{state_key}_apply_increment"):
        st.session_state[state_key] = _apply_flat_increment(
            st.session_state[state_key],
            ("expected_mwh",),
            float(increment_percent),
            minimum=0.0,
        )
    if action_cols[2].button("Add Row", key=f"{state_key}_add"):
        default_month = f"Month {len(st.session_state[state_key]) + 1}"
        st.session_state[state_key].append(
            {
                "id": uuid.uuid4().hex,
                "month": default_month,
                "expected_mwh": 0.0,
            }
        )
        st.session_state[_schedule_edit_state_key(state_key)] = st.session_state[state_key][-1]["id"]

    rows = st.session_state[state_key]
    updated_rows: List[Dict[str, object]] = []
    editing_row_id = st.session_state.get(_schedule_edit_state_key(state_key))

    for idx, row in enumerate(rows):
        row_id = str(row.get("id"))
        summary_cols = st.columns([3.4, 1.2, 1])
        summary_cols[0].markdown(
            f"**{str(row.get('month', '')).strip() or f'Month {idx + 1}'}**  |  "
            f"{float(row.get('expected_mwh', 0.0)):,.2f} MWh"
        )
        if summary_cols[1].button("Edit Row", key=f"{state_key}_edit_{row_id}"):
            st.session_state[_schedule_edit_state_key(state_key)] = row_id
            editing_row_id = row_id
        remove_clicked = summary_cols[2].button("Remove", key=f"{state_key}_remove_{row_id}")
        if remove_clicked:
            if editing_row_id == row_id:
                st.session_state[_schedule_edit_state_key(state_key)] = None
            continue

        updated_row = copy.deepcopy(row)
        if editing_row_id == row_id:
            with st.container(border=True):
                edit_cols = st.columns([3, 2, 1])
                updated_row["month"] = edit_cols[0].text_input(
                    "Month",
                    value=str(row.get("month", "")),
                    key=f"{state_key}_month_{row_id}",
                )
                updated_row["expected_mwh"] = edit_cols[1].number_input(
                    "Expected MWh",
                    value=float(row.get("expected_mwh", 0.0)),
                    key=f"{state_key}_expected_mwh_{row_id}",
                    min_value=0.0,
                    step=1.0,
                )
                if edit_cols[2].button("Done", key=f"{state_key}_done_{row_id}"):
                    st.session_state[_schedule_edit_state_key(state_key)] = None
                    editing_row_id = None

        updated_rows.append(updated_row)

    st.session_state[state_key] = updated_rows

    if updated_rows:
        if len(updated_rows) != 12:
            st.warning("Monthly production table should include 12 months to align with the model timeline.")

    return updated_rows


def _render_projection_horizon_section() -> Tuple[int, int]:
    st.markdown("### Projection Horizon")
    if "projection_start_year" not in st.session_state:
        st.session_state["projection_start_year"] = 2025
    if "projection_end_year" not in st.session_state:
        st.session_state["projection_end_year"] = 2044

    start_year_options = list(range(2020, 2051))
    start_year = st.selectbox(
        "Start Year",
        options=start_year_options,
        index=start_year_options.index(st.session_state["projection_start_year"]),
        key="projection_start_year",
        help="Select the first year in the forecast horizon.",
    )
    end_year_options = list(range(start_year + 1, start_year + 21))
    if st.session_state["projection_end_year"] not in end_year_options:
        st.session_state["projection_end_year"] = end_year_options[-1]
    end_year = st.selectbox(
        "End Year",
        options=end_year_options,
        index=end_year_options.index(st.session_state["projection_end_year"]),
        key="projection_end_year",
        help="Choose the final year for the projection horizon (up to 20 years).",
    )
    return start_year, int(end_year)


def _render_energy_input_mode_selector() -> str:
    st.markdown("### Energy Input Method")
    options = list(ENERGY_INPUT_MODE_OPTIONS.keys())
    default_mode = str(st.session_state.get("energy_input_mode", "capacity_factor"))
    if default_mode not in options:
        default_mode = "capacity_factor"
    selected_mode = st.selectbox(
        "Production Driver",
        options=options,
        index=options.index(default_mode),
        format_func=lambda option: ENERGY_INPUT_MODE_OPTIONS[option],
        key="energy_input_mode",
        help="Choose one production input path. Capacity factor, resource hours, and monthly MWh are mutually exclusive.",
    )
    if selected_mode == "capacity_factor":
        st.caption("Annual hours are fixed at 8,760. Enter capacity factor directly.")
    elif selected_mode == "resource_hours":
        st.caption("Annual hours are fixed at 8,760. Resource hours are converted into implied capacity factor.")
    else:
        st.caption("Provide 12 monthly expected MWh values. The model derives the implied capacity factor.")
    return str(selected_mode)


PANEL_UNIT_COST_DEFAULT = 250.0


def _solar_panels_amount(rows: List[Dict[str, object]]) -> float:
    """Return the amount of the Solar Panels CAPEX line from initial investment rows."""
    for row in rows:
        if str(row.get("name", "")).strip().lower() == "solar panels":
            return float(row.get("amount", 0.0))
    return 0.0


def _derived_panel_count(solar_panels_amount: float, panel_unit_cost: float) -> float:
    """Derive panel count from CAPEX and unit cost."""
    if panel_unit_cost <= 0:
        return 0.0
    return max(0.0, solar_panels_amount) / panel_unit_cost


def _render_initial_investment_section() -> None:
    _ensure_initial_investment_state()
    state_key = "initial_investment"
    st.markdown("### Initial Investment Input Table")

    rows = st.session_state[state_key]
    updated_rows: List[Dict[str, object]] = []
    method_options = ["Straight-Line", "Declining Balance"]
    start_year, _ = _projection_year_bounds()
    year_options = _projection_year_options()
    if not year_options:
        year_options = [start_year]

    month_options = list(range(1, 13))
    month_labels = {
        1: "Jan",
        2: "Feb",
        3: "Mar",
        4: "Apr",
        5: "May",
        6: "Jun",
        7: "Jul",
        8: "Aug",
        9: "Sep",
        10: "Oct",
        11: "Nov",
        12: "Dec",
    }

    for idx, row in enumerate(rows):
        row_id = row.get("id") or uuid.uuid4().hex
        row["id"] = row_id
        with st.container(border=True):
            col_name, col_amount, col_life, col_method = st.columns([3, 2, 1.5, 1.5])
            name = col_name.text_input(
                "Investment Item",
                value=str(row.get("name", "")),
                key=f"{state_key}_name_{row_id}",
            )
            amount = col_amount.number_input(
                "Amount",
                value=float(row.get("amount", 0.0)),
                key=f"{state_key}_amount_{row_id}",
                min_value=0.0,
                step=1000.0,
                format="%.2f",
            )
            depreciation_years = col_life.number_input(
                "Depreciation Years",
                value=float(row.get("depreciation_years", 1.0)),
                key=f"{state_key}_life_{row_id}",
                min_value=0.0,
                step=1.0,
                help="Use 0 for non-depreciable items such as land.",
            )
            method_value = str(row.get("method", method_options[0]))
            if method_value not in method_options:
                method_value = method_options[0]
            method = col_method.selectbox(
                "Method",
                method_options,
                index=method_options.index(method_value),
                key=f"{state_key}_method_{row_id}",
            )

            col_year, col_month, col_service, col_remove = st.columns([1.5, 1.5, 1.5, 1])
            current_year = int(row.get("year", year_options[0]))
            if current_year not in year_options:
                current_year = year_options[0]
            year = col_year.selectbox(
                "Service Year",
                options=year_options,
                index=year_options.index(current_year),
                key=f"{state_key}_year_{row_id}",
            )
            current_month = int(row.get("month", 1))
            if current_month not in month_options:
                current_month = 1
            month = col_month.selectbox(
                "Service Month",
                options=month_options,
                index=month_options.index(current_month),
                format_func=lambda m: month_labels.get(m, str(m)),
                key=f"{state_key}_month_{row_id}",
            )
            service_month = col_service.number_input(
                "Service Month #",
                value=float(row.get("service_month", (year - start_year) * 12 + month)),
                key=f"{state_key}_service_{row_id}",
                min_value=1.0,
                step=1.0,
            )
            remove_clicked = col_remove.button("Remove", key=f"{state_key}_remove_{row_id}")

            col_profile, col_opening, col_rate = st.columns([3, 2, 1.5])
            profile_value = str(row.get("spend_profile", "1.0"))
            spend_profile = col_profile.text_input(
                "Spend Profile",
                value=profile_value,
                key=f"{state_key}_profile_{row_id}",
                help="Comma separated weights (e.g. 0.5, 0.5). Values above 1 are treated as percentages.",
            )
            opening_balance = col_opening.number_input(
                "Opening Balance",
                value=float(row.get("opening_balance", 0.0)),
                key=f"{state_key}_opening_{row_id}",
                min_value=0.0,
                step=1000.0,
                format="%.2f",
            )
            depreciation_rate_percent = col_rate.number_input(
                "Depreciation Rate (%)",
                value=float(row.get("depreciation_rate", 0.0)) * 100.0,
                key=f"{state_key}_rate_{row_id}",
                min_value=0.0,
                max_value=100.0,
                step=0.1,
                format="%.2f",
            )

        if remove_clicked:
            continue

        updated_rows.append(
            {
                "id": row_id,
                "name": name,
                "amount": amount,
                "depreciation_years": depreciation_years,
                "method": method,
                "year": year,
                "month": month,
                "spend_profile": spend_profile,
                "opening_balance": opening_balance,
                "depreciation_rate": depreciation_rate_percent / 100.0,
                "service_month": service_month,
            }
        )

    st.session_state[state_key] = updated_rows
    _sync_initial_investment_to_fixed_assets()

    total_equity = sum(float(row.get("amount", 0.0)) for row in st.session_state[state_key])
    st.markdown(f"**Total Equity:** ${total_equity:,.2f}")

    solar_panels_amount = _solar_panels_amount(st.session_state[state_key])

    panel_unit_cost = st.number_input(
        "Panel Unit Cost ($/panel)",
        min_value=0.0,
        value=float(st.session_state.get("panel_unit_cost_input", PANEL_UNIT_COST_DEFAULT)),
        step=5.0,
        key="panel_unit_cost_input",
        help="Used to derive panel count as Solar Panels amount ÷ panel unit cost.",
    )
    derived_panel_count = _derived_panel_count(solar_panels_amount, panel_unit_cost)
    st.caption(
        f"Derived Panel Count = Solar Panels (${solar_panels_amount:,.2f}) ÷ "
        f"Unit Cost (${panel_unit_cost:,.2f}) = **{derived_panel_count:,.0f} panels**"
    )

    st.caption(
        "Depreciation schedules are generated automatically from these entries; "
        "a separate fixed asset editor is no longer required."
    )

    if st.button("Add Investment Item", key=f"{state_key}_add"):
        st.session_state[state_key].append(
            {
                "id": uuid.uuid4().hex,
                "name": "New Investment",
                "amount": 0.0,
                "depreciation_years": 5.0,
                "method": "Straight-Line",
                "year": year_options[0],
                "month": 1,
                "spend_profile": "1.0",
                "opening_balance": 0.0,
                "depreciation_rate": 0.0,
                "service_month": 1.0,
            }
        )
        _sync_initial_investment_to_fixed_assets()


def _render_labour_structure_section() -> None:
    state_key = "labour_structure"
    if state_key not in st.session_state:
        st.session_state[state_key] = copy.deepcopy(LABOUR_DEFAULTS)

    st.markdown("### Direct Labour Structure")
    rows = st.session_state[state_key]
    updated_rows: List[Dict[str, object]] = []
    for idx, row in enumerate(rows):
        with st.container(border=True):
            col_role, col_cost, col_remove = st.columns([3, 2, 1])
            role = col_role.text_input(
                "Role",
                value=str(row.get("role", "")),
                key=f"{state_key}_role_{idx}",
            )
            annual_cost = col_cost.number_input(
                "Annual Cost",
                value=float(row.get("annual_cost", 0.0)),
                key=f"{state_key}_cost_{idx}",
                step=100.0,
                min_value=0.0,
            )
            remove_clicked = col_remove.button("Remove", key=f"{state_key}_remove_{idx}")
        if remove_clicked:
            continue
        updated_rows.append({"role": role, "annual_cost": annual_cost})
    st.session_state[state_key] = updated_rows

    if st.button("Add Labour Role", key=f"{state_key}_add"):
        st.session_state[state_key].append({"role": "New Role", "annual_cost": 0.0})


def _render_operating_expense_section() -> None:
    state_key = "operating_expenses"
    if state_key not in st.session_state:
        st.session_state[state_key] = copy.deepcopy(OPERATING_EXPENSE_DEFAULTS)

    st.markdown("### Operating Expenses Input Table")
    rows = st.session_state[state_key]
    updated_rows: List[Dict[str, object]] = []
    for row in rows:
        row_id = row.get("id") or uuid.uuid4().hex
        row["id"] = row_id
        with st.container(border=True):
            col_name, col_fixed, col_variable, col_infl, col_remove = st.columns([3, 2, 2, 2, 1])
            name = col_name.text_input(
                "Expense Item",
                value=str(row.get("name", "")),
                key=f"{state_key}_name_{row_id}",
            )
            fixed_cost = col_fixed.number_input(
                "Annual Fixed Cost",
                value=float(row.get("fixed_cost", 0.0)),
                key=f"{state_key}_fixed_{row_id}",
                min_value=0.0,
                step=1000.0,
                format="%.2f",
            )
            variable_cost = col_variable.number_input(
                "Variable Cost per MWh",
                value=float(row.get("variable_cost", 0.0)),
                key=f"{state_key}_variable_{row_id}",
                min_value=0.0,
                step=0.01,
                format="%.4f",
            )

            inflation_percent = col_infl.number_input(
                "Annual Escalation (%)",
                value=float(row.get("inflation_rate", 0.0)) * 100.0,
                key=f"{state_key}_inflation_{row_id}",
                min_value=0.0,
                max_value=100.0,
                step=0.1,
                format="%.2f",
            )

            remove_clicked = col_remove.button("Remove", key=f"{state_key}_remove_{row_id}")

        if remove_clicked:
            for suffix in ("name", "fixed", "variable", "inflation"):
                st.session_state.pop(f"{state_key}_{suffix}_{row_id}", None)
            continue

        updated_rows.append(
            {
                "id": row_id,
                "name": name,
                "fixed_cost": float(fixed_cost),
                "variable_cost": float(variable_cost),
                "inflation_rate": float(inflation_percent) / 100.0,
            }
        )

    st.session_state[state_key] = updated_rows

    if st.button("Add Expense Item", key=f"{state_key}_add"):
        new_id = uuid.uuid4().hex
        st.session_state[state_key].append(
            {
                "id": new_id,
                "name": "New Expense",
                "fixed_cost": 0.0,
                "variable_cost": 0.0,
                "inflation_rate": 0.0,
            }
        )


def _render_accounts_receivable_section() -> None:
    state_key = "accounts_receivable"
    if state_key not in st.session_state:
        st.session_state[state_key] = copy.deepcopy(ACCOUNTS_RECEIVABLE_DEFAULTS)
    _ensure_schedule_row_ids(state_key)

    st.markdown("### Accounts Receivable Input Table")
    start_year, _ = _projection_year_bounds()
    year_options = _projection_year_options()
    if not year_options:
        year_options = [start_year]
    action_cols = st.columns([1.4, 1, 1, 3.6])
    increment_key = _schedule_increment_state_key(state_key)
    st.session_state.setdefault(increment_key, 0.0)
    increment_percent = action_cols[0].number_input(
        "Annual Increment (%)",
        key=increment_key,
        min_value=-100.0,
        step=0.5,
        format="%.2f",
        help="Cascade receivable and working-capital day assumptions year over year from the first row.",
    )
    if action_cols[1].button("Apply Increment", key=f"{state_key}_apply_increment"):
        st.session_state[state_key] = _apply_sequential_annual_increment(
            st.session_state[state_key],
            ("receivable_days", "prepaid_expense_days", "other_asset_days"),
            float(increment_percent),
            year_field="year",
            minimum=0.0,
        )
    if action_cols[2].button("Add Row", key=f"{state_key}_add"):
        template = copy.deepcopy(st.session_state[state_key][-1]) if st.session_state[state_key] else {
            "days_in_year": 365,
            "receivable_days": 45,
            "prepaid_expense_days": 30,
            "other_asset_days": 5,
        }
        template["id"] = uuid.uuid4().hex
        template["year"] = _next_projection_year_for_rows(st.session_state[state_key], year_options, start_year)
        st.session_state[state_key].append(template)
        st.session_state[_schedule_edit_state_key(state_key)] = template["id"]

    rows = st.session_state[state_key]
    updated_rows: List[Dict[str, object]] = []
    editing_row_id = st.session_state.get(_schedule_edit_state_key(state_key))
    for idx, row in enumerate(rows):
        row_id = str(row.get("id"))
        summary_cols = st.columns([3.8, 1.2, 1])
        summary_cols[0].markdown(
            f"**{int(row.get('year', year_options[0]))}**  |  AR {float(row.get('receivable_days', 0.0)):,.1f} days  |  "
            f"Prepaid {float(row.get('prepaid_expense_days', 0.0)):,.1f} days  |  Other {float(row.get('other_asset_days', 0.0)):,.1f} days"
        )
        if summary_cols[1].button("Edit Row", key=f"{state_key}_edit_{row_id}"):
            st.session_state[_schedule_edit_state_key(state_key)] = row_id
            editing_row_id = row_id
        remove_clicked = summary_cols[2].button("Remove", key=f"{state_key}_remove_{row_id}")
        if remove_clicked:
            if editing_row_id == row_id:
                st.session_state[_schedule_edit_state_key(state_key)] = None
            continue

        updated_row = copy.deepcopy(row)
        if editing_row_id == row_id:
            with st.container(border=True):
                col_year, col_days, col_ar_days, col_prepaid, col_other, col_done = st.columns([1.2, 1, 1, 1, 1, 0.8])
                current_year = int(row.get("year", year_options[0]))
                if current_year not in year_options:
                    current_year = year_options[0]
                updated_row["year"] = col_year.selectbox(
                    "Receivable year",
                    options=year_options,
                    index=year_options.index(current_year),
                    key=f"{state_key}_year_{row_id}",
                )
                updated_row["days_in_year"] = col_days.number_input(
                    "Days in Year",
                    value=float(row.get("days_in_year", 365)),
                    key=f"{state_key}_days_{row_id}",
                    min_value=360.0,
                    max_value=370.0,
                    step=1.0,
                )
                updated_row["receivable_days"] = col_ar_days.number_input(
                    "Accounts Receivable Days",
                    value=float(row.get("receivable_days", 45)),
                    key=f"{state_key}_ar_{row_id}",
                    min_value=0.0,
                    step=1.0,
                )
                updated_row["prepaid_expense_days"] = col_prepaid.number_input(
                    "Prepaid Expense Days",
                    value=float(row.get("prepaid_expense_days", 30)),
                    key=f"{state_key}_prepaid_{row_id}",
                    min_value=0.0,
                    step=1.0,
                )
                updated_row["other_asset_days"] = col_other.number_input(
                    "Other Asset Days",
                    value=float(row.get("other_asset_days", 5)),
                    key=f"{state_key}_other_{row_id}",
                    min_value=0.0,
                    step=1.0,
                )
                if col_done.button("Done", key=f"{state_key}_done_{row_id}"):
                    st.session_state[_schedule_edit_state_key(state_key)] = None
                    editing_row_id = None
        updated_rows.append(updated_row)
    st.session_state[state_key] = updated_rows


def _render_inventory_payables_section() -> None:
    state_key = "inventory_payables"
    if state_key not in st.session_state:
        st.session_state[state_key] = copy.deepcopy(INVENTORY_PAYABLE_DEFAULTS)
    _ensure_schedule_row_ids(state_key)

    st.markdown("### Inventory & Accounts Payable Input Table")
    start_year, _ = _projection_year_bounds()
    year_options = _projection_year_options()
    if not year_options:
        year_options = [start_year]
    action_cols = st.columns([1.4, 1, 1, 3.6])
    increment_key = _schedule_increment_state_key(state_key)
    st.session_state.setdefault(increment_key, 0.0)
    increment_percent = action_cols[0].number_input(
        "Annual Increment (%)",
        key=increment_key,
        min_value=-100.0,
        step=0.5,
        format="%.2f",
        help="Cascade inventory and payable day assumptions year over year from the first row.",
    )
    if action_cols[1].button("Apply Increment", key=f"{state_key}_apply_increment"):
        st.session_state[state_key] = _apply_sequential_annual_increment(
            st.session_state[state_key],
            ("inventory_days", "accounts_payable_days"),
            float(increment_percent),
            year_field="year",
            minimum=0.0,
        )
    if action_cols[2].button("Add Row", key=f"{state_key}_add"):
        template = copy.deepcopy(st.session_state[state_key][-1]) if st.session_state[state_key] else {
            "days_in_year": 365,
            "inventory_days": 50,
            "accounts_payable_days": 45,
        }
        template["id"] = uuid.uuid4().hex
        template["year"] = _next_projection_year_for_rows(st.session_state[state_key], year_options, start_year)
        st.session_state[state_key].append(template)
        st.session_state[_schedule_edit_state_key(state_key)] = template["id"]

    rows = st.session_state[state_key]
    updated_rows: List[Dict[str, object]] = []
    editing_row_id = st.session_state.get(_schedule_edit_state_key(state_key))
    for idx, row in enumerate(rows):
        row_id = str(row.get("id"))
        summary_cols = st.columns([3.8, 1.2, 1])
        summary_cols[0].markdown(
            f"**{int(row.get('year', year_options[0]))}**  |  Inventory {float(row.get('inventory_days', 0.0)):,.1f} days  |  "
            f"Payables {float(row.get('accounts_payable_days', 0.0)):,.1f} days"
        )
        if summary_cols[1].button("Edit Row", key=f"{state_key}_edit_{row_id}"):
            st.session_state[_schedule_edit_state_key(state_key)] = row_id
            editing_row_id = row_id
        remove_clicked = summary_cols[2].button("Remove", key=f"{state_key}_remove_{row_id}")
        if remove_clicked:
            if editing_row_id == row_id:
                st.session_state[_schedule_edit_state_key(state_key)] = None
            continue

        updated_row = copy.deepcopy(row)
        if editing_row_id == row_id:
            with st.container(border=True):
                col_year, col_days, col_inventory, col_payable, col_done = st.columns([1.2, 1, 1, 1, 0.8])
                current_year = int(row.get("year", year_options[0]))
                if current_year not in year_options:
                    current_year = year_options[0]
                updated_row["year"] = col_year.selectbox(
                    "Inventory year",
                    options=year_options,
                    index=year_options.index(current_year),
                    key=f"{state_key}_year_{row_id}",
                )
                updated_row["days_in_year"] = col_days.number_input(
                    "Days in Year",
                    value=float(row.get("days_in_year", 365)),
                    key=f"{state_key}_days_{row_id}",
                    min_value=360.0,
                    max_value=370.0,
                    step=1.0,
                )
                updated_row["inventory_days"] = col_inventory.number_input(
                    "Inventory Days",
                    value=float(row.get("inventory_days", 50)),
                    key=f"{state_key}_inventory_{row_id}",
                    min_value=0.0,
                    step=1.0,
                )
                updated_row["accounts_payable_days"] = col_payable.number_input(
                    "Accounts Payable Days",
                    value=float(row.get("accounts_payable_days", 45)),
                    key=f"{state_key}_payable_{row_id}",
                    min_value=0.0,
                    step=1.0,
                )
                if col_done.button("Done", key=f"{state_key}_done_{row_id}"):
                    st.session_state[_schedule_edit_state_key(state_key)] = None
                    editing_row_id = None
        updated_rows.append(updated_row)
    st.session_state[state_key] = updated_rows


def _render_loan_schedule_section() -> None:
    state_key = "loan_schedule"
    if state_key not in st.session_state:
        st.session_state[state_key] = copy.deepcopy(LOAN_SCHEDULE_DEFAULTS)
    _ensure_schedule_row_ids(state_key)
    if "loan_base_rate" not in st.session_state:
        st.session_state["loan_base_rate"] = 0.06

    st.markdown("### Loan Schedule")
    base_rate_col, add_col, _ = st.columns([1, 1, 2])
    st.session_state["loan_base_rate"] = base_rate_col.number_input(
        "Base Interest Rate",
        value=float(st.session_state["loan_base_rate"]),
        min_value=0.0,
        max_value=1.0,
        step=0.005,
        format="%.3f",
    )

    start_year, _ = _projection_year_bounds()
    year_options = _projection_year_options()
    if not year_options:
        year_options = [start_year]

    if add_col.button("Add Row", key=f"{state_key}_add"):
        next_year = _next_projection_year_for_rows(st.session_state[state_key], year_options, start_year)
        new_row = {
            "id": uuid.uuid4().hex,
            "name": f"Facility {len(st.session_state[state_key]) + 1}",
            "year": next_year,
            "duration_years": 5,
            "amount": 1_000_000.0,
            "interest_rate": st.session_state["loan_base_rate"],
        }
        st.session_state[state_key].append(new_row)
        st.session_state[_schedule_edit_state_key(state_key)] = new_row["id"]

    rows = st.session_state[state_key]
    updated_rows: List[Dict[str, object]] = []
    editing_row_id = st.session_state.get(_schedule_edit_state_key(state_key))

    for idx, row in enumerate(rows):
        row_id = str(row.get("id"))
        summary_cols = st.columns([4.2, 1.2, 1])
        summary_cols[0].markdown(
            f"**{str(row.get('name', f'Facility {idx + 1}'))}**  |  "
            f"Year {int(row.get('year', year_options[0]))}  |  "
            f"Duration {int(max(1, row.get('duration_years', 1)))} years  |  "
            f"Amount ${float(row.get('amount', 0.0)):,.2f}  |  "
            f"Rate {float(row.get('interest_rate', st.session_state['loan_base_rate'])):.3f}"
        )
        if summary_cols[1].button("Edit Row", key=f"{state_key}_edit_{row_id}"):
            st.session_state[_schedule_edit_state_key(state_key)] = row_id
            editing_row_id = row_id
        remove_clicked = summary_cols[2].button("Remove", key=f"{state_key}_remove_{row_id}")
        if remove_clicked:
            if editing_row_id == row_id:
                st.session_state[_schedule_edit_state_key(state_key)] = None
            continue

        updated_row = copy.deepcopy(row)
        if editing_row_id == row_id:
            with st.container(border=True):
                col_label, col_year, col_duration, col_amount, col_rate, col_done = st.columns([1.6, 1, 1, 1.2, 1, 0.6])
                updated_row["name"] = col_label.text_input(
                    "Facility Label",
                    value=str(row.get("name", f"Facility {idx + 1}")),
                    key=f"{state_key}_name_{row_id}",
                )
                current_year = int(row.get("year", year_options[0]))
                if current_year not in year_options:
                    current_year = year_options[0]
                updated_row["year"] = col_year.selectbox(
                    "Year",
                    options=year_options,
                    index=year_options.index(current_year),
                    key=f"{state_key}_year_{row_id}",
                )
                duration = int(max(1, row.get("duration_years", 1)))
                updated_row["duration_years"] = int(
                    col_duration.number_input(
                        "Duration",
                        value=float(duration),
                        key=f"{state_key}_duration_{row_id}",
                        min_value=1.0,
                        step=1.0,
                    )
                )
                updated_row["amount"] = float(
                    col_amount.number_input(
                        "Senior Debt Amount",
                        value=float(row.get("amount", 0.0)),
                        key=f"{state_key}_amount_{row_id}",
                        min_value=0.0,
                        step=1000.0,
                    )
                )
                updated_row["interest_rate"] = float(
                    col_rate.number_input(
                        "Interest Rate",
                        value=float(row.get("interest_rate", st.session_state["loan_base_rate"])),
                        key=f"{state_key}_rate_{row_id}",
                        min_value=0.0,
                        max_value=1.0,
                        step=0.005,
                        format="%.3f",
                    )
                )
                if col_done.button("Done", key=f"{state_key}_done_{row_id}"):
                    st.session_state[_schedule_edit_state_key(state_key)] = None
                    editing_row_id = None

        updated_rows.append(updated_row)

    st.session_state[state_key] = updated_rows

    schedule_rows: List[Dict[str, object]] = []
    for facility in st.session_state[state_key]:
        year = int(facility["year"])
        duration = max(1, int(facility["duration_years"]))
        amount = float(facility["amount"])
        rate = float(facility.get("interest_rate", st.session_state["loan_base_rate"]))
        facility_name = str(facility.get("name", "Facility"))

        remaining = amount
        principal_payment = amount / duration if duration else amount

        for offset in range(duration):
            period_year = year + offset
            interest_payment = remaining * rate
            remaining = max(0.0, remaining - principal_payment)
            schedule_rows.append(
                {
                    "Facility": facility_name,
                    "Facility Year": year,
                    "Period": period_year,
                    "Interest Payment": interest_payment,
                    "Principal Payment": principal_payment,
                    "Remaining Balance": remaining,
                }
            )

    if schedule_rows:
        schedule_df = pd.DataFrame(schedule_rows)
        st.markdown("#### Senior Debt Amortisation Schedule")
        st.dataframe(schedule_df, use_container_width=True)
    else:
        st.info("Add a loan facility to build the amortisation schedule.")


def _render_tax_schedule_section() -> None:
    state_key = "tax_schedule"
    if state_key not in st.session_state:
        st.session_state[state_key] = copy.deepcopy(TAX_SCHEDULE_DEFAULTS)
    if "tax_base_rate" not in st.session_state:
        st.session_state["tax_base_rate"] = 0.25
    if "tax_timing_adjustment" not in st.session_state:
        st.session_state["tax_timing_adjustment"] = 0.5

    st.markdown("### Tax Schedule")
    base_col, timing_col = st.columns(2)
    st.session_state["tax_base_rate"] = base_col.number_input(
        "Base tax rate",
        value=float(st.session_state["tax_base_rate"]),
        min_value=0.0,
        max_value=1.0,
        step=0.005,
        format="%.3f",
    )
    st.session_state["tax_timing_adjustment"] = timing_col.number_input(
        "Timing adjustment",
        value=float(st.session_state["tax_timing_adjustment"]),
        min_value=0.0,
        max_value=1.0,
        step=0.005,
        format="%.3f",
    )

    rows = st.session_state[state_key]
    updated_rows: List[Dict[str, object]] = []
    start_year, _ = _projection_year_bounds()
    year_options = _projection_year_options()
    if not year_options:
        year_options = [start_year]

    for idx, row in enumerate(rows):
        with st.container(border=True):
            col_label, col_year, col_rate, col_remove = st.columns([1.4, 1, 1, 0.6])
            label = col_label.text_input(
                "Tax label",
                value=str(row.get("name", f"Tax {idx + 1}")),
                key=f"{state_key}_label_{idx}",
            )
            current_year = int(row.get("year", year_options[0]))
            if current_year not in year_options:
                current_year = year_options[0]
            year = col_year.selectbox(
                "Year",
                options=year_options,
                index=year_options.index(current_year),
                key=f"{state_key}_year_{idx}",
            )
            tax_rate = col_rate.number_input(
                "Tax rate",
                value=float(row.get("tax_rate", st.session_state["tax_base_rate"])),
                key=f"{state_key}_rate_{idx}",
                min_value=0.0,
                max_value=1.0,
                step=0.005,
                format="%.3f",
            )
            remove_clicked = col_remove.button("Remove", key=f"{state_key}_remove_{idx}")
        if remove_clicked:
            continue
        updated_rows.append({"name": label, "year": int(year), "tax_rate": float(tax_rate)})

    st.session_state[state_key] = updated_rows

    if st.button("Add Tax Year", key=f"{state_key}_add"):
        next_year = (
            max(row["year"] for row in st.session_state[state_key]) + 1
            if st.session_state[state_key]
            else start_year
        )
        if next_year > year_options[-1]:
            next_year = year_options[-1]
        st.session_state[state_key].append(
            {
                "name": f"Tax {len(st.session_state[state_key]) + 1}",
                "year": next_year,
                "tax_rate": st.session_state["tax_base_rate"],
            }
        )

    if st.session_state[state_key]:
        tax_df = pd.DataFrame(st.session_state[state_key])
        st.dataframe(tax_df, use_container_width=True)


def _render_inflation_schedule_section() -> None:
    state_key = "inflation_schedule"
    if state_key not in st.session_state:
        st.session_state[state_key] = copy.deepcopy(INFLATION_SCHEDULE_DEFAULTS)
    _ensure_schedule_row_ids(state_key)

    st.markdown("### Inflation Schedule")
    start_year, _ = _projection_year_bounds()
    year_options = _projection_year_options()
    if not year_options:
        year_options = [start_year]
    action_cols = st.columns([1.4, 1, 1, 3.6])
    increment_key = _schedule_increment_state_key(state_key)
    st.session_state.setdefault(increment_key, 0.0)
    increment_percent = action_cols[0].number_input(
        "Annual Increment (%)",
        key=increment_key,
        min_value=-100.0,
        step=0.5,
        format="%.2f",
        help="Cascade inflation rates year over year from the first row.",
    )
    if action_cols[1].button("Apply Increment", key=f"{state_key}_apply_increment"):
        st.session_state[state_key] = _apply_sequential_annual_increment(
            st.session_state[state_key],
            ("rate",),
            float(increment_percent),
            year_field="year",
            minimum=0.0,
            maximums={"rate": 1.0},
        )
    if action_cols[2].button("Add Row", key=f"{state_key}_add"):
        template = copy.deepcopy(st.session_state[state_key][-1]) if st.session_state[state_key] else {
            "name": "Inflation 1",
            "rate": 0.0,
        }
        template["id"] = uuid.uuid4().hex
        template["year"] = _next_projection_year_for_rows(st.session_state[state_key], year_options, start_year)
        template["name"] = str(template.get("name", f"Inflation {len(st.session_state[state_key]) + 1}"))
        st.session_state[state_key].append(template)
        st.session_state[_schedule_edit_state_key(state_key)] = template["id"]

    rows = st.session_state[state_key]
    updated_rows: List[Dict[str, object]] = []
    editing_row_id = st.session_state.get(_schedule_edit_state_key(state_key))
    for idx, row in enumerate(rows):
        row_id = str(row.get("id"))
        summary_cols = st.columns([3.8, 1.2, 1])
        summary_cols[0].markdown(
            f"**{str(row.get('name', f'Inflation {idx + 1}'))}**  |  Year {int(row.get('year', year_options[0]))}  |  "
            f"Rate {float(row.get('rate', 0.0)):.3f}"
        )
        if summary_cols[1].button("Edit Row", key=f"{state_key}_edit_{row_id}"):
            st.session_state[_schedule_edit_state_key(state_key)] = row_id
            editing_row_id = row_id
        remove_clicked = summary_cols[2].button("Remove", key=f"{state_key}_remove_{row_id}")
        if remove_clicked:
            if editing_row_id == row_id:
                st.session_state[_schedule_edit_state_key(state_key)] = None
            continue

        updated_row = copy.deepcopy(row)
        if editing_row_id == row_id:
            with st.container(border=True):
                col_label, col_year, col_rate, col_done = st.columns([1.4, 1, 1, 0.6])
                updated_row["name"] = col_label.text_input(
                    "Inflation label",
                    value=str(row.get("name", f"Inflation {idx + 1}")),
                    key=f"{state_key}_label_{row_id}",
                )
                current_year = int(row.get("year", year_options[0]))
                if current_year not in year_options:
                    current_year = year_options[0]
                updated_row["year"] = col_year.selectbox(
                    "Year",
                    options=year_options,
                    index=year_options.index(current_year),
                    key=f"{state_key}_year_{row_id}",
                )
                updated_row["rate"] = col_rate.number_input(
                    "Rate",
                    value=float(row.get("rate", 0.0)),
                    key=f"{state_key}_rate_{row_id}",
                    min_value=0.0,
                    max_value=1.0,
                    step=0.005,
                    format="%.3f",
                )
                if col_done.button("Done", key=f"{state_key}_done_{row_id}"):
                    st.session_state[_schedule_edit_state_key(state_key)] = None
                    editing_row_id = None
        updated_rows.append(updated_row)

    st.session_state[state_key] = updated_rows

    if st.session_state[state_key]:
        inflation_df = pd.DataFrame(st.session_state[state_key]).drop(columns=["id"], errors="ignore")
        st.dataframe(inflation_df, use_container_width=True)


def _render_risk_schedule_section() -> None:
    state_key = "risk_schedule"
    if state_key not in st.session_state:
        st.session_state[state_key] = copy.deepcopy(RISK_SCHEDULE_DEFAULTS)
    _ensure_schedule_row_ids(state_key)

    st.markdown("### Risk Schedule")
    start_year, _ = _projection_year_bounds()
    year_options = _projection_year_options()
    if not year_options:
        year_options = [start_year]
    action_cols = st.columns([1.4, 1, 1, 3.6])
    increment_key = _schedule_increment_state_key(state_key)
    st.session_state.setdefault(increment_key, 0.0)
    increment_percent = action_cols[0].number_input(
        "Annual Increment (%)",
        key=increment_key,
        min_value=-100.0,
        step=0.5,
        format="%.2f",
        help="Cascade risk scores year over year from the first row.",
    )
    if action_cols[1].button("Apply Increment", key=f"{state_key}_apply_increment"):
        st.session_state[state_key] = _apply_sequential_annual_increment(
            st.session_state[state_key],
            ("inherent_risk", "climate_risk", "political_risk"),
            float(increment_percent),
            year_field="year",
            minimum=0.0,
            maximums={
                "inherent_risk": 1.0,
                "climate_risk": 1.0,
                "political_risk": 1.0,
            },
        )
    if action_cols[2].button("Add Row", key=f"{state_key}_add"):
        template = copy.deepcopy(st.session_state[state_key][-1]) if st.session_state[state_key] else {
            "name": "Risk 1",
            "inherent_risk": 0.0,
            "climate_risk": 0.0,
            "political_risk": 0.0,
        }
        template["id"] = uuid.uuid4().hex
        template["year"] = _next_projection_year_for_rows(st.session_state[state_key], year_options, start_year)
        template["name"] = str(template.get("name", f"Risk {len(st.session_state[state_key]) + 1}"))
        st.session_state[state_key].append(template)
        st.session_state[_schedule_edit_state_key(state_key)] = template["id"]

    rows = st.session_state[state_key]
    updated_rows: List[Dict[str, object]] = []
    editing_row_id = st.session_state.get(_schedule_edit_state_key(state_key))

    for idx, row in enumerate(rows):
        row_id = str(row.get("id"))
        summary_cols = st.columns([4.2, 1.2, 1])
        summary_cols[0].markdown(
            f"**{str(row.get('name', f'Risk {idx + 1}'))}**  |  Year {int(row.get('year', year_options[0]))}  |  "
            f"Inherent {float(row.get('inherent_risk', 0.0)):.3f}  |  "
            f"Climate {float(row.get('climate_risk', 0.0)):.3f}  |  "
            f"Political {float(row.get('political_risk', 0.0)):.3f}"
        )
        if summary_cols[1].button("Edit Row", key=f"{state_key}_edit_{row_id}"):
            st.session_state[_schedule_edit_state_key(state_key)] = row_id
            editing_row_id = row_id
        remove_clicked = summary_cols[2].button("Remove", key=f"{state_key}_remove_{row_id}")
        if remove_clicked:
            if editing_row_id == row_id:
                st.session_state[_schedule_edit_state_key(state_key)] = None
            continue

        updated_row = copy.deepcopy(row)
        if editing_row_id == row_id:
            with st.container(border=True):
                col_label, col_year, col_inherent, col_climate, col_political, col_done = st.columns(
                    [1.4, 1, 1, 1, 1, 0.6]
                )
                updated_row["name"] = col_label.text_input(
                    "Risk label",
                    value=str(row.get("name", f"Risk {idx + 1}")),
                    key=f"{state_key}_label_{row_id}",
                )

                current_year = int(row.get("year", year_options[0]))
                if current_year not in year_options:
                    current_year = year_options[0]
                updated_row["year"] = col_year.selectbox(
                    "Year",
                    options=year_options,
                    index=year_options.index(current_year),
                    key=f"{state_key}_year_{row_id}",
                )

                updated_row["inherent_risk"] = col_inherent.number_input(
                    "Inherent Risk",
                    value=float(row.get("inherent_risk", 0.0)),
                    key=f"{state_key}_inherent_{row_id}",
                    min_value=0.0,
                    max_value=1.0,
                    step=0.005,
                    format="%.3f",
                )
                updated_row["climate_risk"] = col_climate.number_input(
                    "Climate Risk",
                    value=float(row.get("climate_risk", 0.0)),
                    key=f"{state_key}_climate_{row_id}",
                    min_value=0.0,
                    max_value=1.0,
                    step=0.005,
                    format="%.3f",
                )
                updated_row["political_risk"] = col_political.number_input(
                    "Political Risk",
                    value=float(row.get("political_risk", 0.0)),
                    key=f"{state_key}_political_{row_id}",
                    min_value=0.0,
                    max_value=1.0,
                    step=0.005,
                    format="%.3f",
                )

                if col_done.button("Done", key=f"{state_key}_done_{row_id}"):
                    st.session_state[_schedule_edit_state_key(state_key)] = None
                    editing_row_id = None

        updated_rows.append(updated_row)

    st.session_state[state_key] = updated_rows

    if st.session_state[state_key]:
        risk_df = pd.DataFrame(st.session_state[state_key]).drop(columns=["id"], errors="ignore")
        st.dataframe(risk_df, use_container_width=True)


def _get_row_value(state_key: str, field_id: str, default: float | bool, expected_type: type) -> float | bool:
    rows = st.session_state.get(state_key, [])
    for row in rows:
        if row.get("id") == field_id or row.get("label") == field_id:
            value = row.get("value", default)
            break
    else:
        return default

    if expected_type is bool:
        return bool(value)
    try:
        return expected_type(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return default


def _render_assumption_controls() -> tuple[
    bytes | None,
    Dict[str, float | bool | int | str],
    List[Dict[str, object]],
    List[Dict[str, object]],
    List[Dict[str, object]],
    List[Dict[str, object]],
    List[Dict[str, object]],
    List[Dict[str, object]],
    List[Dict[str, object]],
    List[Dict[str, object]],
    List[Dict[str, object]],
    List[Dict[str, object]],
    List[Dict[str, object]],
]:
    """Render the primary assumption inputs and return override values."""

    st.subheader("Assumptions")

    start_year, end_year = _render_projection_horizon_section()
    _render_label_value_table("Core Assumptions", "core_table", CORE_ASSUMPTION_DEFAULTS)
    _render_label_value_table("Global", "global_table", GLOBAL_DEFAULTS)
    _render_label_value_table("Energy", "energy_table", ENERGY_COMMON_DEFAULTS)
    energy_input_mode = _render_energy_input_mode_selector()
    if energy_input_mode == "capacity_factor":
        _render_label_value_table(
            "Energy Driver",
            "energy_capacity_factor_table",
            ENERGY_CAPACITY_FACTOR_DEFAULTS,
        )
        monthly_generation_rows: List[Dict[str, object]] = []
    elif energy_input_mode == "resource_hours":
        _render_label_value_table(
            "Energy Driver",
            "energy_resource_table",
            ENERGY_RESOURCE_HOURS_DEFAULTS,
        )
        monthly_generation_rows = []
    else:
        monthly_generation_rows = _render_monthly_generation_table()
    _render_label_value_table("Revenue Inputs", "revenue_table", REVENUE_DEFAULTS)
    _render_initial_investment_section()
    _render_labour_structure_section()
    _render_operating_expense_section()
    _render_accounts_receivable_section()
    _render_inventory_payables_section()
    _render_loan_schedule_section()
    _render_tax_schedule_section()
    _render_inflation_schedule_section()
    _render_risk_schedule_section()

    excel_bytes = None

    project_name_override = (
        str(st.session_state.get("project_name_override", DEFAULT_PROJECT_NAME)).strip()
        or DEFAULT_PROJECT_NAME
    )

    initial_investment_rows = copy.deepcopy(st.session_state.get("initial_investment", []))
    solar_panels_amount = _solar_panels_amount(initial_investment_rows)
    panel_unit_cost_input = float(st.session_state.get("panel_unit_cost_input", PANEL_UNIT_COST_DEFAULT))
    derived_panel_count = _derived_panel_count(solar_panels_amount, panel_unit_cost_input)

    overrides: Dict[str, float | bool | str] = {
        "discount_rate": float(_get_row_value("core_table", "discount_rate", 0.10, float)),
        "exit_multiple": float(_get_row_value("core_table", "exit_multiple", 5.0, float)),
        "include_terminal": bool(_get_row_value("core_table", "include_terminal", True, bool)),
        "terminal_growth_rate": float(_get_row_value("core_table", "terminal_growth_rate", 0.02, float)),
        "project_name": project_name_override,
        "income_tax_rate": float(_get_row_value("global_table", "income_tax_rate", 0.25, float)),
        "capital_gains_tax_rate": float(
            _get_row_value("global_table", "capital_gains_tax_rate", 0.10, float)
        ),
        "investor_share": float(_get_row_value("global_table", "investor_share", 0.95, float)),
        "owner_share": float(_get_row_value("global_table", "owner_share", 0.05, float)),
        "capacity_mw": float(_get_row_value("energy_table", "capacity_mw", 10.0, float)),
        "capacity_factor": float(
            _get_row_value("energy_capacity_factor_table", "capacity_factor", 0.145, float)
        ),
        "degradation_rate": float(_get_row_value("energy_table", "degradation_rate", 0.005, float)),
        "annual_resource_hours": float(
            _get_row_value(
                "energy_resource_table",
                "annual_resource_hours",
                round(0.145 * CALENDAR_HOURS_PER_YEAR, 1),
                float,
            )
        ),
        "energy_input_mode": energy_input_mode,
        "panel_count": float(derived_panel_count),
        "panel_watt_dc": float(_get_row_value("energy_table", "panel_watt_dc", 550.0, float)),
        "panel_unit_cost": float(panel_unit_cost_input),
        "dc_ac_ratio": float(_get_row_value("energy_table", "dc_ac_ratio", 1.25, float)),
        "annual_production_growth_rate": float(
            _get_row_value("energy_table", "annual_production_growth_rate", 0.0, float)
        ),
        "monthly_min_mwh": float(_get_row_value("energy_table", "monthly_min_mwh", 10.0, float)),
        "ppa_share": float(_get_row_value("revenue_table", "ppa_share", 0.90, float)),
        "ppa_rate": float(_get_row_value("revenue_table", "ppa_rate", 160.0, float)),
        "ppa_escalation": float(_get_row_value("revenue_table", "ppa_escalation", 0.015, float)),
        "merchant_rate": float(_get_row_value("revenue_table", "merchant_rate", 56.58, float)),
        "merchant_escalation": float(_get_row_value("revenue_table", "merchant_escalation", 0.015, float)),
        "rec_rate": float(_get_row_value("revenue_table", "rec_rate", 40.0, float)),
        "rec_escalation": float(_get_row_value("revenue_table", "rec_escalation", 0.02, float)),
        "start_year": int(start_year),
        "end_year": int(end_year),
    }

    labour_rows = [
        {
            "role": str(row.get("role", "")).strip(),
            "annual_cost": float(row.get("annual_cost", 0.0)),
        }
        for row in st.session_state.get("labour_structure", [])
        if str(row.get("role", "")).strip()
    ]

    cost_rows = [
        {
            "name": str(row.get("name", "")).strip(),
            "fixed_cost": float(row.get("fixed_cost", 0.0)),
            "variable_cost": float(row.get("variable_cost", 0.0)),
            "inflation_rate": float(row.get("inflation_rate", 0.0)),
        }
        for row in st.session_state.get("operating_expenses", [])
        if str(row.get("name", "")).strip()
    ]

    receivable_rows = copy.deepcopy(st.session_state.get("accounts_receivable", []))
    inventory_rows = copy.deepcopy(st.session_state.get("inventory_payables", []))
    fixed_asset_rows = copy.deepcopy(st.session_state.get("fixed_assets_schedule", []))
    loan_rows = copy.deepcopy(st.session_state.get("loan_schedule", []))
    tax_rows = copy.deepcopy(st.session_state.get("tax_schedule", []))
    inflation_rows = copy.deepcopy(st.session_state.get("inflation_schedule", []))
    risk_rows = copy.deepcopy(st.session_state.get("risk_schedule", []))

    return (
        excel_bytes,
        overrides,
        monthly_generation_rows,
        labour_rows,
        initial_investment_rows,
        cost_rows,
        receivable_rows,
        inventory_rows,
        fixed_asset_rows,
        loan_rows,
        tax_rows,
        inflation_rows,
        risk_rows,
    )


def _render_input_landing(assumptions: Assumptions, outputs: ModelOutputs) -> None:
    """Present guidance for the Input Landing Page."""

    st.info(
        "Use the editable tables above to update assumptions. "
        "Summary snapshots now live on the Key Metrics Dashboard tab."
    )


def _render_assumption_snapshot(assumptions: Assumptions, outputs: ModelOutputs) -> None:
    """Display core, global, energy, and monthly production summaries."""

    global_cfg = assumptions.global_assumptions
    energy_cfg = assumptions.energy
    metrics = outputs.metrics

    st.header("Assumption Snapshot")

    start_timestamp = pd.Timestamp(global_cfg.start_date)
    horizon_months = max(1, int(global_cfg.forecast_months))
    horizon_end = (start_timestamp + pd.DateOffset(months=horizon_months - 1)).date()
    horizon_years = horizon_months / 12.0
    st.subheader("Core Assumptions")

    default_project_name = st.session_state.get("project_name_override", global_cfg.project_name)
    project_name_input = st.text_input(
        "Project Name",
        value=default_project_name,
        help="Update the project name used across dashboards and exports.",
    )
    project_name_value = project_name_input.strip() or default_project_name
    st.session_state["project_name_override"] = project_name_value
    global_cfg.project_name = project_name_value

    core_df = pd.DataFrame(
        {
            "Metric": [
                "Project Name",
                "Forecast Months",
                "Forecast Years",
                "Start Date",
                "End Date",
                "Include Terminal Value",
                "Exit EBITDA Multiple",
                "Discount Rate",
                "Project NPV",
                "Project IRR",
            ],
            "Value": [
                project_name_value,
                horizon_months,
                f"{horizon_years:.1f}",
                global_cfg.start_date.strftime("%Y-%m-%d"),
                horizon_end.strftime("%Y-%m-%d"),
                "Yes" if global_cfg.include_terminal_value else "No",
                f"{global_cfg.exit_multiple:.2f}x",
                _format_percentage(global_cfg.discount_rate),
                _format_currency(metrics.get("project_npv", float("nan"))),
                _format_percentage(metrics.get("project_irr", float("nan"))),
            ],
        }
    )
    st.dataframe(core_df, use_container_width=True, hide_index=True)

    st.subheader("Global")
    global_col1, global_col2, global_col3 = st.columns(3)
    with global_col1:
        st.metric("Income Tax Rate", _format_percentage(global_cfg.tax.income_tax_rate))
        st.metric("Capital Gains Tax", _format_percentage(global_cfg.tax.capital_gains_tax_rate))
    with global_col2:
        st.metric("Investor Share", _format_percentage(global_cfg.distribution.investor_share))
        st.metric("Owner Share", _format_percentage(global_cfg.distribution.owner_share))
    with global_col3:
        st.metric("Terminal Growth", _format_percentage(assumptions.terminal_growth_rate))
        st.metric(
            "Payback",
            _format_metric(
                "project_payback_months",
                metrics.get("project_payback_months", float("nan")),
            ),
        )

    st.subheader("Energy")
    energy_cols = st.columns(4)
    energy_input_mode = str(
        getattr(energy_cfg, "input_mode", getattr(assumptions, "energy_input_mode", "capacity_factor"))
    )
    energy_cols[0].metric("Capacity (MW)", f"{energy_cfg.capacity_mw:,.2f}")
    energy_cols[1].metric("Capacity Factor", _format_percentage(energy_cfg.capacity_factor))
    energy_cols[2].metric("Degradation Rate", _format_percentage(energy_cfg.degradation_rate))
    energy_cols[3].metric("Input Mode", ENERGY_INPUT_MODE_OPTIONS.get(energy_input_mode, energy_input_mode))
    st.caption(f"Annual hours are fixed at {int(CALENDAR_HOURS_PER_YEAR):,}.")
    if energy_input_mode == "resource_hours":
        annual_resource_hours = float(getattr(energy_cfg, "annual_resource_hours_input", 0.0) or 0.0)
        st.metric("Annual Resource Hours", f"{annual_resource_hours:,.1f}")

    st.markdown("#### Monthly Production Profile")
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    if assumptions.energy.energy_model_mode == "monthly_expected_mwh" and assumptions.energy.monthly_expected_mwh:
        monthly_df = pd.DataFrame(
            {
                "Month": months,
                "Expected MWh (Year 1)": assumptions.energy.monthly_expected_mwh,
            }
        )
        st.metric("Minimum Monthly Generation", f"{assumptions.energy.monthly_min_mwh:,.2f} MWh")
        st.metric(
            "Annual Production Growth",
            _format_percentage(assumptions.energy.annual_production_growth_rate),
        )
    else:
        monthly_df = pd.DataFrame(
            {
                "Month": months,
                "Share of Annual Output": assumptions.energy.seasonality,
            }
        )
        monthly_df["Share of Annual Output"] = monthly_df["Share of Annual Output"].apply(_format_percentage)
    st.dataframe(monthly_df, use_container_width=True, hide_index=True)


def _render_overview(outputs: ModelOutputs, summary_tables: Dict[str, pd.DataFrame]) -> None:
    """Display the overview page with key metrics and highlights."""

    st.subheader("Headline metrics")
    metrics = outputs.metrics
    metric_cols = st.columns(len(MetricLabels))
    for col, (metric_key, label) in zip(metric_cols, MetricLabels.items()):
        value = metrics.get(metric_key, float("nan"))
        col.metric(label, _format_metric(metric_key, value))

    st.divider()
    st.subheader("Latest drivers")
    st.dataframe(summary_tables["key_drivers"], use_container_width=True)

    st.divider()
    energy_chart, cashflow_chart = st.columns(2)
    with energy_chart:
        st.markdown("#### Energy production")
        st.line_chart(outputs.monthly_results["energy_mwh"], use_container_width=True)
    with cashflow_chart:
        st.markdown("#### Equity cash flow")
        st.area_chart(outputs.monthly_results["equity_cash_flow"], use_container_width=True)

    st.divider()
    st.subheader("Annual summary")
    annual = summary_tables["annual_summary"]
    st.dataframe(annual, use_container_width=True)


def _render_revenue_and_energy(outputs: ModelOutputs) -> None:
    """Display revenue composition and energy output analytics."""

    monthly = outputs.monthly_results
    st.subheader("Energy and revenue")

    st.markdown("#### Monthly energy production")
    st.line_chart(monthly["energy_mwh"], use_container_width=True)

    revenue_cols = monthly.filter(like="revenue_")
    if not revenue_cols.empty:
        st.markdown("#### Revenue mix")
        st.area_chart(revenue_cols, use_container_width=True)
        annual_revenue = revenue_cols.resample("YE").sum()
        annual_revenue.index = annual_revenue.index.year
        st.dataframe(annual_revenue, use_container_width=True)
    else:
        st.info("Revenue inputs were not available in the current run.")


def _render_operating_costs(outputs: ModelOutputs) -> None:
    """Display fixed and variable operating cost trends."""

    monthly = outputs.monthly_results
    opex_cols = monthly.filter(like="opex_")

    st.subheader("Operating costs")
    st.markdown("#### Total operating cost")
    st.line_chart(monthly["total_opex"], use_container_width=True)

    if not opex_cols.empty:
        st.markdown("#### Cost breakdown")
        st.area_chart(opex_cols, use_container_width=True)
        annual_opex = opex_cols.resample("YE").sum()
        annual_opex.index = annual_opex.index.year
        st.dataframe(annual_opex, use_container_width=True)
    else:
        st.info("No operating expense items were configured for this run.")


def _render_capital_and_debt(outputs: ModelOutputs) -> None:
    """Display capex profile and debt schedule insights."""

    monthly = outputs.monthly_results
    st.subheader("Capital expenditure and debt")

    if "capex" in monthly.columns:
        st.markdown("#### Monthly capex")
        st.bar_chart(monthly[["capex"]], use_container_width=True)
    else:
        st.info("Capex profile not available.")

    if not outputs.asset_summaries.empty:
        st.markdown("#### Fixed asset summary")
        display_df = outputs.asset_summaries.copy()
        if "service_month" in display_df.columns:
            display_df = display_df.rename(columns={"service_month": "Service Month"})
        st.dataframe(display_df, use_container_width=True)
    else:
        st.info("No fixed asset items were configured.")

    debt_cols = [col for col in ["debt_draw", "debt_principal", "debt_interest", "debt_balance"] if col in monthly.columns]
    if debt_cols:
        st.markdown("#### Debt schedule")
        st.line_chart(monthly[debt_cols], use_container_width=True)
        st.dataframe(monthly[debt_cols], use_container_width=True)
    else:
        st.info("Debt facilities are not part of the current assumption set.")


def _render_cash_flows(outputs: ModelOutputs) -> None:
    """Display free cash flow and equity distribution analytics."""

    monthly = outputs.monthly_results
    st.subheader("Cash flow & returns")

    cash_cols = ["fcff", "equity_cash_flow", "investor_cash_flow", "owner_cash_flow"]
    available_cash = [c for c in cash_cols if c in monthly.columns]

    if available_cash:
        st.markdown("#### Monthly cash flows")
        st.area_chart(monthly[available_cash], use_container_width=True)

        cumulative = monthly[available_cash].cumsum()
        cumulative.columns = [f"cumulative_{col}" for col in cumulative.columns]
        st.markdown("#### Cumulative cash flows")
        st.line_chart(cumulative, use_container_width=True)
        st.dataframe(cumulative, use_container_width=True)
    else:
        st.info("Cash flow outputs are not available for the current run.")


def _render_data_and_downloads(
    outputs: ModelOutputs, summary_tables: Dict[str, pd.DataFrame], assumptions: Assumptions
) -> None:
    """Expose the raw model tables."""

    st.subheader("Model tables")
    st.markdown("#### Monthly detail")
    st.dataframe(outputs.monthly_results, use_container_width=True)

    st.markdown("#### Annual summary")
    st.dataframe(outputs.annual_summary, use_container_width=True)

    st.markdown("#### Metrics")
    st.dataframe(summary_tables["metrics"], use_container_width=True)

def _render_downloads(
    outputs: ModelOutputs, summary_tables: Dict[str, pd.DataFrame], assumptions: Assumptions
) -> None:
    """Render export actions for the current model run."""
    st.write("Export polished presentation outputs for offline analysis.")
    workbook_key = "investor_workbook_bytes"
    workbook_signature = (
        float(outputs.metrics.get("project_npv", float("nan"))),
        float(outputs.metrics.get("project_irr", float("nan"))),
        float(outputs.metrics.get("equity_irr", float("nan"))),
        float(outputs.metrics.get("investor_irr", float("nan"))),
        float(outputs.metrics.get("owner_irr", float("nan"))),
        float(outputs.metrics.get("project_payback_months", float("nan"))),
    )
    signature_key = "investor_workbook_signature"

    if st.button("Prepare Investor Workbook", key="prepare_investor_workbook"):
        with st.spinner("Preparing workbook (optimized cache + reduced simulation load)..."):
            st.session_state[workbook_key] = _downloadable_excel(outputs, summary_tables, assumptions)
            st.session_state[signature_key] = workbook_signature

    if st.session_state.get(signature_key) == workbook_signature and workbook_key in st.session_state:
        st.download_button(
            "Download Investor Presentation Workbook (.xlsx)",
            data=st.session_state[workbook_key],
            file_name="solar_farm_investor_pack.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.info("Click **Prepare Investor Workbook** to generate the latest export for download.")


def _render_financial_performance(outputs: ModelOutputs) -> None:
    """Render detailed income statement schedules."""

    st.header("Statement of Financial Performance")
    income_statement = outputs.annual_summary[
        ["revenue_total", "total_opex", "ebitda", "ebit", "tax_payment", "net_income"]
    ].rename(
        columns={
            "revenue_total": "Revenue",
            "total_opex": "Operating Expenses",
            "ebitda": "EBITDA",
            "ebit": "EBIT",
            "tax_payment": "Taxes",
            "net_income": "Net Income",
        }
    )
    st.dataframe(income_statement, use_container_width=True)

    st.header("Gross Revenue Schedule")
    revenue_schedule = outputs.monthly_results.filter(like="revenue_").resample("YE").sum()
    revenue_schedule.index = revenue_schedule.index.year
    revenue_schedule = revenue_schedule.rename(
        columns=lambda c: c.replace("revenue_", "").replace("_", " ").title()
    )
    st.dataframe(revenue_schedule, use_container_width=True)

    st.header("Total Expense Schedule")
    expense_schedule = outputs.monthly_results.filter(like="opex_").resample("YE").sum()
    expense_schedule.index = expense_schedule.index.year
    expense_schedule = expense_schedule.rename(
        columns=lambda c: c.replace("opex_", "").replace("_", " ").title()
    )
    expense_schedule["Total"] = expense_schedule.sum(axis=1)
    st.dataframe(expense_schedule, use_container_width=True)


def _render_financial_position(outputs: ModelOutputs) -> None:
    """Display a simplified balance sheet view."""

    monthly = outputs.monthly_results
    opening_ppe = monthly.get("ppe_opening_balance", pd.Series(0.0, index=monthly.index)).cumsum()
    net_ppe = (opening_ppe + monthly["capex"].cumsum() - monthly["depreciation"].cumsum()).clip(lower=0)
    cash_balance = monthly["equity_cash_flow"].cumsum()
    debt_balance = monthly.get("debt_balance", pd.Series(0.0, index=monthly.index))
    accounts_receivable = monthly.get("accounts_receivable", pd.Series(0.0, index=monthly.index))
    prepaid_expenses = monthly.get("prepaid_expenses", pd.Series(0.0, index=monthly.index))
    other_current_assets = monthly.get("other_current_assets", pd.Series(0.0, index=monthly.index))
    inventory_balance = monthly.get("inventory_balance", pd.Series(0.0, index=monthly.index))
    accounts_payable = monthly.get("accounts_payable", pd.Series(0.0, index=monthly.index))

    total_assets = (
        cash_balance
        + accounts_receivable
        + prepaid_expenses
        + other_current_assets
        + inventory_balance
        + net_ppe
    )
    total_liabilities = debt_balance + accounts_payable

    balance = pd.DataFrame(
        {
            "Cash": cash_balance,
            "Accounts Receivable": accounts_receivable,
            "Prepaid Expenses": prepaid_expenses,
            "Other Current Assets": other_current_assets,
            "Inventory": inventory_balance,
            "Net PP&E": net_ppe,
            "Total Assets": total_assets,
            "Accounts Payable": accounts_payable,
            "Debt Outstanding": debt_balance,
            "Total Liabilities": total_liabilities,
        }
    )
    balance["Equity"] = balance["Total Assets"] - balance["Total Liabilities"]
    balance["Total Liabilities & Equity"] = balance["Total Liabilities"] + balance["Equity"]

    balance_sheet = balance.resample("YE").last()
    balance_sheet.index = balance_sheet.index.year

    st.header("Statement of Financial Position")
    st.dataframe(balance_sheet, use_container_width=True)


def _render_cash_flow_statement(outputs: ModelOutputs) -> None:
    """Show annual cash flow statement derived from the monthly projection."""

    monthly = outputs.monthly_results
    ebitda = monthly["ebitda"].resample("YE").sum()
    taxes = monthly["tax_payment"].resample("YE").sum()
    interest = monthly["debt_interest"].resample("YE").sum()
    working_cap_change = monthly.get("delta_working_capital", pd.Series(0.0, index=monthly.index)).resample("YE").sum()
    operating_cf = ebitda - taxes - interest - working_cap_change
    investing_cf = (-monthly["capex"]).resample("YE").sum()
    financing_cf = (monthly["debt_draw"] - monthly["debt_principal"]).resample("YE").sum()
    equity_cf = monthly["equity_cash_flow"].resample("YE").sum()

    cash_flow = pd.DataFrame(
        {
            "Operating Cash Flow": operating_cf,
            "Investing Cash Flow": investing_cf,
            "Financing Cash Flow": financing_cf,
            "Equity Cash Flow": equity_cf,
            "Change in Working Capital": -working_cap_change,
        }
    )
    cash_flow["Net Cash Flow"] = cash_flow[["Operating Cash Flow", "Investing Cash Flow", "Financing Cash Flow"]].sum(axis=1)
    cash_flow["Cumulative Net Cash"] = cash_flow["Net Cash Flow"].cumsum()
    cash_flow.index = cash_flow.index.year

    st.header("Statement of Cash Flows")
    st.dataframe(cash_flow, use_container_width=True)


def _parse_multiplier_string(value: str) -> Tuple[List[float], List[str]]:
    """Split a comma-separated multiplier string into floats and invalid tokens."""

    if not value:
        return [], []

    values: List[float] = []
    invalid: List[str] = []
    for raw in value.replace(";", ",").split(","):
        token = raw.strip()
        if not token:
            continue
        try:
            values.append(float(token))
        except ValueError:
            invalid.append(token)
    return values, invalid


def _render_sensitivity_configuration() -> List[Dict[str, object]]:
    """Render editable sensitivity configuration rows and return the stored state."""

    if SENSITIVITY_STATE_KEY not in st.session_state:
        st.session_state[SENSITIVITY_STATE_KEY] = copy.deepcopy(SENSITIVITY_DEFAULTS)

    rows: List[Dict[str, object]] = st.session_state[SENSITIVITY_STATE_KEY]
    updated_rows: List[Dict[str, object]] = []
    option_keys = list(SENSITIVITY_OPTIONS.keys())
    default_key = option_keys[0]

    st.subheader("Sensitivity Analysis Configuration")
    st.caption("Configure the variables and multiplier ranges used for the one-way analysis.")

    for idx, row in enumerate(rows):
        with st.container(border=True):
            col_variable, col_multipliers, col_remove = st.columns([2, 3, 1])
            current_key = str(row.get("variable", default_key))
            if current_key not in option_keys:
                current_key = default_key
            variable_key = col_variable.selectbox(
                "Variable",
                options=option_keys,
                index=option_keys.index(current_key),
                key=f"{SENSITIVITY_STATE_KEY}_variable_{idx}",
                format_func=lambda opt: SENSITIVITY_OPTIONS[opt][0],
            )
            multipliers_value = col_multipliers.text_input(
                "Multipliers",
                value=str(row.get("multipliers", "0.90, 1.00, 1.10")),
                key=f"{SENSITIVITY_STATE_KEY}_multipliers_{idx}",
                placeholder="0.90, 1.00, 1.10",
            )
            remove_clicked = col_remove.button(
                "Remove",
                key=f"{SENSITIVITY_STATE_KEY}_remove_{idx}",
            )
        if remove_clicked:
            continue
        updated_rows.append(
            {
                "id": row.get("id") or uuid.uuid4().hex,
                "variable": variable_key,
                "multipliers": multipliers_value,
            }
        )

    st.session_state[SENSITIVITY_STATE_KEY] = updated_rows

    if st.button("Add Variable", key=f"{SENSITIVITY_STATE_KEY}_add"):
        st.session_state[SENSITIVITY_STATE_KEY].append(
            {
                "id": uuid.uuid4().hex,
                "variable": default_key,
                "multipliers": "0.90, 1.00, 1.10",
            }
        )

    return st.session_state[SENSITIVITY_STATE_KEY]


def _simulate_outputs(base: Assumptions, modifier: Callable[[Assumptions], None]) -> ModelOutputs:
    return _simulate_outputs_helper(base, modifier)


def _simulate_metrics(base: Assumptions, modifier: Callable[[Assumptions], None]) -> Dict[str, float]:
    return _simulate_metrics_helper(base, modifier)


def _goal_seek_value(outputs: ModelOutputs, source: str, metric: str, year: int | None) -> float:
    """Extract a numeric value for goal seek comparisons based on the selected source."""

    if source == "metrics":
        return float(outputs.metrics.get(metric, float("nan")))

    annual = outputs.annual_summary
    if metric not in annual.columns or annual.empty:
        return float("nan")

    if year is not None and year in annual.index:
        return float(annual.at[year, metric])

    if year is None and not annual.empty:
        return float(annual.iloc[0][metric])

    return float("nan")


def _render_sensitivity_analysis(base_assumptions: Assumptions, outputs: ModelOutputs) -> None:
    """Present configurable one-way sensitivity tables for key drivers."""

    st.header("Sensitivity Analyses")
    config_rows = _render_sensitivity_configuration()

    records = [
        {
            "Variable": "Base Case",
            "Multiplier": 1.0,
            "Scenario": "Base Case",
            "Project NPV": outputs.metrics.get("project_npv", float("nan")),
            "Project IRR": outputs.metrics.get("project_irr", float("nan")),
            "Equity IRR": outputs.metrics.get("equity_irr", float("nan")),
            "Investor IRR": outputs.metrics.get("investor_irr", float("nan")),
            "Owner IRR": outputs.metrics.get("owner_irr", float("nan")),
            "Payback (months)": outputs.metrics.get("project_payback_months", float("nan")),
        }
    ]

    produced_results = False

    for row in config_rows:
        key = str(row.get("variable", ""))
        if key not in SENSITIVITY_OPTIONS:
            continue

        label, apply_fn = SENSITIVITY_OPTIONS[key]
        multipliers_text = str(row.get("multipliers", "")).strip()
        multipliers, invalid_tokens = _parse_multiplier_string(multipliers_text)

        if invalid_tokens:
            st.warning(f"Ignoring invalid multipliers for {label}: {', '.join(invalid_tokens)}")

        if not multipliers:
            continue

        for multiplier in multipliers:
            if math.isclose(multiplier, 1.0, rel_tol=1e-9):
                metrics = outputs.metrics
            else:
                metrics = _simulate_metrics(
                    base_assumptions,
                    lambda assumptions, fn=apply_fn, m=multiplier: fn(assumptions, m),
                )

            records.append(
                {
                    "Variable": label,
                    "Multiplier": multiplier,
                    "Scenario": f"{label} x{multiplier:.2f}",
                    "Project NPV": metrics.get("project_npv", float("nan")),
                    "Project IRR": metrics.get("project_irr", float("nan")),
                    "Equity IRR": metrics.get("equity_irr", float("nan")),
                    "Investor IRR": metrics.get("investor_irr", float("nan")),
                    "Owner IRR": metrics.get("owner_irr", float("nan")),
                    "Payback (months)": metrics.get("project_payback_months", float("nan")),
                }
            )
            produced_results = True

    if not produced_results:
        st.info("Add sensitivity variables and multipliers to generate simulation results.")
        return

    st.markdown("#### Simulation Results")
    sensitivity_df = pd.DataFrame(records)
    sensitivity_df = sensitivity_df.sort_values(["Variable", "Multiplier"])
    st.dataframe(sensitivity_df, use_container_width=True)


def _evaluate_goal_seek_rows(
    base_assumptions: Assumptions,
    base_outputs: ModelOutputs,
    rows: List[Dict[str, object]],
) -> List[Dict[str, object]]:
    """Evaluate goal seek configurations and return a list of result records."""

    results: List[Dict[str, object]] = []
    if not rows:
        return results

    option_keys = list(SENSITIVITY_OPTIONS.keys())

    for row in rows:
        variable_key = str(row.get("variable", ""))
        if variable_key not in SENSITIVITY_OPTIONS:
            continue

        source = str(row.get("source", "metrics"))
        if source not in GOAL_SEEK_SOURCE_LABELS:
            source = "metrics"

        metric_options = GOAL_SEEK_METRIC_OPTIONS.get(source, GOAL_SEEK_METRIC_OPTIONS["metrics"])
        metric_key = str(row.get("metric", ""))
        if metric_key not in metric_options:
            continue

        target_value = _coerce_optional_float(row.get("target"))
        if target_value is None:
            continue

        year_value = _coerce_optional_int(row.get("year"))
        apply_label, apply_fn = SENSITIVITY_OPTIONS[variable_key]

        outputs_cache: Dict[float, ModelOutputs] = {1.0: base_outputs}
        best_multiplier: float | None = None
        best_value: float | None = None
        best_diff = float("inf")

        for multiplier in GOAL_SEEK_MULTIPLIERS:
            if multiplier not in outputs_cache:
                outputs_cache[multiplier] = _simulate_outputs(
                    base_assumptions,
                    lambda assumptions, fn=apply_fn, m=multiplier: fn(assumptions, m),
                )

            candidate_outputs = outputs_cache[multiplier]
            value = _goal_seek_value(candidate_outputs, source, metric_key, year_value)
            if pd.isna(value):
                continue

            diff = abs(value - target_value)
            if diff < best_diff:
                best_multiplier = multiplier
                best_value = value
                best_diff = diff

        if best_multiplier is None or best_value is None:
            continue

        tolerance = max(1.0, abs(target_value) * 0.01)
        status = "Within tolerance" if best_diff <= tolerance else "Closest found"
        metric_label = metric_options.get(metric_key, metric_key.replace("_", " ").title())

        results.append(
            {
                "Scenario Variable": apply_label,
                "Metric": metric_label,
                "Target": target_value,
                "Year": year_value if year_value is not None else "N/A",
                "Recommended Multiplier": best_multiplier,
                "Projected Value": best_value,
                "Difference": best_value - target_value,
                "Status": status,
            }
        )

    return results


def _render_goal_seek_configuration(
    base_assumptions: Assumptions,
    base_outputs: ModelOutputs,
) -> List[Dict[str, object]]:
    """Render the goal seek configuration editor and resulting table."""

    if GOAL_SEEK_STATE_KEY not in st.session_state:
        st.session_state[GOAL_SEEK_STATE_KEY] = copy.deepcopy(GOAL_SEEK_DEFAULTS)

    rows: List[Dict[str, object]] = st.session_state[GOAL_SEEK_STATE_KEY]
    updated_rows: List[Dict[str, object]] = []
    option_keys = list(SENSITIVITY_OPTIONS.keys())
    default_variable = option_keys[0]
    source_keys = list(GOAL_SEEK_SOURCE_LABELS.keys())

    st.subheader("Goal Seek Configuration")
    st.caption("Target key outcomes and the model will search for the multiplier that best meets the goal.")

    for idx, row in enumerate(rows):
        row_id = row.get("id") or uuid.uuid4().hex
        current_source = row.get("source") if row.get("source") in source_keys else source_keys[0]
        metric_options = GOAL_SEEK_METRIC_OPTIONS.get(current_source, GOAL_SEEK_METRIC_OPTIONS["metrics"])
        metric_keys = list(metric_options.keys()) or list(MetricLabels.keys())
        current_metric = row.get("metric") if row.get("metric") in metric_keys else metric_keys[0]
        current_variable = row.get("variable") if row.get("variable") in option_keys else default_variable

        with st.container(border=True):
            col_source, col_metric = st.columns([2, 2])
            source_value = col_source.selectbox(
                "Metric Source",
                options=source_keys,
                index=source_keys.index(current_source),
                key=f"{GOAL_SEEK_STATE_KEY}_source_{row_id}",
                format_func=lambda opt: GOAL_SEEK_SOURCE_LABELS[opt],
            )
            metric_mapping = GOAL_SEEK_METRIC_OPTIONS.get(source_value, GOAL_SEEK_METRIC_OPTIONS["metrics"])
            metric_keys = list(metric_mapping.keys()) or list(MetricLabels.keys())
            if current_metric not in metric_keys:
                current_metric = metric_keys[0]
            metric_value = col_metric.selectbox(
                "Metric",
                options=metric_keys,
                index=metric_keys.index(current_metric),
                key=f"{GOAL_SEEK_STATE_KEY}_metric_{row_id}",
                format_func=lambda opt, mapping=metric_mapping: mapping.get(opt, opt.replace("_", " ").title()),
            )

            col_variable, col_target, col_year, col_remove = st.columns([2.5, 2, 1.5, 1])
            variable_value = col_variable.selectbox(
                "Variable",
                options=option_keys,
                index=option_keys.index(current_variable),
                key=f"{GOAL_SEEK_STATE_KEY}_variable_{row_id}",
                format_func=lambda opt: SENSITIVITY_OPTIONS[opt][0],
            )
            target_value = col_target.text_input(
                "Target Value",
                value=str(row.get("target", "")),
                key=f"{GOAL_SEEK_STATE_KEY}_target_{row_id}",
            )
            year_value = col_year.text_input(
                "Year",
                value=str(row.get("year", "")),
                key=f"{GOAL_SEEK_STATE_KEY}_year_{row_id}",
            )
            remove_clicked = col_remove.button("Remove", key=f"{GOAL_SEEK_STATE_KEY}_remove_{row_id}")

        if remove_clicked:
            continue

        updated_rows.append(
            {
                "id": row_id,
                "source": source_value,
                "metric": metric_value,
                "variable": variable_value,
                "target": target_value,
                "year": year_value,
            }
        )

    st.session_state[GOAL_SEEK_STATE_KEY] = updated_rows

    if st.button("Add Goal Seek", key=f"{GOAL_SEEK_STATE_KEY}_add"):
        metric_default_keys = list(GOAL_SEEK_METRIC_OPTIONS["metrics"].keys())
        st.session_state[GOAL_SEEK_STATE_KEY].append(
            {
                "id": uuid.uuid4().hex,
                "source": "metrics",
                "metric": metric_default_keys[0] if metric_default_keys else "project_npv",
                "variable": default_variable,
                "target": "",
                "year": "",
            }
        )

    rows = st.session_state[GOAL_SEEK_STATE_KEY]
    results = _evaluate_goal_seek_rows(base_assumptions, base_outputs, rows)

    if results:
        st.markdown("#### Goal Seek Results")
        result_df = pd.DataFrame(results)
        st.dataframe(result_df, use_container_width=True)
    else:
        st.info("Define at least one target with a variable to run goal seek simulations.")

    return rows


def _render_scenario_configuration() -> List[Dict[str, object]]:
    """Render editable scenario rows for multi-factor analysis."""

    if SCENARIO_STATE_KEY not in st.session_state:
        st.session_state[SCENARIO_STATE_KEY] = copy.deepcopy(SCENARIO_DEFAULTS)

    rows: List[Dict[str, object]] = st.session_state[SCENARIO_STATE_KEY]
    updated_rows: List[Dict[str, object]] = []

    st.subheader("Scenario / IFs Configuration")
    st.caption("Adjust inflation, financing, and operating multipliers to build composite scenarios.")

    for idx, row in enumerate(rows):
        row_id = row.get("id") or uuid.uuid4().hex
        with st.container(border=True):
            col_name, col_infl, col_int, col_rev, col_opex, col_capex, col_cap, col_remove = st.columns([3, 2, 2, 2, 2, 2, 2, 1])

            name_value = col_name.text_input(
                "Scenario Name",
                value=str(row.get("name", f"Scenario {idx + 1}")),
                key=f"{SCENARIO_STATE_KEY}_name_{row_id}",
            )
            inflation_value = col_infl.number_input(
                "Inflation Rate",
                min_value=0.0,
                max_value=1.0,
                value=float(row.get("inflation_rate", 0.02)),
                step=0.001,
                format="%.3f",
                key=f"{SCENARIO_STATE_KEY}_inflation_{row_id}",
            )
            interest_value = col_int.number_input(
                "Interest Rate",
                min_value=0.0,
                max_value=1.0,
                value=float(row.get("interest_rate", 0.05)),
                step=0.001,
                format="%.3f",
                key=f"{SCENARIO_STATE_KEY}_interest_{row_id}",
            )
            revenue_multiplier = col_rev.number_input(
                "Revenue Multiplier",
                min_value=0.0,
                value=float(row.get("revenue_multiplier", 1.0)),
                step=0.01,
                format="%.2f",
                key=f"{SCENARIO_STATE_KEY}_revenue_{row_id}",
            )
            opex_multiplier = col_opex.number_input(
                "Opex Multiplier",
                min_value=0.0,
                value=float(row.get("opex_multiplier", 1.0)),
                step=0.01,
                format="%.2f",
                key=f"{SCENARIO_STATE_KEY}_opex_{row_id}",
            )
            capex_multiplier = col_capex.number_input(
                "Capex Multiplier",
                min_value=0.0,
                value=float(row.get("capex_multiplier", 1.0)),
                step=0.01,
                format="%.2f",
                key=f"{SCENARIO_STATE_KEY}_capex_{row_id}",
            )
            capacity_multiplier = col_cap.number_input(
                "Capacity Multiplier",
                min_value=0.0,
                value=float(row.get("capacity_multiplier", 1.0)),
                step=0.01,
                format="%.2f",
                key=f"{SCENARIO_STATE_KEY}_capacity_{row_id}",
            )
            remove_clicked = col_remove.button("Remove", key=f"{SCENARIO_STATE_KEY}_remove_{row_id}")

        if remove_clicked:
            continue

        updated_rows.append(
            {
                "id": row_id,
                "name": name_value,
                "inflation_rate": inflation_value,
                "interest_rate": interest_value,
                "revenue_multiplier": revenue_multiplier,
                "opex_multiplier": opex_multiplier,
                "capex_multiplier": capex_multiplier,
                "capacity_multiplier": capacity_multiplier,
            }
        )

    st.session_state[SCENARIO_STATE_KEY] = updated_rows

    if st.button("Add Scenario", key=f"{SCENARIO_STATE_KEY}_add"):
        st.session_state[SCENARIO_STATE_KEY].append(
            {
                "id": uuid.uuid4().hex,
                "name": "New Scenario",
                "inflation_rate": 0.02,
                "interest_rate": 0.05,
                "revenue_multiplier": 1.0,
                "opex_multiplier": 1.0,
                "capex_multiplier": 1.0,
                "capacity_multiplier": 1.0,
            }
        )

    return st.session_state[SCENARIO_STATE_KEY]


def _render_scenario_tool_configuration() -> Dict[str, Dict[str, object]]:
    """Render configuration blocks for advanced scenario analysis tools."""

    if SCENARIO_TOOL_STATE_KEY not in st.session_state:
        st.session_state[SCENARIO_TOOL_STATE_KEY] = copy.deepcopy(SCENARIO_TOOL_DEFAULTS)

    st.subheader("Scenario Tool Configuration")
    st.caption("Document additional analytical approaches that accompany each scenario run.")

    state: Dict[str, Dict[str, object]] = st.session_state[SCENARIO_TOOL_STATE_KEY]
    updated_state: Dict[str, Dict[str, object]] = {}

    for key, meta in SCENARIO_TOOL_SECTIONS:
        section_state = state.get(key, {"variables": [], "iterate": ""})
        section_variables = section_state.get("variables", [])
        iterate_value = section_state.get("iterate", "")

        with st.container(border=True):
            st.markdown(f"**{meta['label']}**")
            updated_variables: List[Dict[str, object]] = []

            for var in section_variables:
                row_id = var.get("id") or uuid.uuid4().hex
                col_var, col_remove = st.columns([4, 1])
                variable_value = col_var.text_input(
                    "Variable",
                    value=str(var.get("variable", "")),
                    key=f"{SCENARIO_TOOL_STATE_KEY}_{key}_variable_{row_id}",
                )
                remove_clicked = col_remove.button(
                    "Remove",
                    key=f"{SCENARIO_TOOL_STATE_KEY}_{key}_remove_{row_id}",
                )
                if remove_clicked:
                    continue
                updated_variables.append({"id": row_id, "variable": variable_value})

            if st.button("Add Variable", key=f"{SCENARIO_TOOL_STATE_KEY}_{key}_add"):
                updated_variables.append({"id": uuid.uuid4().hex, "variable": ""})

            if meta.get("show_iterate", False):
                iterate_value = st.text_input(
                    "Iterate",
                    value=str(iterate_value),
                    key=f"{SCENARIO_TOOL_STATE_KEY}_{key}_iterate",
                )

        updated_state[key] = {"variables": updated_variables, "iterate": iterate_value}

    st.session_state[SCENARIO_TOOL_STATE_KEY] = updated_state
    return updated_state


def _build_scenario_dataframe(
    base_assumptions: Assumptions,
    base_outputs: ModelOutputs,
    scenario_rows: List[Dict[str, object]],
) -> pd.DataFrame:
    """Construct a dataframe comparing configured scenarios."""

    records: List[Dict[str, object]] = [
        {
            "Scenario": "Base Case",
            "Inflation Rate": float("nan"),
            "Interest Rate": float("nan"),
            "Revenue Multiplier": 1.0,
            "Opex Multiplier": 1.0,
            "Capex Multiplier": 1.0,
            "Capacity Multiplier": 1.0,
            "Project NPV": base_outputs.metrics.get("project_npv", float("nan")),
            "Project IRR": base_outputs.metrics.get("project_irr", float("nan")),
            "Equity IRR": base_outputs.metrics.get("equity_irr", float("nan")),
            "Investor IRR": base_outputs.metrics.get("investor_irr", float("nan")),
            "Owner IRR": base_outputs.metrics.get("owner_irr", float("nan")),
            "Payback (months)": base_outputs.metrics.get("project_payback_months", float("nan")),
        }
    ]

    for idx, row in enumerate(scenario_rows):
        name = str(row.get("name", f"Scenario {idx + 1}")).strip() or f"Scenario {idx + 1}"
        inflation_rate = _coerce_optional_float(row.get("inflation_rate"))
        interest_rate = _coerce_optional_float(row.get("interest_rate"))
        revenue_multiplier = _coerce_float(row.get("revenue_multiplier"), 1.0)
        opex_multiplier = _coerce_float(row.get("opex_multiplier"), 1.0)
        capex_multiplier = _coerce_float(row.get("capex_multiplier"), 1.0)
        capacity_multiplier = _coerce_float(row.get("capacity_multiplier"), 1.0)

        requires_simulation = not (
            math.isclose(revenue_multiplier, 1.0, rel_tol=1e-9)
            and math.isclose(opex_multiplier, 1.0, rel_tol=1e-9)
            and math.isclose(capex_multiplier, 1.0, rel_tol=1e-9)
            and math.isclose(capacity_multiplier, 1.0, rel_tol=1e-9)
            and inflation_rate is None
            and interest_rate is None
        )

        if requires_simulation:

            def modifier(assumptions: Assumptions) -> None:
                if not math.isclose(revenue_multiplier, 1.0, rel_tol=1e-9):
                    assumptions.revenue.ppa.rate_curve.initial *= revenue_multiplier
                    assumptions.revenue.merchant.rate_curve.initial *= revenue_multiplier
                    assumptions.revenue.rec.initial *= revenue_multiplier

                if not math.isclose(opex_multiplier, 1.0, rel_tol=1e-9):
                    for item in assumptions.fixed_opex:
                        item.annual_cost = max(0.0, item.annual_cost * opex_multiplier)
                        if hasattr(item, "cost_per_mwh"):
                            item.cost_per_mwh = max(0.0, item.cost_per_mwh * opex_multiplier)
                    for item in assumptions.variable_opex:
                        item.cost_per_mwh = max(0.0, item.cost_per_mwh * opex_multiplier)

                if not math.isclose(capex_multiplier, 1.0, rel_tol=1e-9):
                    for item in assumptions.capex_items:
                        item.amount = max(0.0, item.amount * capex_multiplier)

                if not math.isclose(capacity_multiplier, 1.0, rel_tol=1e-9):
                    assumptions.energy.capacity_factor = max(
                        0.01,
                        min(1.0, assumptions.energy.capacity_factor * capacity_multiplier),
                    )

                if inflation_rate is not None:
                    for item in assumptions.fixed_opex:
                        item.inflation_rate = inflation_rate
                    for item in assumptions.variable_opex:
                        item.escalation_rate = inflation_rate

                if interest_rate is not None:
                    for facility in assumptions.debt_facilities:
                        facility.interest_rate = interest_rate

            scenario_outputs = _simulate_outputs(base_assumptions, modifier)
        else:
            scenario_outputs = base_outputs

        metrics = scenario_outputs.metrics
        records.append(
            {
                "Scenario": name,
                "Inflation Rate": inflation_rate if inflation_rate is not None else float("nan"),
                "Interest Rate": interest_rate if interest_rate is not None else float("nan"),
                "Revenue Multiplier": revenue_multiplier,
                "Opex Multiplier": opex_multiplier,
                "Capex Multiplier": capex_multiplier,
                "Capacity Multiplier": capacity_multiplier,
                "Project NPV": metrics.get("project_npv", float("nan")),
                "Project IRR": metrics.get("project_irr", float("nan")),
                "Equity IRR": metrics.get("equity_irr", float("nan")),
                "Investor IRR": metrics.get("investor_irr", float("nan")),
                "Owner IRR": metrics.get("owner_irr", float("nan")),
                "Payback (months)": metrics.get("project_payback_months", float("nan")),
            }
        )

    return pd.DataFrame(records)


def _render_scenario_analysis(base_assumptions: Assumptions, outputs: ModelOutputs) -> None:
    """Display configurable goal seek, scenario, and analysis tool controls."""

    st.header("Scenario / IFs Analysis")
    _render_goal_seek_configuration(base_assumptions, outputs)
    scenario_rows = _render_scenario_configuration()
    _render_scenario_tool_configuration()

    scenario_df = _build_scenario_dataframe(base_assumptions, outputs, scenario_rows)

    if scenario_df.empty:
        st.info("Add scenarios to generate comparison results.")
        return

    st.markdown("#### Scenario Comparison")
    st.dataframe(scenario_df, use_container_width=True)

    selected = st.selectbox(
        "Select scenario for highlight",
        scenario_df["Scenario"].tolist(),
        key="scenario_highlight",
    )
    selected_row = scenario_df[scenario_df["Scenario"] == selected].iloc[0]
    highlight_cols = st.columns(3)
    highlight_cols[0].metric(
        "Project NPV",
        _format_currency(selected_row.get("Project NPV", float("nan"))),
    )
    highlight_cols[1].metric(
        "Project IRR",
        _format_percentage(selected_row.get("Project IRR", float("nan"))),
    )
    highlight_cols[2].metric(
        "Payback",
        _format_metric(
            "project_payback_months",
            selected_row.get("Payback (months)", float("nan")),
        ),
    )


def _render_monte_carlo_configuration() -> Dict[str, object]:
    """Capture Monte Carlo configuration, drivers, and metric outputs."""

    if MONTE_CARLO_STATE_KEY not in st.session_state:
        st.session_state[MONTE_CARLO_STATE_KEY] = copy.deepcopy(MONTE_CARLO_CONFIG_DEFAULT)

    config: Dict[str, object] = st.session_state[MONTE_CARLO_STATE_KEY]
    config.setdefault("drivers", copy.deepcopy(MONTE_CARLO_DRIVER_DEFAULTS))
    config.setdefault("metrics", MONTE_CARLO_METRIC_DEFAULTS.copy())
    config.setdefault("iterations", MONTE_CARLO_CONFIG_DEFAULT["iterations"])
    config.setdefault("seed", MONTE_CARLO_CONFIG_DEFAULT["seed"])

    st.subheader("Monte Carlo Simulation Configuration")
    st.caption("Control the randomised drivers, iteration count, and captured metrics.")

    col_iterations, col_seed = st.columns([1, 1])
    iterations = col_iterations.number_input(
        "Iterations",
        min_value=100,
        max_value=5000,
        step=100,
        value=int(config.get("iterations", MONTE_CARLO_CONFIG_DEFAULT["iterations"])),
        help="Number of random scenarios to evaluate.",
    )
    seed = col_seed.number_input(
        "Random Seed",
        min_value=0,
        max_value=10_000,
        step=1,
        value=int(config.get("seed", MONTE_CARLO_CONFIG_DEFAULT["seed"])),
        help="Seed used to reproduce simulation results.",
    )
    config["iterations"] = int(iterations)
    config["seed"] = int(seed)

    metric_keys = list(MetricLabels.keys())
    selected_metrics = st.multiselect(
        "Metric outputs",
        options=metric_keys,
        default=[m for m in config.get("metrics", []) if m in metric_keys] or MONTE_CARLO_METRIC_DEFAULTS,
        format_func=lambda key: MetricLabels.get(key, key.replace("_", " ").title()),
        help="Choose which metrics to summarise from the simulation results.",
    )
    config["metrics"] = selected_metrics or MONTE_CARLO_METRIC_DEFAULTS.copy()

    driver_rows: List[Dict[str, object]] = list(config.get("drivers", []))
    updated_drivers: List[Dict[str, object]] = []
    option_keys = list(SENSITIVITY_OPTIONS.keys())
    if not option_keys:
        st.warning("No variables available for simulation.")
        config["drivers"] = []
    else:
        for idx, row in enumerate(driver_rows):
            with st.container(border=True):
                col_variable, col_dist, col_p1, col_p2, col_p3, col_p4, col_remove = st.columns(
                    [2.2, 1.6, 1.2, 1.2, 1.2, 1.2, 0.8]
                )
                current_var = str(row.get("variable", option_keys[0]))
                if current_var not in option_keys:
                    current_var = option_keys[0]
                variable_key = col_variable.selectbox(
                    "Variable",
                    options=option_keys,
                    index=option_keys.index(current_var),
                    key=f"{MONTE_CARLO_STATE_KEY}_variable_{idx}",
                    format_func=lambda opt: SENSITIVITY_OPTIONS[opt][0],
                )

                distribution_options = list(MONTE_CARLO_DISTRIBUTION_OPTIONS.keys())
                current_dist = str(row.get("distribution", distribution_options[0]))
                if current_dist not in distribution_options:
                    current_dist = distribution_options[0]
                distribution = col_dist.selectbox(
                    "Distribution",
                    options=distribution_options,
                    index=distribution_options.index(current_dist),
                    key=f"{MONTE_CARLO_STATE_KEY}_distribution_{idx}",
                    format_func=lambda opt: MONTE_CARLO_DISTRIBUTION_OPTIONS[opt],
                )

                params: Dict[str, float] = {}
                if distribution == "normal":
                    params["mean"] = col_p1.number_input(
                        "Mean multiplier",
                        value=float(row.get("mean", 1.0)),
                        step=0.01,
                        key=f"{MONTE_CARLO_STATE_KEY}_mean_{idx}",
                    )
                    params["std_dev"] = col_p2.number_input(
                        "Std dev",
                        value=float(row.get("std_dev", 0.05)),
                        min_value=0.0,
                        step=0.005,
                        key=f"{MONTE_CARLO_STATE_KEY}_std_{idx}",
                    )
                    params["min"] = col_p3.number_input(
                        "Min clip",
                        value=float(row.get("min", 0.5)),
                        step=0.01,
                        key=f"{MONTE_CARLO_STATE_KEY}_min_{idx}",
                    )
                    params["max"] = col_p4.number_input(
                        "Max clip",
                        value=float(row.get("max", 1.5)),
                        step=0.01,
                        key=f"{MONTE_CARLO_STATE_KEY}_max_{idx}",
                    )
                elif distribution == "uniform":
                    params["min"] = col_p1.number_input(
                        "Min multiplier",
                        value=float(row.get("min", 0.9)),
                        step=0.01,
                        key=f"{MONTE_CARLO_STATE_KEY}_umin_{idx}",
                    )
                    params["max"] = col_p2.number_input(
                        "Max multiplier",
                        value=float(row.get("max", 1.1)),
                        step=0.01,
                        key=f"{MONTE_CARLO_STATE_KEY}_umax_{idx}",
                    )
                else:  # triangular
                    params["min"] = col_p1.number_input(
                        "Min multiplier",
                        value=float(row.get("min", 0.9)),
                        step=0.01,
                        key=f"{MONTE_CARLO_STATE_KEY}_tmin_{idx}",
                    )
                    params["mode"] = col_p2.number_input(
                        "Mode",
                        value=float(row.get("mode", 1.0)),
                        step=0.01,
                        key=f"{MONTE_CARLO_STATE_KEY}_tmode_{idx}",
                    )
                    params["max"] = col_p3.number_input(
                        "Max multiplier",
                        value=float(row.get("max", 1.1)),
                        step=0.01,
                        key=f"{MONTE_CARLO_STATE_KEY}_tmax_{idx}",
                    )
                    params["std_dev"] = float(row.get("std_dev", 0.0))

                remove_clicked = col_remove.button(
                    "Remove",
                    key=f"{MONTE_CARLO_STATE_KEY}_remove_{idx}",
                )

            if remove_clicked:
                continue

            updated_row = {
                "id": row.get("id") or uuid.uuid4().hex,
                "variable": variable_key,
                "distribution": distribution,
            }
            updated_row.update(params)
            updated_drivers.append(updated_row)

        st.session_state[MONTE_CARLO_STATE_KEY]["drivers"] = updated_drivers

        if st.button("Add Simulation Variable", key=f"{MONTE_CARLO_STATE_KEY}_add"):
            st.session_state[MONTE_CARLO_STATE_KEY]["drivers"].append(
                {
                    "id": uuid.uuid4().hex,
                    "variable": option_keys[0],
                    "distribution": "normal",
                    "mean": 1.0,
                    "std_dev": 0.05,
                    "min": 0.8,
                    "max": 1.2,
                }
            )

    return copy.deepcopy(st.session_state[MONTE_CARLO_STATE_KEY])


def _sample_monte_carlo_multiplier(rng: np.random.Generator, row: Dict[str, object]) -> float:
    """Generate a multiplier for the selected driver based on its configuration."""

    distribution = str(row.get("distribution", "normal"))
    if distribution == "normal":
        mean = _coerce_float(row.get("mean"), 1.0)
        std_dev = max(0.0, _coerce_float(row.get("std_dev"), 0.05))
        value = float(rng.normal(mean, std_dev))
        min_clip = row.get("min")
        max_clip = row.get("max")
        if min_clip is not None:
            value = max(_coerce_float(min_clip, value), value)
        if max_clip is not None:
            value = min(_coerce_float(max_clip, value), value)
        return value

    if distribution == "uniform":
        low = _coerce_float(row.get("min"), 0.9)
        high = _coerce_float(row.get("max"), 1.1)
        if high <= low:
            high = low + 1e-6
        return float(rng.uniform(low, high))

    # Triangular distribution
    low = _coerce_float(row.get("min"), 0.9)
    high = _coerce_float(row.get("max"), max(low + 1e-6, 1.1))
    mode = _coerce_float(row.get("mode"), (low + high) / 2)
    mode = min(max(mode, low), high)
    return float(rng.triangular(low, mode, high))


def _render_break_even_configuration() -> List[Dict[str, object]]:
    """Render editable break-even input rows and return the stored configuration."""

    if BREAK_EVEN_STATE_KEY not in st.session_state:
        st.session_state[BREAK_EVEN_STATE_KEY] = copy.deepcopy(BREAK_EVEN_DEFAULTS)

    rows: List[Dict[str, object]] = st.session_state[BREAK_EVEN_STATE_KEY]
    updated_rows: List[Dict[str, object]] = []

    st.subheader("Break-even Analysis Inputs")
    st.caption("Capture fixed costs, contribution margins, and target profit by product or revenue stream.")

    for idx, row in enumerate(rows):
        with st.container(border=True):
            cols = st.columns([2.0, 1.2, 1.2, 1.2, 1.2, 1.2, 0.8])
            product = cols[0].text_input(
                "Product",
                value=str(row.get("product", "")),
                key=f"{BREAK_EVEN_STATE_KEY}_product_{idx}",
            )
            fixed_cost = cols[1].number_input(
                "Fixed Cost",
                value=float(row.get("fixed_cost", 0.0)),
                min_value=0.0,
                step=1_000.0,
                key=f"{BREAK_EVEN_STATE_KEY}_fixed_{idx}",
            )
            variable_cost = cols[2].number_input(
                "Variable Cost",
                value=float(row.get("variable_cost", 0.0)),
                min_value=0.0,
                step=0.01,
                key=f"{BREAK_EVEN_STATE_KEY}_variable_{idx}",
            )
            selling_price = cols[3].number_input(
                "Selling Price",
                value=float(row.get("selling_price", 0.0)),
                min_value=0.0,
                step=0.01,
                key=f"{BREAK_EVEN_STATE_KEY}_price_{idx}",
            )
            target_profit = cols[4].number_input(
                "Target Profit",
                value=float(row.get("target_profit", 0.0)),
                min_value=0.0,
                step=1_000.0,
                key=f"{BREAK_EVEN_STATE_KEY}_profit_{idx}",
            )
            expected_volume = cols[5].number_input(
                "Expected Volume",
                value=float(row.get("expected_volume", 0.0)),
                min_value=0.0,
                step=1_000.0,
                key=f"{BREAK_EVEN_STATE_KEY}_volume_{idx}",
            )
            remove_clicked = cols[6].button("Remove", key=f"{BREAK_EVEN_STATE_KEY}_remove_{idx}")

        if remove_clicked:
            continue

        updated_rows.append(
            {
                "id": row.get("id") or uuid.uuid4().hex,
                "product": product,
                "fixed_cost": fixed_cost,
                "variable_cost": variable_cost,
                "selling_price": selling_price,
                "target_profit": target_profit,
                "expected_volume": expected_volume,
            }
        )

    st.session_state[BREAK_EVEN_STATE_KEY] = updated_rows

    with st.container(border=True):
        st.markdown("#### Add Break-even Input")
        add_cols = st.columns([2.0, 1.2, 1.2, 1.2, 1.2, 1.2, 0.8])
        new_product = add_cols[0].text_input(
            "Product Name",
            value=st.session_state.get(f"{BREAK_EVEN_STATE_KEY}_new_product", ""),
            key=f"{BREAK_EVEN_STATE_KEY}_new_product",
        )
        new_fixed = add_cols[1].number_input(
            "Fixed Cost",
            value=st.session_state.get(f"{BREAK_EVEN_STATE_KEY}_new_fixed", 0.0),
            min_value=0.0,
            step=1_000.0,
            key=f"{BREAK_EVEN_STATE_KEY}_new_fixed",
        )
        new_variable = add_cols[2].number_input(
            "Variable Cost",
            value=st.session_state.get(f"{BREAK_EVEN_STATE_KEY}_new_variable", 0.0),
            min_value=0.0,
            step=0.01,
            key=f"{BREAK_EVEN_STATE_KEY}_new_variable",
        )
        new_price = add_cols[3].number_input(
            "Selling Price",
            value=st.session_state.get(f"{BREAK_EVEN_STATE_KEY}_new_price", 0.0),
            min_value=0.0,
            step=0.01,
            key=f"{BREAK_EVEN_STATE_KEY}_new_price",
        )
        new_target = add_cols[4].number_input(
            "Target Profit",
            value=st.session_state.get(f"{BREAK_EVEN_STATE_KEY}_new_target", 0.0),
            min_value=0.0,
            step=1_000.0,
            key=f"{BREAK_EVEN_STATE_KEY}_new_target",
        )
        new_volume = add_cols[5].number_input(
            "Expected Volume",
            value=st.session_state.get(f"{BREAK_EVEN_STATE_KEY}_new_volume", 0.0),
            min_value=0.0,
            step=1_000.0,
            key=f"{BREAK_EVEN_STATE_KEY}_new_volume",
        )

        add_clicked = add_cols[6].button("Add", key=f"{BREAK_EVEN_STATE_KEY}_add")

    if add_clicked and new_product:
        st.session_state[BREAK_EVEN_STATE_KEY].append(
            {
                "id": uuid.uuid4().hex,
                "product": new_product,
                "fixed_cost": float(new_fixed),
                "variable_cost": float(new_variable),
                "selling_price": float(new_price),
                "target_profit": float(new_target),
                "expected_volume": float(new_volume),
            }
        )
        for suffix, reset_value in (
            ("new_product", ""),
            ("new_fixed", 0.0),
            ("new_variable", 0.0),
            ("new_price", 0.0),
            ("new_target", 0.0),
            ("new_volume", 0.0),
        ):
            st.session_state[f"{BREAK_EVEN_STATE_KEY}_{suffix}"] = reset_value
        st.experimental_rerun()

    return copy.deepcopy(st.session_state[BREAK_EVEN_STATE_KEY])


def _render_monte_carlo(base_assumptions: Assumptions) -> None:
    """Run a Monte Carlo analysis across core drivers."""

    st.header("Monte Carlo Simulation")
    config = _render_monte_carlo_configuration()

    iterations = int(config.get("iterations", MONTE_CARLO_CONFIG_DEFAULT["iterations"]))
    seed = int(config.get("seed", MONTE_CARLO_CONFIG_DEFAULT["seed"]))
    drivers: List[Dict[str, object]] = list(config.get("drivers", []))
    metric_keys: List[str] = [m for m in config.get("metrics", []) if m in MetricLabels]

    if not drivers:
        st.info("Add at least one simulation variable to run the Monte Carlo analysis.")
        return

    if not metric_keys:
        metric_keys = MONTE_CARLO_METRIC_DEFAULTS

    rng = np.random.default_rng(seed)
    results: List[Dict[str, float]] = []
    progress = st.progress(0.0)

    for i in range(iterations):

        def modifier(a: Assumptions, driver_rows: List[Dict[str, object]] = drivers) -> None:
            for driver in driver_rows:
                variable = driver.get("variable")
                if not variable or variable not in SENSITIVITY_OPTIONS:
                    continue
                multiplier = _sample_monte_carlo_multiplier(rng, driver)
                apply_fn = SENSITIVITY_OPTIONS[variable][1]
                apply_fn(a, multiplier)

        metrics = _simulate_metrics(base_assumptions, modifier)
        results.append({MetricLabels[key]: metrics.get(key, float("nan")) for key in metric_keys})
        progress.progress((i + 1) / max(1, iterations))

    progress.empty()

    if not results:
        st.warning("No simulation results generated.")
        return

    mc_df = pd.DataFrame(results)

    st.markdown("#### Distribution Summary")
    summary = mc_df.agg(["mean", "std", "min", "max"])
    quantiles = mc_df.quantile([0.1, 0.5, 0.9]).rename(index={0.1: "P10", 0.5: "Median", 0.9: "P90"})
    st.dataframe(pd.concat([summary, quantiles]), use_container_width=True)

    chart_metric_options = list(mc_df.columns)
    selected_metric = st.selectbox(
        "Distribution metric",
        options=chart_metric_options,
        index=0,
        key="monte_carlo_metric_selection",
    )

    st.markdown(f"#### {selected_metric} Distribution")
    hist, bins = np.histogram(mc_df[selected_metric].dropna(), bins=20)
    if hist.size:
        midpoints = (bins[:-1] + bins[1:]) / 2
        hist_df = pd.DataFrame({"Count": hist}, index=pd.Index(midpoints, name=selected_metric))
        st.bar_chart(hist_df)
    else:
        st.info("Not enough data to plot histogram for the selected metric.")

    st.download_button(
        "Download Simulation Results",
        data=mc_df.to_csv(index=False).encode("utf-8"),
        file_name="monte_carlo_results.csv",
        mime="text/csv",
    )


def _render_break_even(outputs: ModelOutputs) -> None:
    """Show break-even and payback diagnostics."""

    st.header("Break-Even & Payback")
    break_even_rows = _render_break_even_configuration()
    monthly = outputs.monthly_results
    cumulative_equity = monthly["equity_cash_flow"].cumsum()
    breakeven_mask = cumulative_equity >= 0
    breakeven_date = cumulative_equity.index[breakeven_mask.argmax()] if breakeven_mask.any() else None

    metrics = outputs.metrics
    metric_cols = st.columns(3)
    metric_cols[0].metric("Project NPV", _format_currency(metrics.get("project_npv", float("nan"))))
    metric_cols[1].metric("Project IRR", _format_percentage(metrics.get("project_irr", float("nan"))))
    metric_cols[2].metric(
        "Payback",
        _format_metric("project_payback_months", metrics.get("project_payback_months", float("nan"))),
    )

    if break_even_rows:
        records: List[Dict[str, object]] = []
        issues: List[str] = []
        for row in break_even_rows:
            product = str(row.get("product", "")).strip() or "Unlabelled"
            fixed_cost = max(0.0, _coerce_float(row.get("fixed_cost")))
            variable_cost = max(0.0, _coerce_float(row.get("variable_cost")))
            price = max(0.0, _coerce_float(row.get("selling_price")))
            target_profit = max(0.0, _coerce_float(row.get("target_profit")))
            expected_volume = max(0.0, _coerce_float(row.get("expected_volume")))

            contribution = price - variable_cost
            contribution_ratio = contribution / price if price > 0 else float("nan")

            if contribution <= 0:
                break_even_units = float("nan")
                break_even_revenue = float("nan")
                margin_of_safety = float("nan")
                margin_of_safety_ratio = float("nan")
                issues.append(f"{product}: Contribution margin must be positive to compute break-even.")
            else:
                break_even_units = (fixed_cost + target_profit) / contribution
                break_even_revenue = break_even_units * price
                margin_of_safety = expected_volume - break_even_units if expected_volume > 0 else float("nan")
                if expected_volume > 0 and not math.isnan(margin_of_safety):
                    margin_of_safety_ratio = margin_of_safety / expected_volume
                else:
                    margin_of_safety_ratio = float("nan")

            records.append(
                {
                    "Product": product,
                    "Fixed Cost": fixed_cost,
                    "Variable Cost": variable_cost,
                    "Selling Price": price,
                    "Contribution Margin": contribution,
                    "Contribution Margin %": contribution_ratio,
                    "Target Profit": target_profit,
                    "Expected Volume": expected_volume,
                    "Break-even Units": break_even_units,
                    "Break-even Revenue": break_even_revenue,
                    "Margin of Safety": margin_of_safety,
                    "Margin of Safety %": margin_of_safety_ratio,
                }
            )

        if records:
            st.markdown("#### Break-even Results")
            results_df = pd.DataFrame(records)
            formatters = {
                "Fixed Cost": "{:,.0f}".format,
                "Variable Cost": "{:,.4f}".format,
                "Selling Price": "{:,.4f}".format,
                "Contribution Margin": "{:,.4f}".format,
                "Contribution Margin %": "{:.2%}".format,
                "Target Profit": "{:,.0f}".format,
                "Expected Volume": "{:,.0f}".format,
                "Break-even Units": "{:,.2f}".format,
                "Break-even Revenue": "{:,.0f}".format,
                "Margin of Safety": "{:,.2f}".format,
                "Margin of Safety %": "{:.2%}".format,
            }
            st.dataframe(results_df.style.format(formatters), use_container_width=True)

            chart_df = results_df.set_index("Product")[[
                col for col in ["Break-even Units", "Expected Volume"] if col in results_df.columns
            ]]
            if not chart_df.empty:
                st.bar_chart(chart_df, use_container_width=True)

            st.download_button(
                "Download Break-even Table",
                data=results_df.to_csv(index=False).encode("utf-8"),
                file_name="break_even_analysis.csv",
                mime="text/csv",
            )

        for message in issues:
            st.error(message)

    if breakeven_date is not None:
        st.success(f"Break-even reached in {breakeven_date.strftime('%B %Y')}")
    else:
        st.warning("Break-even is not reached within the projection horizon.")

    st.markdown("#### Cumulative Equity Cash Flow")
    cumulative_df = pd.DataFrame(
        {
            "Equity Cash Flow": monthly["equity_cash_flow"],
            "Cumulative Equity": cumulative_equity,
            "Cumulative FCFF": monthly["fcff"].cumsum(),
        }
    )
    st.line_chart(cumulative_df[["Cumulative Equity", "Cumulative FCFF"]], use_container_width=True)
    st.dataframe(cumulative_df, use_container_width=True)

_inject_app_theme()
_render_model_hero()
_configure_llm_secrets()

DEFAULT_PROJECT_NAME = "Solar Numquants Ltd"

PAGE_OPTIONS = [
    "Input Landing Page",
    "Key Metrics Dashboard",
    "Financials",
    "Key Analytics",
    "AI Benchmark Assistant",
]

tabs = st.tabs(PAGE_OPTIONS)

with tabs[0]:
    (
        excel_bytes,
        override_dict,
        monthly_generation_rows,
        labour_rows,
        initial_investment_rows,
        cost_rows,
        receivable_rows,
        inventory_rows,
        fixed_asset_rows,
        loan_rows,
        tax_rows,
        inflation_rows,
        risk_rows,
    ) = _render_assumption_controls()

monthly_generation_tuple = _tupleize(monthly_generation_rows, ("month", "expected_mwh"))
labour_tuple = _tupleize(labour_rows, ("role", "annual_cost"))
initial_investment_tuple = _tupleize(
    initial_investment_rows,
    (
        "id",
        "name",
        "amount",
        "depreciation_years",
        "method",
        "year",
        "month",
        "spend_profile",
        "opening_balance",
        "depreciation_rate",
        "service_month",
    ),
)
cost_tuple = _tupleize(cost_rows, ("name", "fixed_cost", "variable_cost", "inflation_rate"))
receivable_tuple = _tupleize(
    receivable_rows,
    ("year", "days_in_year", "receivable_days", "prepaid_expense_days", "other_asset_days"),
)
inventory_tuple = _tupleize(
    inventory_rows,
    ("year", "days_in_year", "inventory_days", "accounts_payable_days"),
)
fixed_asset_tuple = _tupleize(
    fixed_asset_rows,
    (
        "asset_type",
        "method",
        "year",
        "acquisition",
        "asset_life",
        "net_book_value",
        "depreciation_rate",
        "total_asset_cost",
        "total_depreciation",
        "cumulative_depreciation",
        "ending_book_value",
    ),
)
loan_tuple = _tupleize(loan_rows, ("name", "year", "duration_years", "amount", "interest_rate"))
tax_tuple = _tupleize(tax_rows, ("name", "year", "tax_rate"))
inflation_tuple = _tupleize(inflation_rows, ("name", "year", "rate"))
risk_tuple = _tupleize(
    risk_rows,
    ("name", "year", "inherent_risk", "climate_risk", "political_risk"),
)

override_tuple = tuple(sorted(override_dict.items()))

outputs, summary_tables, assumptions = _run_model(
    excel_bytes,
    override_tuple,
    monthly_generation_tuple,
    labour_tuple,
    initial_investment_tuple,
    cost_tuple,
    receivable_tuple,
    inventory_tuple,
    fixed_asset_tuple,
    loan_tuple,
    tax_tuple,
    inflation_tuple,
    risk_tuple,
)

projection_caption = _format_projection_caption(assumptions)

with tabs[0]:
    st.divider()
    st.caption(projection_caption)
    _render_input_landing(assumptions, outputs)

for page_name, tab in zip(PAGE_OPTIONS[1:], tabs[1:]):
    with tab:
        st.caption(projection_caption)
        if page_name == "Key Metrics Dashboard":
            st.header("Downloads")
            _render_downloads(outputs, summary_tables, assumptions)
            st.divider()
            _render_assumption_snapshot(assumptions, outputs)
            st.divider()
            st.header("Overview")
            _render_overview(outputs, summary_tables)
            st.header("Revenue & Energy")
            _render_revenue_and_energy(outputs)
            st.header("Operating Costs")
            _render_operating_costs(outputs)
            st.header("Capital & Debt")
            _render_capital_and_debt(outputs)
            st.header("Cash Flow & Returns")
            _render_cash_flows(outputs)
            st.header("Data")
            _render_data_and_downloads(outputs, summary_tables, assumptions)
        elif page_name == "Financials":
            _render_financial_performance(outputs)
            st.divider()
            _render_financial_position(outputs)
            st.divider()
            _render_cash_flow_statement(outputs)
        elif page_name == "Key Analytics":
            _render_sensitivity_analysis(assumptions, outputs)
            st.divider()
            _render_scenario_analysis(assumptions, outputs)
            st.divider()
            _render_monte_carlo(assumptions)
            st.divider()
            _render_break_even(outputs)
        elif page_name == "AI Benchmark Assistant":
            _render_ai_benchmark_assistant(outputs, assumptions)
        else:
            st.info("Select a page tab to view analytics.")
